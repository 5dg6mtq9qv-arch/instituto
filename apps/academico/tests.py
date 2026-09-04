import json
import shutil
import tempfile
from datetime import timedelta, time
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.academico.models import (
    Aula,
    AulaCurso,
    Clase,
    ClaseAsistencia,
    ClaseEstudianteMovimiento,
    ClaseHoraDocente,
    Competencia,
    Curso,
    CursoPeriodo,
    Dia,
    Estrategia,
    GrupoEstudiante,
    Horario,
    HorarioAulaCurso,
    HorarioDia,
    Materia,
    MateriaCurso,
    MateriaSubtema,
    MateriaTema,
    Periodo,
    PlanificacionDocente,
    PlanificacionTema,
    ProfesorMateriaCurso,
    Recurso,
    Subtema,
    Tema,
)
from apps.academico.views import DocenteClaseAsistenciaView
from apps.core.current_user import set_current_request
from apps.core.models import Empresa, Partner, TipoIdentificacion
from apps.matricula.models import FichaInscripcion


class DocenteHorariosPanelTests(TestCase):
    def setUp(self):
        set_current_request(None)
        self.media_root = tempfile.mkdtemp()
        self.media_override = override_settings(MEDIA_ROOT=self.media_root)
        self.media_override.enable()
        self.user = get_user_model().objects.create_user(username="docente", password="ClaveActual987!")
        self.empresa = Empresa.objects.create(ruc="0999999999001", razon_social="Instituto Prueba")
        self.tipo_identificacion = TipoIdentificacion.objects.create(nombre="Cedula", codigo="CED")
        self.docente = Partner.objects.create(
            tipo_identificacion=self.tipo_identificacion,
            identificacion="DOC-001",
            nombre="Docente Prueba",
            usuario=self.user,
            es_docente=True,
            activo=True,
        )
        self.curso = Curso.objects.create(nombre="Grupo A", activo=True)
        self.aula = Aula.objects.create(nombre="Aula 1")
        aula_curso = AulaCurso.objects.create(aula=self.aula, curso=self.curso)
        dia, _ = Dia.objects.get_or_create(dia="Lunes")
        self.horario = Horario.objects.create(hora_inicio=time(8, 0), hora_fin=time(9, 0))
        horario_dia = HorarioDia.objects.create(dia=dia, horario=self.horario)
        self.horario_aula_curso = HorarioAulaCurso.objects.create(aula_curso=aula_curso, horario_dia=horario_dia)
        self.materia = Materia.objects.create(nombre="Matematicas", nombre_corto="MAT", color="#0f766e")
        self.materia_curso = MateriaCurso.objects.create(materia=self.materia, grupo=self.curso)
        self.profesor_materia_curso = ProfesorMateriaCurso.objects.create(
            partner=self.docente,
            materia_curso=self.materia_curso,
        )
        self.planificacion = PlanificacionDocente.objects.create(
            materia_curso=self.materia_curso,
            nombre="Plan base",
        )
        self.tema = Tema.objects.create(planificacion=self.planificacion, nombre="Numeros", orden=1)
        self.subtema = Subtema.objects.create(tema=self.tema, nombre="Suma", orden=1)
        self.planificacion_tema = PlanificacionTema.objects.create(
            profesor_materia_curso=self.profesor_materia_curso,
            tema=self.tema,
            nombre="Grupo A - Matematicas - Numeros",
        )
        self.competencia = Competencia.objects.create(nombre="Resolver problemas")
        self.estrategia = Estrategia.objects.create(nombre="Aprendizaje guiado")
        self.recurso = Recurso.objects.create(nombre="Pizarra")
        today = timezone.localdate()
        next_monday = today + timedelta(days=(7 - today.weekday()) % 7 or 7)
        self.pendiente = Clase.objects.create(
            horario_aula_curso=self.horario_aula_curso,
            materia_curso=self.materia_curso,
            fecha=next_monday,
            estado_planificacion="pendiente",
        )
        self.revision = Clase.objects.create(
            horario_aula_curso=self.horario_aula_curso,
            materia_curso=self.materia_curso,
            fecha=next_monday + timedelta(days=7),
            estado_planificacion="revision",
            descripcion="Clase enviada a revision.",
        )
        self.rechazada = Clase.objects.create(
            horario_aula_curso=self.horario_aula_curso,
            materia_curso=self.materia_curso,
            fecha=next_monday + timedelta(days=14),
            estado_planificacion="rechazada",
            descripcion="Clase con observaciones.",
            notas_revision="Completar recursos.",
        )
        self.atrasada = Clase.objects.create(
            horario_aula_curso=self.horario_aula_curso,
            materia_curso=self.materia_curso,
            fecha=today - timedelta(days=1),
            estado_planificacion="pendiente",
        )

    def tearDown(self):
        set_current_request(None)
        self.media_override.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)

    def create_coordinator(self):
        user = get_user_model().objects.create_user(username="coordinador", password="ClaveActual987!")
        user.groups.add(Group.objects.get_or_create(name="Coordinacion")[0])
        return user

    def create_director(self):
        user = get_user_model().objects.create_user(username="director", password="ClaveActual987!")
        user.groups.add(Group.objects.get_or_create(name="Director")[0])
        return user

    def create_periodo_for_course(self):
        today = timezone.localdate()
        periodo = Periodo.objects.create(
            nombre="Periodo de prueba",
            fecha_inicio=today,
            fecha_fin=today + timedelta(days=35),
        )
        CursoPeriodo.objects.create(curso=self.curso, periodo=periodo)
        return periodo

    def date_for_weekday(self, periodo, weekday):
        current = periodo.fecha_inicio
        while current <= periodo.fecha_fin:
            if current.weekday() == weekday:
                return current
            current += timedelta(days=1)
        return periodo.fecha_inicio

    def make_superuser(self):
        self.user.is_staff = True
        self.user.is_superuser = True
        self.user.save(update_fields=["is_staff", "is_superuser"])

    def create_docente(self, username="docente-reemplazo", identificacion="DOC-002", nombre="Docente Reemplazo"):
        user = get_user_model().objects.create_user(username=username, password="ClaveActual987!")
        docente = Partner.objects.create(
            tipo_identificacion=self.tipo_identificacion,
            identificacion=identificacion,
            nombre=nombre,
            usuario=user,
            es_docente=True,
            activo=True,
        )
        return docente, user

    def find_cell_containing(self, sheet, text):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and text in str(cell.value):
                    return cell
        self.fail(f"No se encontro una celda con el texto {text!r}")

    def find_cell_containing_all(self, sheet, *texts):
        for row in sheet.iter_rows():
            for cell in row:
                value = str(cell.value or "")
                if all(text in value for text in texts):
                    return cell
        self.fail(f"No se encontro una celda con los textos {texts!r}")

    def create_class_for_date(
        self,
        fecha,
        hora_inicio,
        hora_fin,
        aula_nombre="Aula extra",
        materia_curso=None,
        curso=None,
    ):
        curso = curso or self.curso
        aula = Aula.objects.create(nombre=aula_nombre)
        aula_curso = AulaCurso.objects.create(aula=aula, curso=curso)
        weekday_names = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]
        dia, _ = Dia.objects.get_or_create(dia=weekday_names[fecha.weekday()])
        horario, _ = Horario.objects.get_or_create(hora_inicio=hora_inicio, hora_fin=hora_fin)
        horario_dia, _ = HorarioDia.objects.get_or_create(dia=dia, horario=horario)
        horario_aula_curso = HorarioAulaCurso.objects.create(aula_curso=aula_curso, horario_dia=horario_dia)
        return Clase.objects.create(
            horario_aula_curso=horario_aula_curso,
            materia_curso=materia_curso or self.materia_curso,
            fecha=fecha,
        )

    def create_student_ficha(self, nombre="Estudiante Prueba", identificacion="EST-001", numero="F-001"):
        estudiante = Partner.objects.create(
            tipo_identificacion=self.tipo_identificacion,
            identificacion=identificacion,
            nombre=nombre,
            es_estudiante=True,
            activo=True,
        )
        representante = Partner.objects.create(
            tipo_identificacion=self.tipo_identificacion,
            identificacion=f"REP-{identificacion}",
            nombre=f"Representante {nombre}",
            es_representante=True,
            activo=True,
        )
        ficha = FichaInscripcion.objects.create(
            empresa=self.empresa,
            numero=numero,
            fecha=timezone.localdate(),
            cliente=representante,
            estudiante=estudiante,
            representante=representante,
            estado="activa",
            activo=True,
        )
        return estudiante, ficha

    def test_group_student_assignment_view_creates_academic_group_assignment(self):
        self.make_superuser()
        estudiante, ficha = self.create_student_ficha()
        self.client.force_login(self.user)

        page_response = self.client.get(reverse("academico:grupo_estudiantes"), HTTP_HOST="localhost")

        self.assertEqual(page_response.status_code, 200)
        self.assertContains(page_response, "student-transfer-board")
        self.assertContains(page_response, "data-transfer-action=\"assign\"")

        response = self.client.post(
            reverse("academico:grupo_estudiantes"),
            {
                "assignment_action": "sync_students",
                "grupo": self.curso.pk,
                "fecha_asignacion": timezone.localdate().isoformat(),
                "fichas": [ficha.pk],
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        asignacion = GrupoEstudiante.objects.get(ficha_inscripcion=ficha)
        self.assertEqual(asignacion.estudiante, estudiante)
        self.assertEqual(asignacion.grupo, self.curso)
        self.assertEqual(asignacion.estado, "activo")

    def test_group_student_assignment_view_can_return_student_to_unassigned(self):
        self.make_superuser()
        estudiante, ficha = self.create_student_ficha()
        asignacion = GrupoEstudiante.objects.create(
            ficha_inscripcion=ficha,
            estudiante=estudiante,
            grupo=self.curso,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:grupo_estudiantes"),
            {
                "assignment_action": "sync_students",
                "grupo": self.curso.pk,
                "fecha_asignacion": timezone.localdate().isoformat(),
                "asignaciones_remover": [asignacion.pk],
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(GrupoEstudiante.objects.filter(pk=asignacion.pk).exists())

    def test_group_assignment_is_default_roster_for_all_group_classes(self):
        today = timezone.localdate()
        first_class = self.create_class_for_date(today, time(10, 0), time(11, 0), "Aula asistencia 1")
        second_class = self.create_class_for_date(today, time(11, 0), time(12, 0), "Aula asistencia 2")
        estudiante, ficha = self.create_student_ficha()
        GrupoEstudiante.objects.create(
            ficha_inscripcion=ficha,
            estudiante=estudiante,
            grupo=self.curso,
        )
        self.client.force_login(self.user)

        first_response = self.client.get(
            reverse("academico:docente_clase_asistencia", args=[first_class.pk]),
            HTTP_HOST="localhost",
        )
        second_response = self.client.get(
            reverse("academico:docente_clase_asistencia", args=[second_class.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(second_response.status_code, 200)
        self.assertEqual([row["estudiante"] for row in first_response.context["rows"]], [estudiante])
        self.assertEqual([row["estudiante"] for row in second_response.context["rows"]], [estudiante])
        self.assertFalse(ClaseAsistencia.objects.filter(estudiante=estudiante).exists())

    def test_docente_attendance_only_opens_on_class_date(self):
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:docente_clase_asistencia", args=[self.pendiente.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("academico:docente_horarios"))

    def test_class_student_movement_requires_same_materia_in_another_group(self):
        estudiante, ficha = self.create_student_ficha()
        asignacion = GrupoEstudiante.objects.create(
            ficha_inscripcion=ficha,
            estudiante=estudiante,
            grupo=self.curso,
        )
        grupo_destino = Curso.objects.create(nombre="Grupo B", activo=True)
        materia_curso_destino = MateriaCurso.objects.create(materia=self.materia, grupo=grupo_destino)
        target_class = self.create_class_for_date(
            self.pendiente.fecha,
            time(10, 0),
            time(11, 0),
            "Aula destino",
            materia_curso=materia_curso_destino,
            curso=grupo_destino,
        )
        valid_movement = ClaseEstudianteMovimiento(
            asignacion=asignacion,
            clase_origen=self.pendiente,
            clase_destino=target_class,
        )

        valid_movement.full_clean()

        same_group_movement = ClaseEstudianteMovimiento(
            asignacion=asignacion,
            clase_origen=self.pendiente,
            clase_destino=self.revision,
        )
        with self.assertRaises(ValidationError) as same_group_error:
            same_group_movement.full_clean()

        self.assertIn(
            "Selecciona una clase destino de otro grupo.",
            same_group_error.exception.message_dict["clase_destino"],
        )

        otra_materia = Materia.objects.create(nombre="Lenguaje", nombre_corto="LEN", color="#2563eb")
        otra_materia_curso = MateriaCurso.objects.create(materia=otra_materia, grupo=grupo_destino)
        wrong_class = self.create_class_for_date(
            self.pendiente.fecha + timedelta(days=1),
            time(11, 0),
            time(12, 0),
            "Aula materia distinta",
            materia_curso=otra_materia_curso,
            curso=grupo_destino,
        )
        invalid_movement = ClaseEstudianteMovimiento(
            asignacion=asignacion,
            clase_origen=self.pendiente,
            clase_destino=wrong_class,
        )

        with self.assertRaises(ValidationError) as error:
            invalid_movement.full_clean()

        self.assertIn(
            "Solo puedes mover entre grupos que tengan la misma materia.",
            error.exception.message_dict["clase_destino"],
        )

    def test_group_student_movement_view_changes_subject_to_equivalent_group(self):
        self.make_superuser()
        today = timezone.localdate()
        estudiante, ficha = self.create_student_ficha()
        asignacion = GrupoEstudiante.objects.create(
            ficha_inscripcion=ficha,
            estudiante=estudiante,
            grupo=self.curso,
        )
        grupo_destino = Curso.objects.create(nombre="Grupo B", activo=True)
        materia_curso_destino = MateriaCurso.objects.create(materia=self.materia, grupo=grupo_destino)
        origin_class = self.create_class_for_date(today, time(10, 0), time(11, 0), "Aula origen movimiento")
        destination_class = self.create_class_for_date(
            today,
            time(12, 0),
            time(13, 0),
            "Aula destino movimiento",
            materia_curso=materia_curso_destino,
            curso=grupo_destino,
        )
        self.client.force_login(self.user)

        page_response = self.client.get(
            reverse("academico:grupo_estudiantes"),
            {"grupo": self.curso.pk},
            HTTP_HOST="localhost",
        )

        self.assertEqual(page_response.status_code, 200)
        self.assertContains(page_response, "id_materia_origen")
        self.assertContains(page_response, "id_materia_destino")
        self.assertNotContains(page_response, "id_clase_origen")

        response = self.client.post(
            reverse("academico:grupo_estudiantes"),
            {
                "assignment_action": "move_student",
                "grupo": self.curso.pk,
                "asignacion": asignacion.pk,
                "materia_origen": self.materia_curso.pk,
                "materia_destino": materia_curso_destino.pk,
                "fecha_inicio": today.isoformat(),
                "motivo": "Cambio de horario.",
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        movimiento = ClaseEstudianteMovimiento.objects.get(asignacion=asignacion)
        self.assertEqual(movimiento.clase_origen, origin_class)
        self.assertEqual(movimiento.clase_destino, destination_class)
        self.assertEqual(movimiento.fecha_inicio, today)
        self.assertEqual(movimiento.motivo, "Cambio de horario.")
        self.assertEqual(asignacion.grupo, self.curso)

    def test_docente_attendance_uses_group_roster_and_movement_exceptions(self):
        today = timezone.localdate()
        origin_class = self.create_class_for_date(today, time(10, 0), time(11, 0), "Aula origen")
        grupo_destino = Curso.objects.create(nombre="Grupo B", activo=True)
        materia_curso_destino = MateriaCurso.objects.create(materia=self.materia, grupo=grupo_destino)
        ProfesorMateriaCurso.objects.create(partner=self.docente, materia_curso=materia_curso_destino)
        destination_class = self.create_class_for_date(
            today,
            time(11, 0),
            time(12, 0),
            "Aula destino",
            materia_curso=materia_curso_destino,
            curso=grupo_destino,
        )
        future_origin_class = self.create_class_for_date(
            today + timedelta(days=7),
            time(10, 0),
            time(11, 0),
            "Aula origen futura",
        )
        future_destination_class = self.create_class_for_date(
            today + timedelta(days=7),
            time(11, 0),
            time(12, 0),
            "Aula destino futura",
            materia_curso=materia_curso_destino,
            curso=grupo_destino,
        )
        estudiante_uno, ficha_uno = self.create_student_ficha(
            nombre="Ana Estudiante",
            identificacion="EST-101",
            numero="F-101",
        )
        estudiante_dos, ficha_dos = self.create_student_ficha(
            nombre="Luis Estudiante",
            identificacion="EST-102",
            numero="F-102",
        )
        asignacion_uno = GrupoEstudiante.objects.create(
            ficha_inscripcion=ficha_uno,
            estudiante=estudiante_uno,
            grupo=self.curso,
        )
        asignacion_dos = GrupoEstudiante.objects.create(
            ficha_inscripcion=ficha_dos,
            estudiante=estudiante_dos,
            grupo=self.curso,
        )
        ClaseEstudianteMovimiento.objects.create(
            asignacion=asignacion_dos,
            clase_origen=origin_class,
            clase_destino=destination_class,
        )
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:docente_clase_asistencia", args=[origin_class.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "attendance-status-group")
        self.assertEqual([row["estudiante"] for row in response.context["rows"]], [estudiante_uno])
        self.assertEqual(response.context["moved_out_rows"][0]["estudiante"], estudiante_dos)

        response = self.client.post(
            reverse("academico:docente_clase_asistencia", args=[origin_class.pk]),
            {
                f"estado_{asignacion_uno.pk}": "ausente",
                f"observacion_{asignacion_uno.pk}": "No asistio.",
                f"estado_{asignacion_dos.pk}": "presente",
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        asistencia = ClaseAsistencia.objects.get(clase=origin_class, estudiante=estudiante_uno)
        self.assertEqual(asistencia.estado, "ausente")
        self.assertEqual(asistencia.observacion, "No asistio.")
        self.assertFalse(ClaseAsistencia.objects.filter(clase=origin_class, estudiante=estudiante_dos).exists())

        destination_response = self.client.get(
            reverse("academico:docente_clase_asistencia", args=[destination_class.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(destination_response.status_code, 200)
        self.assertEqual(len(destination_response.context["rows"]), 1)
        incoming_rows = [row for row in destination_response.context["rows"] if row["incoming"]]
        self.assertEqual(incoming_rows[0]["estudiante"], estudiante_dos)

        roster_view = DocenteClaseAsistenciaView()
        future_origin_rows, future_origin_moved_out = roster_view.get_roster_rows(future_origin_class)
        future_destination_rows, _ = roster_view.get_roster_rows(future_destination_class)

        self.assertEqual([row["estudiante"] for row in future_origin_rows], [estudiante_uno])
        self.assertEqual(future_origin_moved_out[0]["estudiante"], estudiante_dos)
        self.assertEqual([row["estudiante"] for row in future_destination_rows], [estudiante_dos])
        self.assertTrue(future_destination_rows[0]["incoming"])

    def test_docente_can_close_attendance_after_save_and_lock_changes(self):
        today_class = self.create_class_for_date(timezone.localdate(), time(10, 0), time(11, 0), "Aula cierre")
        estudiante, ficha = self.create_student_ficha()
        asignacion = GrupoEstudiante.objects.create(
            ficha_inscripcion=ficha,
            estudiante=estudiante,
            grupo=self.curso,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_clase_asistencia", args=[today_class.pk]),
            {
                "attendance_action": "save",
                f"estado_{asignacion.pk}": "presente",
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        page_response = self.client.get(
            reverse("academico:docente_clase_asistencia", args=[today_class.pk]),
            HTTP_HOST="localhost",
        )
        self.assertContains(page_response, "Cerrar asistencia")

        response = self.client.post(
            reverse("academico:docente_clase_asistencia", args=[today_class.pk]),
            {"attendance_action": "close"},
            HTTP_HOST="localhost",
        )
        today_class.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertTrue(today_class.asistencia_cerrada)
        self.assertEqual(today_class.asistencia_cerrada_por, self.docente)
        self.assertIsNotNone(today_class.fecha_cierre_asistencia)

        response = self.client.post(
            reverse("academico:docente_clase_asistencia", args=[today_class.pk]),
            {
                "attendance_action": "save",
                f"estado_{asignacion.pk}": "ausente",
            },
            HTTP_HOST="localhost",
        )
        asistencia = ClaseAsistencia.objects.get(clase=today_class, estudiante=estudiante)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(asistencia.estado, "presente")
        locked_response = self.client.get(
            reverse("academico:docente_clase_asistencia", args=[today_class.pk]),
            HTTP_HOST="localhost",
        )
        self.assertContains(locked_response, "Asistencia cerrada")
        self.assertContains(locked_response, "Registro bloqueado")

    def test_docente_dashboard_renders_panel_cards(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("academico:docente_horarios"), HTTP_HOST="localhost")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mis planificaciones")
        self.assertContains(response, "Por atender")
        self.assertContains(response, "docente-planning-card")
        self.assertEqual(response.context["planificacion_stats"]["total"], 4)
        self.assertEqual(response.context["planificacion_stats"]["por_atender"], 3)

    def test_docente_dashboard_groups_planifications_by_topic(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("academico:docente_horarios"), HTTP_HOST="localhost")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Planificaciones por tema")
        self.assertContains(response, self.tema.nombre)
        self.assertEqual(len(response.context["tema_cards"]), 1)
        self.assertEqual(response.context["tema_cards"][0]["tema"], self.tema)
        self.assertEqual(response.context["tema_cards"][0]["available_count"], 3)

    def test_docente_calendar_view_renders_week_calendar_and_export_link(self):
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:docente_calendario"),
            {"fecha": self.revision.fecha.isoformat()},
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mi calendario")
        self.assertContains(response, "docente-calendar-full-panel")
        self.assertContains(response, "Descargar Excel")
        self.assertContains(response, "agendaWeek")
        self.assertTrue(response.context["has_events"])
        self.assertGreaterEqual(response.context["week_count"], 1)

    def test_docente_calendar_export_returns_weekly_excel(self):
        self.revision.tema = self.tema
        self.revision.subtema = self.subtema
        self.revision.save(update_fields=["tema", "subtema"])
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:docente_calendario_exportar"),
            {"fecha": self.revision.fecha.isoformat()},
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value]
        schedule_cell = self.find_cell_containing(sheet, "Matematicas")

        self.assertTrue(any("Matematicas" in str(value) for value in values))
        self.assertTrue(any("Grupo A" in str(value) for value in values))
        self.assertIn("Numeros", schedule_cell.value)
        self.assertIn("Suma", schedule_cell.value)
        self.assertTrue(schedule_cell.alignment.wrap_text)
        self.assertEqual(schedule_cell.alignment.vertical, "top")
        self.assertEqual(sheet.column_dimensions["B"].width, 30)
        self.assertGreaterEqual(sheet.row_dimensions[schedule_cell.row].height, 78)

    def test_academic_planning_export_fits_and_combines_overlapping_slots(self):
        self.make_superuser()
        another_course = Curso.objects.create(nombre="Grupo B", activo=True)
        another_aula = Aula.objects.create(nombre="Aula 2")
        another_aula_curso = AulaCurso.objects.create(aula=another_aula, curso=another_course)
        another_horario_aula_curso = HorarioAulaCurso.objects.create(
            aula_curso=another_aula_curso,
            horario_dia=self.horario_aula_curso.horario_dia,
        )
        another_materia_curso = MateriaCurso.objects.create(materia=self.materia, grupo=another_course)
        ProfesorMateriaCurso.objects.create(partner=self.docente, materia_curso=another_materia_curso)
        Clase.objects.create(
            horario_aula_curso=another_horario_aula_curso,
            materia_curso=another_materia_curso,
            fecha=self.pendiente.fecha,
        )
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:planificacion_academica_exportar"),
            {"tipo": "general"},
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook["General"]
        schedule_cell = self.find_cell_containing_all(sheet, "Grupo A", "Grupo B")

        self.assertIn("Grupo B", schedule_cell.value)
        self.assertIn("\n\n", schedule_cell.value)
        self.assertTrue(schedule_cell.alignment.wrap_text)
        self.assertEqual(schedule_cell.alignment.vertical, "top")
        self.assertEqual(sheet.column_dimensions["B"].width, 30)
        self.assertGreaterEqual(sheet.row_dimensions[schedule_cell.row].height, 130)

    def test_academic_planning_updates_future_class_without_delete_error(self):
        self.create_periodo_for_course()
        self.make_superuser()
        self.pendiente.tema = self.tema
        self.pendiente.subtema = self.subtema
        self.pendiente.descripcion = "Planificacion previa."
        self.pendiente.estado_planificacion = "rechazada"
        self.pendiente.save(update_fields=["tema", "subtema", "descripcion", "estado_planificacion"])
        self.pendiente.competencias.add(self.competencia)
        self.pendiente.estrategias.add(self.estrategia)
        self.pendiente.recursos.add(self.recurso)
        nueva_materia = Materia.objects.create(nombre="Lenguaje", nombre_corto="LEN", color="#2563eb")
        nueva_materia_curso = MateriaCurso.objects.create(materia=nueva_materia, grupo=self.curso)
        nueva_planificacion = PlanificacionDocente.objects.create(
            materia_curso=nueva_materia_curso,
            nombre="Plan lenguaje",
        )
        Tema.objects.create(planificacion=nueva_planificacion, nombre="Lectura", orden=1)
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "curso": self.curso.pk,
                "horario_aula_curso": self.horario_aula_curso.pk,
                "fecha": self.pendiente.fecha.isoformat(),
                "materia": nueva_materia.pk,
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.pendiente.materia_curso.materia, nueva_materia)
        self.assertEqual(self.pendiente.estado_planificacion, "pendiente")
        self.assertIsNone(self.pendiente.tema)
        self.assertFalse(self.pendiente.recursos.exists())

    def test_academic_planning_selector_only_lists_subjects_with_topics(self):
        self.create_periodo_for_course()
        self.make_superuser()
        nueva_materia = Materia.objects.create(nombre="Lenguaje", nombre_corto="LEN", color="#2563eb")
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:planificacion_academica"),
            {"curso": self.curso.pk},
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f'<option value="{self.materia.pk}">Matematicas</option>', html=True)
        self.assertNotContains(response, f'<option value="{nueva_materia.pk}">Lenguaje</option>', html=True)
        self.assertEqual(response.context["materias_asignables"], [self.materia])

    def test_academic_planning_rejects_subject_without_topics(self):
        self.create_periodo_for_course()
        self.make_superuser()
        nueva_materia = Materia.objects.create(nombre="Lenguaje", nombre_corto="LEN", color="#2563eb")
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "curso": self.curso.pk,
                "horario_aula_curso": self.horario_aula_curso.pk,
                "fecha": self.pendiente.fecha.isoformat(),
                "materia": nueva_materia.pk,
            },
            follow=True,
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.pendiente.materia_curso, self.materia_curso)
        self.assertFalse(MateriaCurso.objects.filter(materia=nueva_materia, grupo=self.curso).exists())
        self.assertContains(response, "Solo puedes asignar materias que ya tienen temas cargados para este grupo.")

    def test_academic_planning_can_assign_subject_with_base_topics_to_group(self):
        self.create_periodo_for_course()
        self.make_superuser()
        nueva_materia = Materia.objects.create(nombre="Lenguaje", nombre_corto="LEN", color="#2563eb")
        materia_tema = MateriaTema.objects.create(materia=nueva_materia, nombre="Lectura", orden=1)
        MateriaSubtema.objects.create(tema=materia_tema, nombre="Comprension", orden=1)
        self.client.force_login(self.user)

        page_response = self.client.get(
            reverse("academico:planificacion_academica"),
            {"curso": self.curso.pk},
            HTTP_HOST="localhost",
        )
        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "curso": self.curso.pk,
                "horario_aula_curso": self.horario_aula_curso.pk,
                "fecha": self.pendiente.fecha.isoformat(),
                "materia": nueva_materia.pk,
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()

        self.assertEqual(page_response.status_code, 200)
        self.assertContains(page_response, f'<option value="{nueva_materia.pk}">Lenguaje</option>', html=True)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.pendiente.materia_curso.materia, nueva_materia)
        generated_tema = self.pendiente.materia_curso.planificaciones.get().temas_planificacion.get(nombre="Lectura")
        self.assertEqual(generated_tema.materia_tema, materia_tema)
        self.assertTrue(generated_tema.subtemas_planificacion.filter(nombre="Comprension").exists())

    def test_academic_planning_calendar_opens_on_today_inside_period(self):
        today = timezone.localdate()
        periodo = Periodo.objects.create(
            nombre="Periodo vigente",
            fecha_inicio=today - timedelta(days=14),
            fecha_fin=today + timedelta(days=35),
        )
        CursoPeriodo.objects.create(curso=self.curso, periodo=periodo)
        self.make_superuser()
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:planificacion_academica"),
            {"curso": self.curso.pk},
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["calendar_default_date"], today.isoformat())

    def test_periodo_edit_loads_dates_with_modern_datepicker(self):
        self.make_superuser()
        today = timezone.localdate()
        periodo = Periodo.objects.create(
            nombre="Periodo editable",
            fecha_inicio=today,
            fecha_fin=today + timedelta(days=30),
        )
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:periodo_editar", args=[periodo.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "flatpickr.min.css")
        self.assertContains(response, "InstitutoDatePicker.init")
        self.assertContains(response, "js-date-picker")
        self.assertContains(response, f'value="{today.isoformat()}"')
        self.assertContains(response, f'value="{(today + timedelta(days=30)).isoformat()}"')

    def test_materia_form_uses_custom_color_picker(self):
        self.make_superuser()
        self.client.force_login(self.user)

        response = self.client.get(reverse("academico:materia_nueva"), HTTP_HOST="localhost")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "subject-color-picker")
        self.assertContains(response, "data-color-trigger")
        self.assertContains(response, "data-color-popover")
        self.assertContains(response, "data-color-honeycomb")
        self.assertContains(response, "subject-color-row")
        self.assertContains(response, "dataset.colorTone")
        self.assertContains(response, "data-color-preview")
        self.assertContains(response, 'data-color-input=""')
        self.assertNotContains(response, 'type="color"')

    def test_materia_form_saves_hex_color(self):
        self.make_superuser()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:materia_nueva"),
            {
                "nombre": "Historia",
                "nombre_corto": "HIS",
                "color": "0F766E",
                "descripcion": "",
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(Materia.objects.get(nombre="Historia").color, "#0f766e")

    def test_coordinacion_topic_create_only_selects_subject(self):
        coordinator = self.create_coordinator()
        self.client.force_login(coordinator)

        response = self.client.get(
            reverse("academico:coordinacion_planificacion_nueva"),
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Materia")
        self.assertContains(response, 'name="materia"')
        self.assertNotContains(response, "Materia / grupo")
        self.assertNotContains(response, 'name="materia_curso"')
        self.assertNotContains(response, "Docente asignado")

    def test_coordinacion_topic_create_applies_subject_topics_to_existing_groups(self):
        coordinator = self.create_coordinator()
        grupo_b = Curso.objects.create(nombre="Grupo B", activo=True)
        materia_curso_b = MateriaCurso.objects.create(materia=self.materia, grupo=grupo_b)
        self.client.force_login(coordinator)

        response = self.client.post(
            reverse("academico:coordinacion_planificacion_nueva"),
            {
                "materia": self.materia.pk,
                "form-TOTAL_FORMS": "1",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "1000",
                "form-0-tema_id": "",
                "form-0-nombre": "Algebra",
                "form-0-detalle": "",
                "form-0-orden": "1",
                "form-0-subtemas-TOTAL_FORMS": "1",
                "form-0-subtemas-0-id": "",
                "form-0-subtemas-0-nombre": "Polinomios",
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            reverse("academico:coordinacion_planificacion_materia_editar", args=[self.materia.pk]),
        )
        materia_tema = MateriaTema.objects.get(materia=self.materia, nombre="Algebra")
        self.assertTrue(materia_tema.subtemas_base.filter(nombre="Polinomios").exists())
        for materia_curso in [self.materia_curso, materia_curso_b]:
            planificacion = PlanificacionDocente.objects.get(materia_curso=materia_curso)
            tema = planificacion.temas_planificacion.get(nombre="Algebra")
            self.assertEqual(tema.materia_tema, materia_tema)
            self.assertTrue(tema.subtemas_planificacion.filter(nombre="Polinomios").exists())

    def test_coordinacion_topic_create_allows_subject_without_existing_group_link(self):
        coordinator = self.create_coordinator()
        nueva_materia = Materia.objects.create(nombre="Fisica", nombre_corto="FIS", color="#0891b2")
        self.client.force_login(coordinator)

        response = self.client.post(
            reverse("academico:coordinacion_planificacion_nueva"),
            {
                "materia": nueva_materia.pk,
                "form-TOTAL_FORMS": "1",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "1000",
                "form-0-tema_id": "",
                "form-0-nombre": "Movimiento",
                "form-0-detalle": "",
                "form-0-orden": "1",
                "form-0-subtemas-TOTAL_FORMS": "1",
                "form-0-subtemas-0-id": "",
                "form-0-subtemas-0-nombre": "Velocidad",
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            reverse("academico:coordinacion_planificacion_materia_editar", args=[nueva_materia.pk]),
        )
        materia_tema = MateriaTema.objects.get(materia=nueva_materia, nombre="Movimiento")
        self.assertTrue(materia_tema.subtemas_base.filter(nombre="Velocidad").exists())
        self.assertFalse(MateriaCurso.objects.filter(materia=nueva_materia).exists())

    def test_academic_planning_can_remove_subject_assignment_with_cleared_selector(self):
        self.create_periodo_for_course()
        self.make_superuser()
        self.pendiente.tema = self.tema
        self.pendiente.save(update_fields=["tema"])
        self.pendiente.sync_subtemas_planificados([self.subtema])
        self.client.force_login(self.user)

        page_response = self.client.get(
            reverse("academico:planificacion_academica"),
            {"curso": self.curso.pk},
            HTTP_HOST="localhost",
        )

        self.assertEqual(page_response.status_code, 200)
        self.assertContains(page_response, "allowClear: true")
        self.assertContains(page_response, "select2:clearing")
        self.assertContains(page_response, "select2:opening")
        self.assertNotContains(page_response, "Quitar materia de esta clase")

        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "curso": self.curso.pk,
                "horario_aula_curso": self.horario_aula_curso.pk,
                "fecha": self.pendiente.fecha.isoformat(),
                "materia": "",
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Clase.objects.filter(pk=self.pendiente.pk).exists())

    def test_academic_planning_does_not_update_approved_class(self):
        self.create_periodo_for_course()
        self.make_superuser()
        nueva_materia = Materia.objects.create(nombre="Lenguaje", nombre_corto="LEN", color="#2563eb")
        self.revision.estado_planificacion = "aprobada"
        self.revision.save(update_fields=["estado_planificacion"])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "curso": self.curso.pk,
                "horario_aula_curso": self.horario_aula_curso.pk,
                "fecha": self.revision.fecha.isoformat(),
                "materia": nueva_materia.pk,
            },
            follow=True,
            HTTP_HOST="localhost",
        )
        self.revision.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.revision.materia_curso, self.materia_curso)
        self.assertFalse(MateriaCurso.objects.filter(materia=nueva_materia, grupo=self.curso).exists())
        self.assertContains(response, "La clase no se puede modificar porque tiene una planificacion enviada o aprobada.")
        self.assertContains(response, "approved-locked-event")

    def test_academic_planning_assigns_single_day_docente_override(self):
        self.create_periodo_for_course()
        self.make_superuser()
        reemplazo, reemplazo_user = self.create_docente()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "curso": self.curso.pk,
                "horario_aula_curso": self.horario_aula_curso.pk,
                "fecha": self.pendiente.fecha.isoformat(),
                "materia": self.materia.pk,
                "docente": reemplazo.pk,
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()
        self.revision.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.pendiente.docente, reemplazo)
        self.assertTrue(self.pendiente.docente_override)
        self.assertIsNone(self.revision.docente)
        profesor_materia_curso = ProfesorMateriaCurso.objects.get(partner=reemplazo, materia_curso=self.materia_curso)
        self.assertTrue(profesor_materia_curso.auto_generada_por_clases)
        self.assertTrue(PlanificacionTema.objects.filter(profesor_materia_curso=profesor_materia_curso, tema=self.tema).exists())

        self.client.force_login(reemplazo_user)
        docente_response = self.client.get(reverse("academico:docente_horarios"), HTTP_HOST="localhost")

        self.assertEqual(docente_response.status_code, 200)
        self.assertEqual(docente_response.context["planificacion_stats"]["total"], 1)
        self.assertEqual(len(docente_response.context["tema_cards"]), 1)

    def test_academic_planning_removing_class_cleans_auto_docente_topic_plans(self):
        self.create_periodo_for_course()
        self.make_superuser()
        reemplazo, _ = self.create_docente()
        self.client.force_login(self.user)

        assign_response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "curso": self.curso.pk,
                "horario_aula_curso": self.horario_aula_curso.pk,
                "fecha": self.pendiente.fecha.isoformat(),
                "materia": self.materia.pk,
                "docente": reemplazo.pk,
            },
            HTTP_HOST="localhost",
        )
        profesor_materia_curso = ProfesorMateriaCurso.objects.get(partner=reemplazo, materia_curso=self.materia_curso)

        remove_response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "curso": self.curso.pk,
                "horario_aula_curso": self.horario_aula_curso.pk,
                "fecha": self.pendiente.fecha.isoformat(),
                "materia": "",
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(assign_response.status_code, 302)
        self.assertEqual(remove_response.status_code, 302)
        self.assertFalse(Clase.objects.filter(pk=self.pendiente.pk).exists())
        self.assertFalse(ProfesorMateriaCurso.objects.filter(pk=profesor_materia_curso.pk).exists())
        self.assertFalse(PlanificacionTema.objects.filter(profesor_materia_curso_id=profesor_materia_curso.pk).exists())

    def test_docente_subject_assignment_removal_blocks_locked_class_planifications(self):
        self.make_superuser()
        self.revision.estado_planificacion = "aprobada"
        self.revision.save(update_fields=["estado_planificacion"])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_docente"),
            {
                "materia_curso": self.materia_curso.pk,
                "docente": "",
                "grupo": self.curso.pk,
            },
            follow=True,
            HTTP_HOST="localhost",
        )
        self.revision.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            ProfesorMateriaCurso.objects.filter(
                pk=self.profesor_materia_curso.pk,
                auto_generada_por_clases=False,
            ).exists()
        )
        self.assertFalse(self.revision.docente_override)
        self.assertContains(response, "No se puede quitar el docente porque hay clases con planificacion enviada o aprobada.")

    def test_docente_subject_assignment_replacement_preserves_locked_classes_for_previous_teacher(self):
        self.make_superuser()
        reemplazo, reemplazo_user = self.create_docente()
        self.revision.estado_planificacion = "aprobada"
        self.revision.save(update_fields=["estado_planificacion"])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_docente"),
            {
                "materia_curso": self.materia_curso.pk,
                "docente": reemplazo.pk,
                "grupo": self.curso.pk,
            },
            follow=True,
            HTTP_HOST="localhost",
        )
        self.revision.refresh_from_db()
        self.pendiente.refresh_from_db()
        previous_assignment = ProfesorMateriaCurso.objects.get(partner=self.docente, materia_curso=self.materia_curso)
        new_assignment = ProfesorMateriaCurso.objects.get(partner=reemplazo, materia_curso=self.materia_curso)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(previous_assignment.auto_generada_por_clases)
        self.assertFalse(new_assignment.auto_generada_por_clases)
        self.assertEqual(self.revision.docente, self.docente)
        self.assertTrue(self.revision.docente_override)
        self.assertIsNone(self.pendiente.docente)
        self.assertFalse(self.pendiente.docente_override)
        self.assertTrue(
            PlanificacionTema.objects.filter(profesor_materia_curso=previous_assignment, tema=self.tema).exists()
        )
        self.assertTrue(
            PlanificacionTema.objects.filter(profesor_materia_curso=new_assignment, tema=self.tema).exists()
        )

        self.client.force_login(reemplazo_user)
        docente_response = self.client.get(reverse("academico:docente_horarios"), HTTP_HOST="localhost")

        self.assertEqual(docente_response.status_code, 200)
        self.assertEqual(docente_response.context["planificacion_stats"]["total"], 3)
        self.assertNotIn(self.revision, [card["clase"] for card in docente_response.context["planificacion_cards"]])

        self.client.force_login(self.user)
        previous_docente_response = self.client.get(reverse("academico:docente_horarios"), HTTP_HOST="localhost")

        self.assertEqual(previous_docente_response.status_code, 200)
        self.assertEqual(previous_docente_response.context["planificacion_stats"]["total"], 1)

    def test_docente_bulk_assignment_removal_blocks_locked_class_planifications(self):
        self.make_superuser()
        self.revision.estado_planificacion = "aprobada"
        self.revision.save(update_fields=["estado_planificacion"])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_docente_editar", args=[self.docente.pk]),
            {
                "docente": self.docente.pk,
                "form-TOTAL_FORMS": "1",
                "form-INITIAL_FORMS": "1",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "1000",
                "form-0-grupo": self.curso.pk,
                "form-0-materia_curso": self.materia_curso.pk,
                "form-0-DELETE": "on",
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            ProfesorMateriaCurso.objects.filter(
                pk=self.profesor_materia_curso.pk,
                auto_generada_por_clases=False,
            ).exists()
        )
        self.assertContains(response, "No se puede quitar el docente porque hay clases con planificacion enviada o aprobada")

    def test_academic_planning_assigns_docente_from_selected_date_forward(self):
        self.create_periodo_for_course()
        self.make_superuser()
        reemplazo, reemplazo_user = self.create_docente()
        selected_date = self.rechazada.fecha
        future_date = selected_date + timedelta(days=7)
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "curso": self.curso.pk,
                "horario_aula_curso": self.horario_aula_curso.pk,
                "fecha": selected_date.isoformat(),
                "materia": self.materia.pk,
                "docente": reemplazo.pk,
                "asignar_periodo": "on",
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()
        self.revision.refresh_from_db()
        self.rechazada.refresh_from_db()
        future_class = Clase.objects.get(horario_aula_curso=self.horario_aula_curso, fecha=future_date)
        previous_assignment = ProfesorMateriaCurso.objects.get(partner=self.docente, materia_curso=self.materia_curso)
        new_assignment = ProfesorMateriaCurso.objects.get(partner=reemplazo, materia_curso=self.materia_curso)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.pendiente.docente, self.docente)
        self.assertTrue(self.pendiente.docente_override)
        self.assertEqual(self.revision.docente, self.docente)
        self.assertTrue(self.revision.docente_override)
        self.assertIsNone(self.rechazada.docente)
        self.assertFalse(self.rechazada.docente_override)
        self.assertIsNone(future_class.docente)
        self.assertFalse(future_class.docente_override)
        self.assertTrue(previous_assignment.auto_generada_por_clases)
        self.assertFalse(new_assignment.auto_generada_por_clases)

        self.client.force_login(reemplazo_user)
        replacement_response = self.client.get(
            reverse("academico:docente_clase_planificar", args=[self.rechazada.pk]),
            HTTP_HOST="localhost",
        )
        locked_response = self.client.get(
            reverse("academico:docente_clase_planificar", args=[self.revision.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(replacement_response.status_code, 200)
        self.assertEqual(locked_response.status_code, 404)

        self.client.force_login(self.user)
        assignment_response = self.client.get(
            reverse("academico:planificacion_docente"),
            {"grupo": self.curso.pk},
            HTTP_HOST="localhost",
        )

        self.assertEqual(assignment_response.status_code, 200)
        self.assertEqual(assignment_response.context["stats"]["asignadas"], 1)
        self.assertEqual(assignment_response.context["stats"]["pendientes"], 0)
        self.assertContains(assignment_response, "Docente Reemplazo")

    def test_academic_planning_can_leave_single_class_without_docente(self):
        self.create_periodo_for_course()
        self.make_superuser()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "curso": self.curso.pk,
                "horario_aula_curso": self.horario_aula_curso.pk,
                "fecha": self.pendiente.fecha.isoformat(),
                "materia": self.materia.pk,
                "docente": "__none__",
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertIsNone(self.pendiente.docente)
        self.assertTrue(self.pendiente.docente_override)

        old_docente_response = self.client.get(
            reverse("academico:docente_clase_planificar", args=[self.pendiente.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(old_docente_response.status_code, 404)

    def test_academic_planning_can_restore_materia_docente_assignment(self):
        self.create_periodo_for_course()
        self.make_superuser()
        self.pendiente.docente = None
        self.pendiente.docente_override = True
        self.pendiente.save(update_fields=["docente", "docente_override"])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "curso": self.curso.pk,
                "horario_aula_curso": self.horario_aula_curso.pk,
                "fecha": self.pendiente.fecha.isoformat(),
                "materia": self.materia.pk,
                "docente": "",
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertIsNone(self.pendiente.docente)
        self.assertFalse(self.pendiente.docente_override)

        docente_response = self.client.get(
            reverse("academico:docente_clase_planificar", args=[self.pendiente.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(docente_response.status_code, 200)

    def test_academic_planning_does_not_change_submitted_class_docente(self):
        self.create_periodo_for_course()
        self.make_superuser()
        reemplazo, _ = self.create_docente()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "curso": self.curso.pk,
                "horario_aula_curso": self.horario_aula_curso.pk,
                "fecha": self.revision.fecha.isoformat(),
                "materia": self.materia.pk,
                "docente": reemplazo.pk,
            },
            follow=True,
            HTTP_HOST="localhost",
        )
        self.revision.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(self.revision.docente)
        self.assertFalse(self.revision.docente_override)
        self.assertContains(response, "La clase no se puede modificar porque tiene una planificacion enviada o aprobada.")

    def test_academic_planning_adds_schedule_block(self):
        periodo = self.create_periodo_for_course()
        self.make_superuser()
        aula = Aula.objects.create(nombre="Aula 2")
        dia, _ = Dia.objects.get_or_create(dia="Martes")
        selected_date = self.date_for_weekday(periodo, 1)
        self.client.force_login(self.user)

        page_response = self.client.get(
            reverse("academico:planificacion_academica"),
            {"curso": self.curso.pk},
            HTTP_HOST="localhost",
        )

        self.assertContains(page_response, "planificacionHorarioOffcanvas")
        self.assertContains(page_response, "selectable: canAddSchedule")
        self.assertContains(page_response, "Generar en todo el periodo preseleccionado")
        self.assertContains(page_response, "data-schedule-periodo-switch")
        self.assertNotContains(page_response, "Agregar bloque al calendario")

        response = self.client.post(
            f"{reverse('academico:planificacion_academica')}?curso={self.curso.pk}",
            {
                "planning_action": "add_schedule",
                "generar_periodo": "on",
                "schedule_fecha": selected_date.isoformat(),
                "schedule-aula": aula.pk,
                "schedule-dia": dia.pk,
                "schedule-hora_inicio": "10:00",
                "schedule-hora_fin": "11:00",
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            HorarioAulaCurso.objects.filter(
                aula_curso__curso=self.curso,
                aula_curso__aula=aula,
                fecha__isnull=True,
                horario_dia__dia=dia,
                horario_dia__horario__hora_inicio=time(10, 0),
                horario_dia__horario__hora_fin=time(11, 0),
            ).exists()
        )

    def test_academic_planning_rejects_schedule_aula_overlap(self):
        periodo = self.create_periodo_for_course()
        self.make_superuser()
        dia = self.horario_aula_curso.horario_dia.dia
        selected_date = self.date_for_weekday(periodo, 0)
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "planning_action": "add_schedule",
                "curso": self.curso.pk,
                "generar_periodo": "on",
                "schedule_fecha": selected_date.isoformat(),
                "schedule-aula": self.aula.pk,
                "schedule-dia": dia.pk,
                "schedule-hora_inicio": "08:30",
                "schedule-hora_fin": "09:30",
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "El aula Aula 1 ya esta asignada")
        self.assertFalse(
            HorarioAulaCurso.objects.filter(
                aula_curso__curso=self.curso,
                aula_curso__aula=self.aula,
                horario_dia__dia=dia,
                horario_dia__horario__hora_inicio=time(8, 30),
                horario_dia__horario__hora_fin=time(9, 30),
            ).exists()
        )

    def test_academic_planning_adds_single_date_schedule_when_period_switch_off(self):
        periodo = self.create_periodo_for_course()
        self.make_superuser()
        aula = Aula.objects.create(nombre="Aula 3")
        dia, _ = Dia.objects.get_or_create(dia="Jueves")
        selected_date = self.date_for_weekday(periodo, 3)
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "planning_action": "add_schedule",
                "curso": self.curso.pk,
                "generar_periodo": "off",
                "schedule_fecha": selected_date.isoformat(),
                "schedule-aula": aula.pk,
                "schedule-dia": dia.pk,
                "schedule-hora_inicio": "12:00",
                "schedule-hora_fin": "13:00",
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        horario_aula_curso = HorarioAulaCurso.objects.get(
            aula_curso__curso=self.curso,
            aula_curso__aula=aula,
            fecha=selected_date,
            horario_dia__dia=dia,
            horario_dia__horario__hora_inicio=time(12, 0),
            horario_dia__horario__hora_fin=time(13, 0),
        )
        page_response = self.client.get(
            reverse("academico:planificacion_academica"),
            {"curso": self.curso.pk},
            HTTP_HOST="localhost",
        )
        events = [
            event
            for event in json.loads(page_response.context["calendar_events_json"])
            if event["horarioId"] == horario_aula_curso.pk
        ]

        self.assertEqual(len(events), 1)
        self.assertEqual(events[0]["fecha"], selected_date.isoformat())
        self.assertTrue(events[0]["singleDate"])

    def test_academic_planning_updates_schedule_block(self):
        periodo = self.create_periodo_for_course()
        self.make_superuser()
        aula_origen = Aula.objects.create(nombre="Aula 4")
        aula_destino = Aula.objects.create(nombre="Aula 5")
        dia, _ = Dia.objects.get_or_create(dia="Martes")
        horario = Horario.objects.create(hora_inicio=time(10, 0), hora_fin=time(11, 0))
        horario_dia = HorarioDia.objects.create(dia=dia, horario=horario)
        aula_curso = AulaCurso.objects.create(aula=aula_origen, curso=self.curso)
        horario_aula_curso = HorarioAulaCurso.objects.create(aula_curso=aula_curso, horario_dia=horario_dia)
        selected_date = self.date_for_weekday(periodo, 1)
        self.client.force_login(self.user)

        page_response = self.client.get(
            reverse("academico:planificacion_academica"),
            {"curso": self.curso.pk},
            HTTP_HOST="localhost",
        )

        self.assertContains(page_response, "data-edit-schedule-button")
        self.assertContains(page_response, "aulaId")
        self.assertContains(page_response, "horaInicio")

        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "planning_action": "update_schedule",
                "curso": self.curso.pk,
                "schedule_horario_aula_curso": horario_aula_curso.pk,
                "generar_periodo": "on",
                "schedule_fecha": selected_date.isoformat(),
                "schedule-aula": aula_destino.pk,
                "schedule-dia": dia.pk,
                "schedule-hora_inicio": "10:30",
                "schedule-hora_fin": "11:30",
            },
            HTTP_HOST="localhost",
        )
        horario_aula_curso.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(horario_aula_curso.aula_curso.aula, aula_destino)
        self.assertIsNone(horario_aula_curso.fecha)
        self.assertEqual(horario_aula_curso.horario_dia.horario.hora_inicio, time(10, 30))
        self.assertEqual(horario_aula_curso.horario_dia.horario.hora_fin, time(11, 30))

    def test_academic_planning_updates_schedule_block_to_single_date(self):
        periodo = self.create_periodo_for_course()
        self.make_superuser()
        aula = Aula.objects.create(nombre="Aula 6")
        dia, _ = Dia.objects.get_or_create(dia="Viernes")
        horario = Horario.objects.create(hora_inicio=time(14, 0), hora_fin=time(15, 0))
        horario_dia = HorarioDia.objects.create(dia=dia, horario=horario)
        aula_curso = AulaCurso.objects.create(aula=aula, curso=self.curso)
        horario_aula_curso = HorarioAulaCurso.objects.create(aula_curso=aula_curso, horario_dia=horario_dia)
        selected_date = self.date_for_weekday(periodo, 4)
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "planning_action": "update_schedule",
                "curso": self.curso.pk,
                "schedule_horario_aula_curso": horario_aula_curso.pk,
                "generar_periodo": "off",
                "schedule_fecha": selected_date.isoformat(),
                "schedule-aula": aula.pk,
                "schedule-dia": dia.pk,
                "schedule-hora_inicio": "14:00",
                "schedule-hora_fin": "15:00",
            },
            HTTP_HOST="localhost",
        )
        horario_aula_curso.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(horario_aula_curso.fecha, selected_date)

    def test_academic_planning_does_not_update_schedule_with_approved_class(self):
        periodo = self.create_periodo_for_course()
        self.make_superuser()
        aula = Aula.objects.create(nombre="Aula 7")
        dia, _ = Dia.objects.get_or_create(dia="Martes")
        horario = Horario.objects.create(hora_inicio=time(15, 0), hora_fin=time(16, 0))
        horario_dia = HorarioDia.objects.create(dia=dia, horario=horario)
        aula_curso = AulaCurso.objects.create(aula=aula, curso=self.curso)
        horario_aula_curso = HorarioAulaCurso.objects.create(aula_curso=aula_curso, horario_dia=horario_dia)
        selected_date = self.date_for_weekday(periodo, 1)
        Clase.objects.create(
            horario_aula_curso=horario_aula_curso,
            materia_curso=self.materia_curso,
            fecha=selected_date,
            estado_planificacion="aprobada",
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:planificacion_academica"),
            {
                "planning_action": "update_schedule",
                "curso": self.curso.pk,
                "schedule_horario_aula_curso": horario_aula_curso.pk,
                "generar_periodo": "on",
                "schedule_fecha": selected_date.isoformat(),
                "schedule-aula": aula.pk,
                "schedule-dia": dia.pk,
                "schedule-hora_inicio": "15:30",
                "schedule-hora_fin": "16:30",
            },
            HTTP_HOST="localhost",
        )
        horario_aula_curso.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No se puede modificar el horario porque tiene una planificacion enviada o aprobada.")
        self.assertEqual(horario_aula_curso.horario_dia.horario.hora_inicio, time(15, 0))
        self.assertEqual(horario_aula_curso.horario_dia.horario.hora_fin, time(16, 0))

    def test_docente_status_filter_limits_cards(self):
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:docente_horarios"),
            {"estado": "revision"},
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_filter"], "revision")
        self.assertEqual(len(response.context["planificacion_cards"]), 1)
        self.assertEqual(response.context["planificacion_cards"][0]["clase"], self.revision)

    def test_docente_observed_filter_limits_cards(self):
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:docente_horarios"),
            {"estado": "rechazada"},
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_filter"], "rechazada")
        self.assertEqual(len(response.context["planificacion_cards"]), 1)
        self.assertEqual(response.context["planificacion_cards"][0]["clase"], self.rechazada)

    def test_docente_class_planning_renders_workstation(self):
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:docente_clase_planificar", args=[self.pendiente.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "teacher-plan-hero")
        self.assertContains(response, "teacher-submit-panel")
        self.assertContains(response, "Guardar borrador")
        self.assertContains(response, "teacher-chip-pool")
        self.assertContains(response, "data-available-chip")
        self.assertContains(response, "subtemas_nuevos")
        self.assertContains(response, "data-selected-list")
        self.assertNotContains(response, "Usar disponible")
        self.assertNotContains(response, "Crear nuevo")
        self.assertEqual(response.context["planning_total"], 4)

    def test_docente_topic_planning_assigns_available_class(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_tema_planificar", args=[self.planificacion_tema.pk]),
            {
                "tema_action": "assign",
                "clase_id": self.pendiente.pk,
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            f"{reverse('academico:docente_tema_planificar', args=[self.planificacion_tema.pk])}#clase-{self.pendiente.pk}",
        )
        self.assertEqual(self.pendiente.tema, self.tema)
        self.assertIsNone(self.pendiente.subtema)
        self.assertEqual(self.pendiente.get_subtemas_planificados(), [])

    def test_docente_topic_planning_saves_multiple_subtopics_to_assigned_class(self):
        otro_subtema = Subtema.objects.create(tema=self.tema, nombre="Resta", orden=2)
        self.pendiente.tema = self.tema
        self.pendiente.save(update_fields=["tema"])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_tema_planificar", args=[self.planificacion_tema.pk]),
            {
                "tema_action": "save_class",
                "clase_id": self.pendiente.pk,
                "subtema_ids": [self.subtema.pk, otro_subtema.pk],
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            f"{reverse('academico:docente_tema_planificar', args=[self.planificacion_tema.pk])}#clase-{self.pendiente.pk}",
        )
        self.assertEqual(self.pendiente.tema, self.tema)
        self.assertEqual(self.pendiente.subtema, self.subtema)
        self.assertEqual(self.pendiente.get_subtemas_planificados(), [self.subtema, otro_subtema])

    def test_docente_topic_planning_renders_summary_cards_for_assigned_class(self):
        self.pendiente.tema = self.tema
        self.pendiente.save(update_fields=["tema"])
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:docente_tema_planificar", args=[self.planificacion_tema.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "topic-class-summary")
        self.assertContains(response, "topic-summary-block")
        self.assertContains(response, "Editar clase")
        self.assertContains(response, reverse("academico:docente_clase_planificar", args=[self.pendiente.pk]))
        self.assertContains(response, "from_planificacion_tema")
        self.assertContains(response, "Envia a revision la clase agregada antes de tomar otra clase")
        self.assertNotContains(response, "data-topic-class-inline-form")
        self.assertNotContains(response, "data-open-class-picker")
        self.assertNotContains(response, "data-class-picker")
        self.assertNotContains(response, "Tomar clases para este tema")

    def test_docente_topic_planning_creates_new_subtema_from_inline_class(self):
        self.pendiente.tema = self.tema
        self.pendiente.save(update_fields=["tema"])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_tema_planificar", args=[self.planificacion_tema.pk]),
            {
                "tema_action": "save_class",
                "clase_id": self.pendiente.pk,
                "subtemas_nuevos": "Multiplicacion de polinomios",
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()
        nuevo_subtema = Subtema.objects.get(tema=self.tema, nombre="Multiplicacion de polinomios")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(nuevo_subtema.orden, 2)
        self.assertEqual(self.pendiente.get_subtemas_planificados(), [nuevo_subtema])

    def test_docente_topic_planning_hides_subtopics_used_by_another_class(self):
        self.pendiente.tema = self.tema
        self.pendiente.save(update_fields=["tema"])
        self.pendiente.sync_subtemas_planificados([self.subtema])
        otra_clase = self.create_class_for_date(
            self.pendiente.fecha + timedelta(days=21),
            time(10, 0),
            time(11, 0),
            aula_nombre="Aula subtema disponible",
        )
        otra_clase.tema = self.tema
        otra_clase.save(update_fields=["tema"])
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:docente_tema_planificar", args=[self.planificacion_tema.pk]),
            HTTP_HOST="localhost",
        )
        assigned_slots = response.context["assigned_classes"]
        otra_clase_slot = next(item for item in assigned_slots if item["clase"] == otra_clase)
        visible_subtemas = [item["subtema"] for item in otra_clase_slot["subtema_options"]]

        self.assertEqual(response.status_code, 200)
        self.assertNotIn(self.subtema, visible_subtemas)

    def test_docente_topic_planning_hides_available_classes_when_subtopics_are_done(self):
        self.pendiente.tema = self.tema
        self.pendiente.estado_planificacion = "revision"
        self.pendiente.save(update_fields=["tema", "estado_planificacion"])
        self.pendiente.sync_subtemas_planificados([self.subtema])
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:docente_tema_planificar", args=[self.planificacion_tema.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["pending_subtema_count"], 0)
        self.assertEqual(response.context["available_classes"], [])
        self.assertContains(response, "Todos los subtemas del tema ya fueron planificados.")

    def test_docente_topic_planning_blocks_next_class_until_current_is_sent(self):
        self.pendiente.tema = self.tema
        self.pendiente.save(update_fields=["tema"])
        otra_clase = self.create_class_for_date(
            self.pendiente.fecha + timedelta(days=21),
            time(10, 0),
            time(11, 0),
            aula_nombre="Aula clase bloqueada",
        )
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:docente_tema_planificar", args=[self.planificacion_tema.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["available_classes"], [])
        self.assertContains(response, "Envia a revision la clase agregada antes de tomar otra clase")

        response = self.client.post(
            reverse("academico:docente_tema_planificar", args=[self.planificacion_tema.pk]),
            {
                "tema_action": "assign",
                "clase_id": otra_clase.pk,
            },
            HTTP_HOST="localhost",
        )
        otra_clase.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertIsNone(otra_clase.tema)

    def test_docente_topic_planning_does_not_assign_more_classes_when_subtopics_are_done(self):
        self.pendiente.tema = self.tema
        self.pendiente.save(update_fields=["tema"])
        self.pendiente.sync_subtemas_planificados([self.subtema])
        otra_clase = self.create_class_for_date(
            self.pendiente.fecha + timedelta(days=21),
            time(10, 0),
            time(11, 0),
            aula_nombre="Aula tema completo",
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_tema_planificar", args=[self.planificacion_tema.pk]),
            {
                "tema_action": "assign",
                "clase_id": otra_clase.pk,
            },
            HTTP_HOST="localhost",
        )
        otra_clase.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertIsNone(otra_clase.tema)

    def test_docente_topic_planning_does_not_reuse_subtopic_in_another_class(self):
        self.pendiente.tema = self.tema
        self.pendiente.save(update_fields=["tema"])
        self.pendiente.sync_subtemas_planificados([self.subtema])
        otra_clase = self.create_class_for_date(
            self.pendiente.fecha + timedelta(days=21),
            time(10, 0),
            time(11, 0),
            aula_nombre="Aula subtema duplicado",
        )
        otra_clase.tema = self.tema
        otra_clase.save(update_fields=["tema"])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_tema_planificar", args=[self.planificacion_tema.pk]),
            {
                "tema_action": "save_class",
                "clase_id": otra_clase.pk,
                "subtema_ids": [self.subtema.pk],
            },
            HTTP_HOST="localhost",
        )
        otra_clase.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(otra_clase.tema, self.tema)
        self.assertEqual(otra_clase.get_subtemas_planificados(), [])

    def test_docente_topic_planning_sends_inline_class_to_review(self):
        self.pendiente.tema = self.tema
        self.pendiente.save(update_fields=["tema"])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_tema_planificar", args=[self.planificacion_tema.pk]),
            {
                "tema_action": "send_class",
                "clase_id": self.pendiente.pk,
                "subtema_ids": [self.subtema.pk],
                "competencias_existentes": [self.competencia.pk],
                "estrategias_existentes": [self.estrategia.pk],
                "recursos_existentes": [self.recurso.pk],
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            f"{reverse('academico:docente_tema_planificar', args=[self.planificacion_tema.pk])}#clase-{self.pendiente.pk}",
        )
        self.assertEqual(self.pendiente.estado_planificacion, "revision")
        self.assertEqual(self.pendiente.get_subtemas_planificados(), [self.subtema])
        self.assertIn(self.competencia, self.pendiente.competencias.all())
        self.assertIn(self.estrategia, self.pendiente.estrategias.all())
        self.assertIn(self.recurso, self.pendiente.recursos.all())

    def test_docente_class_planning_from_topic_opens_class_editor(self):
        self.pendiente.tema = self.tema
        self.pendiente.save(update_fields=["tema"])
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:docente_clase_planificar", args=[self.pendiente.pk]),
            {"from_planificacion_tema": self.planificacion_tema.pk},
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Planificar clase")
        self.assertContains(response, self.tema.nombre)

    def test_docente_topic_planning_does_not_take_class_from_other_topic(self):
        other_topic = Tema.objects.create(planificacion=self.planificacion, nombre="Geometria", orden=2)
        self.pendiente.tema = other_topic
        self.pendiente.save(update_fields=["tema"])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_tema_planificar", args=[self.planificacion_tema.pk]),
            {
                "tema_action": "assign",
                "clase_id": self.pendiente.pk,
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.pendiente.tema, other_topic)

    def test_docente_topic_planning_unassigns_draft_class(self):
        self.pendiente.tema = self.tema
        self.pendiente.subtema = self.subtema
        self.pendiente.save(update_fields=["tema", "subtema"])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_tema_planificar", args=[self.planificacion_tema.pk]),
            {
                "tema_action": "unassign",
                "clase_id": self.pendiente.pk,
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertIsNone(self.pendiente.tema)
        self.assertIsNone(self.pendiente.subtema)
        self.assertEqual(self.pendiente.get_subtemas_planificados(), [])

    def test_docente_class_planning_uses_single_class_docente_override(self):
        reemplazo, reemplazo_user = self.create_docente()
        self.pendiente.docente = reemplazo
        self.pendiente.docente_override = True
        self.pendiente.save(update_fields=["docente", "docente_override"])

        self.client.force_login(self.user)
        old_docente_response = self.client.get(
            reverse("academico:docente_clase_planificar", args=[self.pendiente.pk]),
            HTTP_HOST="localhost",
        )

        self.client.force_login(reemplazo_user)
        reemplazo_response = self.client.get(
            reverse("academico:docente_clase_planificar", args=[self.pendiente.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(old_docente_response.status_code, 404)
        self.assertEqual(reemplazo_response.status_code, 200)
        self.assertContains(reemplazo_response, "teacher-plan-hero")

    def test_docente_class_planning_draft_keeps_planification_pending(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_clase_planificar", args=[self.pendiente.pk]),
            {
                "plan_action": "draft",
                "descripcion": "Borrador en progreso.",
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.pendiente.estado_planificacion, "pendiente")
        self.assertEqual(self.pendiente.descripcion, "Borrador en progreso.")

    def test_docente_class_planning_send_requires_complete_content(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_clase_planificar", args=[self.pendiente.pk]),
            {
                "plan_action": "send",
                "descripcion": "Solo detalle.",
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Selecciona el tema de la clase.")
        self.assertEqual(self.pendiente.estado_planificacion, "pendiente")

    def test_docente_class_planning_send_complete_content_to_review(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_clase_planificar", args=[self.pendiente.pk]),
            {
                "plan_action": "send",
                "tema": self.tema.pk,
                "subtema": self.subtema.pk,
                "descripcion": "Desarrollo completo de la clase.",
                "competencias_existentes": [self.competencia.pk],
                "estrategias_existentes": [self.estrategia.pk],
                "recursos_existentes": [self.recurso.pk],
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.pendiente.estado_planificacion, "revision")
        self.assertEqual(self.pendiente.tema, self.tema)
        self.assertEqual(self.pendiente.subtema, self.subtema)
        self.assertEqual(self.pendiente.get_subtemas_planificados(), [self.subtema])
        self.assertIn(self.competencia, self.pendiente.competencias.all())
        self.assertIn(self.estrategia, self.pendiente.estrategias.all())
        self.assertIn(self.recurso, self.pendiente.recursos.all())

    def test_docente_class_planning_creates_written_tags_and_subtopics(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_clase_planificar", args=[self.pendiente.pk]),
            {
                "from_planificacion_tema": self.planificacion_tema.pk,
                "plan_action": "send",
                "tema": self.tema.pk,
                "subtemas_nuevos": "Terminos semejantes\nPolinomios",
                "competencias_nuevos": "Opera polinomios con precision",
                "estrategias_nuevos": "Ejercicios en parejas",
                "recursos_nuevos": "Guia impresa",
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.pendiente.estado_planificacion, "revision")
        self.assertEqual(
            [subtema.nombre for subtema in self.pendiente.get_subtemas_planificados()],
            ["Terminos semejantes", "Polinomios"],
        )
        self.assertTrue(self.pendiente.competencias.filter(nombre="Opera polinomios con precision").exists())
        self.assertTrue(self.pendiente.estrategias.filter(nombre="Ejercicios en parejas").exists())
        self.assertTrue(self.pendiente.recursos.filter(nombre="Guia impresa").exists())

    def test_docente_class_planning_returns_to_topic_after_send(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_clase_planificar", args=[self.pendiente.pk]),
            {
                "from_tema": self.tema.pk,
                "from_planificacion_tema": self.planificacion_tema.pk,
                "plan_action": "send",
                "tema": self.tema.pk,
                "subtema": self.subtema.pk,
                "descripcion": "Desarrollo completo de la clase.",
                "competencias_existentes": [self.competencia.pk],
                "estrategias_existentes": [self.estrategia.pk],
                "recursos_existentes": [self.recurso.pk],
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            f"{reverse('academico:docente_tema_planificar', args=[self.planificacion_tema.pk])}#clase-{self.pendiente.pk}",
        )

    def test_docente_class_planning_creates_new_resource_with_file(self):
        self.client.force_login(self.user)
        uploaded = SimpleUploadedFile(
            "guia.pdf",
            b"contenido",
            content_type="application/pdf",
        )

        response = self.client.post(
            reverse("academico:docente_clase_planificar", args=[self.pendiente.pk]),
            {
                "plan_action": "send",
                "tema": self.tema.pk,
                "subtema": self.subtema.pk,
                "descripcion": "Desarrollo completo de la clase.",
                "competencias_existentes": [self.competencia.pk],
                "estrategias_existentes": [self.estrategia.pk],
                "recursos-TOTAL_FORMS": "1",
                "recursos-0-nombre": "Guia de ejercicios",
                "recursos-0-archivo": uploaded,
            },
            HTTP_HOST="localhost",
        )
        self.pendiente.refresh_from_db()
        recurso = Recurso.objects.get(nombre="Guia de ejercicios")
        clase_recurso = self.pendiente.clase_recursos.get(recurso=recurso)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.pendiente.estado_planificacion, "revision")
        self.assertIn(recurso, self.pendiente.recursos.all())
        self.assertTrue(clase_recurso.archivo.name.endswith(".pdf"))

        response = self.client.get(
            reverse("academico:docente_clase_planificar", args=[self.pendiente.pk]),
            HTTP_HOST="localhost",
        )

        self.assertContains(response, "teacher-file-current")
        self.assertContains(response, "file-kind-pdf")
        self.assertContains(response, "ri-file-pdf-line")

    def test_docente_class_planning_approved_planification_is_locked(self):
        self.revision.estado_planificacion = "aprobada"
        self.revision.save(update_fields=["estado_planificacion"])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("academico:docente_clase_planificar", args=[self.revision.pk]),
            {
                "plan_action": "send",
                "descripcion": "Cambio posterior.",
            },
            HTTP_HOST="localhost",
        )
        self.revision.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "La planificacion aprobada no se puede editar")
        self.assertEqual(self.revision.estado_planificacion, "aprobada")
        self.assertEqual(self.revision.descripcion, "Clase enviada a revision.")

    def test_coordinacion_review_dashboard_filters_by_docente(self):
        coordinator = self.create_coordinator()
        self.client.force_login(coordinator)

        response = self.client.get(
            reverse("academico:coordinacion_revision_planificaciones"),
            {"docente": self.docente.pk, "estado": "revision"},
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Enviadas")
        self.assertContains(response, "coordination-review-card")
        self.assertEqual(response.context["selected_docente"], self.docente)
        self.assertEqual(response.context["selected_estado"], "revision")
        self.assertEqual(len(response.context["revision_cards"]), 1)
        self.assertEqual(response.context["revision_cards"][0]["clase"], self.revision)

    def test_coordinacion_review_dashboard_shows_teacher_topic_and_subject_progress(self):
        otro_subtema = Subtema.objects.create(tema=self.tema, nombre="Resta", orden=2)
        self.revision.tema = self.tema
        self.revision.save(update_fields=["tema"])
        self.revision.sync_subtemas_planificados([self.subtema])
        coordinator = self.create_coordinator()
        self.client.force_login(coordinator)

        response = self.client.get(
            reverse("academico:coordinacion_revision_planificaciones"),
            {"docente": self.docente.pk, "estado": "revision"},
            HTTP_HOST="localhost",
        )
        card = response.context["revision_cards"][0]

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Historial del docente")
        self.assertContains(response, "Avance tema")
        self.assertContains(response, "Avance materia")
        self.assertEqual(card["topic_progress"]["covered"], 1)
        self.assertEqual(card["topic_progress"]["total"], 2)
        self.assertEqual(card["topic_progress"]["progress"], 50)
        self.assertEqual(card["materia_progress"]["covered"], 1)
        self.assertEqual(card["materia_progress"]["total"], 2)
        self.assertEqual(card["materia_progress"]["progress"], 50)
        self.assertEqual(response.context["docente_history"]["progress"]["progress"], 50)
        self.assertIn(otro_subtema, list(self.tema.subtemas_planificacion.all()))

    def test_coordinacion_review_dashboard_filters_by_docente_override(self):
        coordinator = self.create_coordinator()
        reemplazo, _ = self.create_docente()
        self.pendiente.docente = reemplazo
        self.pendiente.docente_override = True
        self.pendiente.save(update_fields=["docente", "docente_override"])
        self.client.force_login(coordinator)

        response = self.client.get(
            reverse("academico:coordinacion_revision_planificaciones"),
            {"docente": reemplazo.pk, "estado": "pendiente"},
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_docente"], reemplazo)
        self.assertEqual(len(response.context["revision_cards"]), 1)
        self.assertEqual(response.context["revision_cards"][0]["clase"], self.pendiente)

    def test_coordinacion_review_dashboard_shows_late_cards(self):
        coordinator = self.create_coordinator()
        self.client.force_login(coordinator)

        response = self.client.get(
            reverse("academico:coordinacion_revision_planificaciones"),
            {"docente": self.docente.pk, "estado": "atrasadas"},
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Atrasada")
        self.assertEqual(response.context["selected_estado"], "atrasadas")
        self.assertEqual(len(response.context["revision_cards"]), 1)
        self.assertEqual(response.context["revision_cards"][0]["clase"], self.atrasada)

    def test_coordinacion_attendance_review_filters_by_group_and_docente(self):
        coordinator = self.create_coordinator()
        estudiante_uno, ficha_uno = self.create_student_ficha(
            nombre="Ana Asistencia",
            identificacion="AST-001",
            numero="AST-001",
        )
        estudiante_dos, ficha_dos = self.create_student_ficha(
            nombre="Luis Asistencia",
            identificacion="AST-002",
            numero="AST-002",
        )
        GrupoEstudiante.objects.create(ficha_inscripcion=ficha_uno, estudiante=estudiante_uno, grupo=self.curso)
        GrupoEstudiante.objects.create(ficha_inscripcion=ficha_dos, estudiante=estudiante_dos, grupo=self.curso)
        ClaseAsistencia.objects.create(
            clase=self.revision,
            estudiante=estudiante_uno,
            estado="presente",
            registrado_por=self.docente,
        )
        ClaseAsistencia.objects.create(
            clase=self.revision,
            estudiante=estudiante_dos,
            estado="ausente",
            observacion="No asistio.",
            registrado_por=self.docente,
        )
        self.client.force_login(coordinator)

        response = self.client.get(
            reverse("academico:coordinacion_revision_asistencia"),
            {"grupo": self.curso.pk, "docente": self.docente.pk, "estado": "ausentes"},
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Revision de asistencia")
        self.assertContains(response, "attendance-review-card")
        self.assertContains(response, "Ver reporte")
        self.assertNotContains(response, "Ana Asistencia")
        self.assertNotContains(response, "No asistio.")
        self.assertEqual(response.context["selected_grupo"], self.curso)
        self.assertEqual(response.context["selected_docente"], self.docente)
        self.assertEqual(response.context["selected_estado"], "ausentes")
        self.assertEqual(len(response.context["attendance_cards"]), 1)
        card = response.context["attendance_cards"][0]
        self.assertEqual(card["clase"], self.revision)
        self.assertEqual(card["counts"]["presente"], 1)
        self.assertEqual(card["counts"]["ausente"], 1)
        self.assertEqual(card["observation_count"], 1)

        report_response = self.client.get(str(card["report_url"]), HTTP_HOST="localhost")

        self.assertEqual(report_response.status_code, 200)
        self.assertContains(report_response, "Reporte de asistencia")
        self.assertContains(report_response, "Ana Asistencia")
        self.assertContains(report_response, "No asistio.")
        self.assertEqual(report_response.context["card"]["counts"]["ausente"], 1)

    def test_coordinacion_student_attendance_report_exports_parent_excel(self):
        coordinator = self.create_coordinator()
        estudiante_uno, ficha_uno = self.create_student_ficha(
            nombre="Ana Padres",
            identificacion="PAD-001",
            numero="PAD-001",
        )
        estudiante_dos, ficha_dos = self.create_student_ficha(
            nombre="Luis Padres",
            identificacion="PAD-002",
            numero="PAD-002",
        )
        GrupoEstudiante.objects.create(ficha_inscripcion=ficha_uno, estudiante=estudiante_uno, grupo=self.curso)
        GrupoEstudiante.objects.create(ficha_inscripcion=ficha_dos, estudiante=estudiante_dos, grupo=self.curso)
        ClaseAsistencia.objects.create(
            clase=self.revision,
            estudiante=estudiante_uno,
            estado="ausente",
            observacion="No asistio por cita medica.",
            registrado_por=self.docente,
        )
        ClaseAsistencia.objects.create(
            clase=self.revision,
            estudiante=estudiante_dos,
            estado="presente",
            observacion="Otro alumno.",
            registrado_por=self.docente,
        )
        params = {
            "grupo": self.curso.pk,
            "estudiante": estudiante_uno.pk,
            "desde": self.pendiente.fecha.isoformat(),
            "hasta": self.revision.fecha.isoformat(),
        }
        self.client.force_login(coordinator)

        response = self.client.get(reverse("academico:coordinacion_reporte_asistencia_alumno"), params, HTTP_HOST="localhost")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Reporte de asistencia del alumno")
        self.assertContains(response, "Ana Padres")
        self.assertContains(response, "Matematicas")
        self.assertContains(response, "No asistio por cita medica.")
        self.assertNotContains(response, "Otro alumno.")
        self.assertEqual(response.context["stats"]["total"], 2)
        self.assertEqual(response.context["stats"]["ausente"], 1)
        self.assertEqual(response.context["stats"]["pendiente"], 1)
        self.assertFalse(
            any(row["observacion"] == "Otro alumno." for row in response.context["rows"])
        )

        export_response = self.client.get(
            reverse("academico:coordinacion_reporte_asistencia_alumno"),
            {**params, "export": "excel"},
            HTTP_HOST="localhost",
        )

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(export_response.content))
        sheet = workbook.active
        values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value]

        self.assertIn("Reporte de asistencia del alumno", values)
        self.assertIn("Ana Padres", values)
        self.assertIn("Matematicas", values)
        self.assertIn("Ausente", values)
        self.assertIn("No asistio por cita medica.", values)
        self.assertNotIn("Luis Padres", values)

    def test_director_attendance_review_opens_without_docente_partner(self):
        director = get_user_model().objects.create_user(username="director-asistencia", password="ClaveActual987!")
        director.groups.add(Group.objects.get_or_create(name="Director")[0])
        estudiante, ficha = self.create_student_ficha()
        GrupoEstudiante.objects.create(ficha_inscripcion=ficha, estudiante=estudiante, grupo=self.curso)
        self.client.force_login(director)

        response = self.client.get(
            reverse("academico:coordinacion_revision_asistencia"),
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Revision de asistencia")
        self.assertContains(response, "Revision asistencia")
        self.assertGreaterEqual(response.context["attendance_stats"]["total"], 4)
        self.assertGreaterEqual(response.context["attendance_stats"]["pendientes_registro"], 4)

    def test_director_can_register_teacher_replacement_hours(self):
        director = self.create_director()
        reemplazo, _ = self.create_docente()
        clase = self.create_class_for_date(timezone.localdate(), time(10, 0), time(12, 0), aula_nombre="Aula pago")
        self.client.force_login(director)

        response = self.client.post(
            reverse("academico:direccion_horas_docente"),
            {
                "fecha": clase.fecha.isoformat(),
                "clase": clase.pk,
                f"hora_{clase.pk}-estado": "reemplazo",
                f"hora_{clase.pk}-docente": reemplazo.pk,
                f"hora_{clase.pk}-horas": "1.50",
                f"hora_{clase.pk}-observacion": "Reemplazo autorizado.",
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        registro = ClaseHoraDocente.objects.get(clase=clase)
        self.assertEqual(registro.estado, "reemplazo")
        self.assertEqual(registro.docente, reemplazo)
        self.assertEqual(registro.docente_reemplazado, self.docente)
        self.assertEqual(registro.horas, Decimal("1.50"))
        self.assertEqual(registro.registrado_por, director)

    def test_teacher_hours_asistio_ignores_posted_replacement_teacher(self):
        director = self.create_director()
        reemplazo, _ = self.create_docente()
        clase = self.create_class_for_date(timezone.localdate(), time(10, 0), time(12, 0), aula_nombre="Aula pago")
        self.client.force_login(director)

        response = self.client.post(
            reverse("academico:direccion_horas_docente"),
            {
                "fecha": clase.fecha.isoformat(),
                "clase": clase.pk,
                f"hora_{clase.pk}-estado": "asistio",
                f"hora_{clase.pk}-docente": reemplazo.pk,
                f"hora_{clase.pk}-horas": "2.00",
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 302)
        registro = ClaseHoraDocente.objects.get(clase=clase)
        self.assertEqual(registro.estado, "asistio")
        self.assertEqual(registro.docente, self.docente)
        self.assertIsNone(registro.docente_reemplazado)

    def test_docente_cannot_open_director_teacher_hours(self):
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("academico:direccion_horas_docente"),
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 403)

    def test_coordinator_cannot_open_director_teacher_hours(self):
        self.client.force_login(self.create_coordinator())

        response = self.client.get(
            reverse("academico:direccion_horas_docente"),
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 403)

    def test_teacher_hours_report_renders_and_exports_excel(self):
        director = self.create_director()
        clase = self.create_class_for_date(timezone.localdate(), time(14, 0), time(16, 0), aula_nombre="Aula reporte")
        ClaseHoraDocente.objects.create(
            clase=clase,
            docente=self.docente,
            estado="asistio",
            horas=Decimal("2.00"),
            registrado_por=director,
            fecha_registro=timezone.now(),
            usuario_updated=director,
        )
        self.client.force_login(director)
        params = {
            "desde": clase.fecha.isoformat(),
            "hasta": clase.fecha.isoformat(),
        }

        response = self.client.get(
            reverse("academico:direccion_horas_docente_reporte"),
            params,
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Reporte horas docente")
        self.assertContains(response, "Docente Prueba")
        self.assertEqual(response.context["stats"]["horas"], Decimal("2.00"))

        export_response = self.client.get(
            reverse("academico:direccion_horas_docente_reporte"),
            {**params, "export": "excel"},
            HTTP_HOST="localhost",
        )

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(export_response.content))
        values = [cell.value for row in workbook.active.iter_rows() for cell in row if cell.value]
        self.assertIn("Docente Prueba", values)
        self.assertIn("Matematicas", values)
        self.assertIn(2, values)

    def test_coordinacion_review_detail_renders_visual_review_panel(self):
        coordinator = self.create_coordinator()
        self.client.force_login(coordinator)

        response = self.client.get(
            reverse("academico:coordinacion_revision_planificacion_detalle", args=[self.revision.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "review-hero-panel")
        self.assertContains(response, "Checklist de revision")
        self.assertContains(response, "review-decision-panel")
        self.assertEqual(response.context["review_total"], 4)

    def test_coordinacion_review_detail_shows_topic_and_subject_progress(self):
        Subtema.objects.create(tema=self.tema, nombre="Resta", orden=2)
        self.revision.tema = self.tema
        self.revision.save(update_fields=["tema"])
        self.revision.sync_subtemas_planificados([self.subtema])
        coordinator = self.create_coordinator()
        self.client.force_login(coordinator)

        response = self.client.get(
            reverse("academico:coordinacion_revision_planificacion_detalle", args=[self.revision.pk]),
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Avance del tema")
        self.assertContains(response, "Avance de la materia")
        self.assertEqual(response.context["topic_progress"]["progress"], 50)
        self.assertEqual(response.context["materia_progress"]["progress"], 50)

    def test_coordinacion_review_detail_approval_requires_all_sections_checked(self):
        coordinator = self.create_coordinator()
        self.client.force_login(coordinator)

        response = self.client.post(
            reverse("academico:coordinacion_revision_planificacion_detalle", args=[self.revision.pk]),
            {"review_action": "aprobar"},
            HTTP_HOST="localhost",
        )
        self.revision.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Marca todos los puntos como correctos")
        self.assertEqual(self.revision.estado_planificacion, "revision")

    def test_coordinacion_review_detail_rejects_one_observed_section(self):
        coordinator = self.create_coordinator()
        self.client.force_login(coordinator)

        response = self.client.post(
            reverse("academico:coordinacion_revision_planificacion_detalle", args=[self.revision.pk]),
            {
                "review_action": "rechazar",
                "notas_revision": "Corregir recursos.",
                "revision_tema_ok": "on",
                "revision_competencias_ok": "on",
                "revision_estrategias_ok": "on",
                "observacion_recursos": "Agregar un recurso verificable.",
            },
            HTTP_HOST="localhost",
        )
        self.revision.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.revision.estado_planificacion, "rechazada")
        self.assertFalse(self.revision.revision_recursos_ok)
        self.assertEqual(
            self.revision.observaciones_revision,
            {"recursos": "Agregar un recurso verificable."},
        )
