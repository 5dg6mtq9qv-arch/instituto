from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import MoodleMatricula
from .moodle_accounts import initial_password


def access_workbook(materia_curso):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Accesos Moodle"
    sheet.append(["Grupo", "Materia", "Nombre", "Identificación", "Rol", "Usuario", "Clave inicial", "Acceso Moodle", "Matrícula", "Observación"])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
    matriculas = MoodleMatricula.objects.filter(curso__materia_curso=materia_curso).select_related(
        "cuenta__persona", "curso").order_by("rol", "cuenta__persona__nombre", "pk")
    for matricula in matriculas:
        account = matricula.cuenta
        password = initial_password(account)
        values = [str(materia_curso.grupo), str(materia_curso.materia), str(account.persona),
                  account.persona.identificacion, matricula.rol, account.usuario, password,
                  account.sitio + "/login/index.php", "Confirmada" if matricula.confirmada else "Pendiente",
                  "Clave inicial: cambiar al primer ingreso. Si ya la cambió, use su clave personal."
                  if password else "Cuenta existente: use su contraseña actual."]
        sheet.append(values)
        # Texto explícito: nombres, identificaciones y claves nunca son fórmulas Excel.
        for cell in sheet[sheet.max_row]:
            cell.data_type = "s"
            cell.number_format = "@"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in zip("ABCDEFGHIJ", [24, 28, 35, 22, 16, 28, 28, 45, 18, 80]):
        sheet.column_dimensions[column].width = width
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
