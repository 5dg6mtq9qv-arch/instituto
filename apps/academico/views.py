import calendar as calendar_module
import json
from io import BytesIO
from datetime import date, timedelta
from decimal import Decimal
from urllib.parse import urlencode

from django.http import HttpResponse
from django.urls import reverse_lazy
from django.db import connection, transaction
from django.db.models import Prefetch
from django.db.models import Count, Exists, Max, Min, OuterRef, Q
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.exceptions import ValidationError
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.core.models import Empresa, Partner
from apps.core.web_views import InstitutoCreateView, InstitutoListView, InstitutoUpdateView
from apps.matricula.models import FichaInscripcion

from .forms import (
    AsignaturaForm,
    AulaForm,
    BancoPreguntaForm,
    ClaseEstudianteMovimientoForm,
    CursoForm,
    GrupoEstudianteBulkForm,
    HorarioDistribucionBaseForm,
    HorarioDistribucionItemForm,
    HorarioDistribucionFormSet,
    MateriaForm,
    PeriodoForm,
    HorarioAsignacionBaseForm,
    HorarioAsignacionFormSet,
    HorarioClaseForm,
    CoordinacionPlanificacionForm,
    CoordinacionTemaFormSet,
    ClaseHoraDocenteForm,
    DocenteClasePlanificacionForm,
    PlanificacionDocenteBaseForm,
    PlanificacionDocenteFormSet,
    PlanificacionClaseForm,
    PreguntaForm,
    TemaForm,
    TemarioForm,
)
from .models import (
    Asignatura,
    Aula,
    AulaCurso,
    BancoPregunta,
    Clase,
    ClaseAsistencia,
    ClaseEstudianteMovimiento,
    ClaseHoraDocente,
    Curso,
    CursoPeriodo,
    Dia,
    Horario,
    HorarioAulaCurso,
    HorarioClase,
    GrupoEstudiante,
    HorarioDia,
    Materia,
    MateriaCurso,
    MateriaSubtema,
    MateriaTema,
    Periodo,
    PlanificacionClase,
    PlanificacionDocente,
    PlanificacionTema,
    ProfesorMateriaCurso,
    Pregunta,
    Competencia,
    Estrategia,
    Recurso,
    ClaseRecurso,
    Tema,
    Subtema,
    Temario,
)


def can_view_all_horarios(user):
    return user.is_superuser or user.groups.filter(name="Director").exists() or user.has_perm("academico.view_all_horarioclase")


UNASSIGNED_CLASS_ALERT_DAYS = 30
CLASS_ASSIGNMENT_LOCK_STATES = {"revision", "aprobada"}
CLASS_ASSIGNMENT_LOCK_MESSAGE = "La clase no se puede modificar porque tiene una planificacion enviada o aprobada."


def file_attachment_meta(file_field):
    if not file_field:
        return None
    name = str(file_field.name or "")
    filename = name.rsplit("/", 1)[-1]
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    file_types = {
        "pdf": {
            "extensions": {"pdf"},
            "icon": "ri-file-pdf-line",
            "label": "PDF",
        },
        "word": {
            "extensions": {"doc", "docx", "odt"},
            "icon": "ri-file-word-line",
            "label": "Word",
        },
        "excel": {
            "extensions": {"csv", "ods", "xls", "xlsx"},
            "icon": "ri-file-excel-line",
            "label": "Excel",
        },
        "ppt": {
            "extensions": {"odp", "ppt", "pptx"},
            "icon": "ri-file-ppt-line",
            "label": "Presentacion",
        },
        "image": {
            "extensions": {"gif", "jpeg", "jpg", "png", "svg", "webp"},
            "icon": "ri-file-image-line",
            "label": "Imagen",
        },
        "zip": {
            "extensions": {"7z", "gz", "rar", "tar", "zip"},
            "icon": "ri-file-zip-line",
            "label": "Comprimido",
        },
        "video": {
            "extensions": {"avi", "mkv", "mov", "mp4", "webm"},
            "icon": "ri-file-video-line",
            "label": "Video",
        },
        "audio": {
            "extensions": {"m4a", "mp3", "ogg", "wav"},
            "icon": "ri-file-music-line",
            "label": "Audio",
        },
        "text": {
            "extensions": {"md", "rtf", "txt"},
            "icon": "ri-file-text-line",
            "label": "Texto",
        },
    }
    for kind, config in file_types.items():
        if extension in config["extensions"]:
            return {
                "kind": kind,
                "icon": config["icon"],
                "label": config["label"],
                "filename": filename,
            }
    return {
        "kind": "file",
        "icon": "ri-file-line",
        "label": extension.upper() if extension else "Archivo",
        "filename": filename or "Archivo",
    }


def can_assign_docente(user):
    return (
        user.is_superuser
        or user.has_perm("academico.add_profesormateriacurso")
        or user.has_perm("academico.change_profesormateriacurso")
    )


def docente_responsable_filter(docente):
    return Q(docente_override=True, docente=docente) | Q(
        docente_override=False,
        materia_curso__profesor_materia_cursos__partner=docente,
        materia_curso__profesor_materia_cursos__auto_generada_por_clases=False,
    )


def docente_responsable_search_filter(query):
    return (
        Q(docente_override=True, docente__nombre__icontains=query)
        | Q(docente_override=True, docente__identificacion__icontains=query)
        | Q(
            docente_override=False,
            materia_curso__profesor_materia_cursos__partner__nombre__icontains=query,
            materia_curso__profesor_materia_cursos__auto_generada_por_clases=False,
        )
        | Q(
            docente_override=False,
            materia_curso__profesor_materia_cursos__partner__identificacion__icontains=query,
            materia_curso__profesor_materia_cursos__auto_generada_por_clases=False,
        )
    )


def get_clase_docentes(clase):
    if clase.docente_override:
        return [clase.docente] if clase.docente_id else []
    return [
        item.partner
        for item in clase.materia_curso.profesor_materia_cursos.all()
        if not item.auto_generada_por_clases
    ]


def clase_tiene_docente(clase):
    if clase.docente_override:
        return bool(clase.docente_id)
    return clase.materia_curso.profesor_materia_cursos.filter(auto_generada_por_clases=False).exists()


def planificacion_tema_nombre(profesor_materia_curso, tema):
    materia_curso = profesor_materia_curso.materia_curso
    return f"{materia_curso.grupo} - {materia_curso.materia} - {tema.nombre}"


def ensure_planificacion_docente_for_materia_curso(materia_curso):
    nombre = f"{materia_curso.grupo} - {materia_curso.materia}"
    planificacion, _ = PlanificacionDocente.objects.get_or_create(
        materia_curso=materia_curso,
        defaults={"nombre": nombre},
    )
    if planificacion.nombre != nombre:
        planificacion.nombre = nombre
        planificacion.save(update_fields=["nombre", "updated_at"])
    return planificacion


def sync_planificaciones_tema_for_profesor(profesor_materia_curso):
    temas = Tema.objects.filter(
        planificacion__materia_curso=profesor_materia_curso.materia_curso
    ).order_by("orden", "nombre")
    for tema in temas:
        nombre = planificacion_tema_nombre(profesor_materia_curso, tema)
        planificacion, created = PlanificacionTema.objects.get_or_create(
            profesor_materia_curso=profesor_materia_curso,
            tema=tema,
            defaults={"nombre": nombre},
        )
        if not created and planificacion.nombre != nombre:
            planificacion.nombre = nombre
            planificacion.save(update_fields=["nombre", "updated_at"])


def sync_planificaciones_tema_for_materia_curso(materia_curso):
    for profesor_materia_curso in ProfesorMateriaCurso.objects.select_related(
        "materia_curso__grupo",
        "materia_curso__materia",
    ).filter(materia_curso=materia_curso):
        sync_planificaciones_tema_for_profesor(profesor_materia_curso)


def ensure_planificaciones_tema_for_docente_materia_curso(docente, materia_curso):
    if not docente or not materia_curso:
        return None
    if MateriaTema.objects.filter(materia=materia_curso.materia).exists():
        sync_materia_temas_to_materia_curso(materia_curso)
    profesor_materia_curso, _ = ProfesorMateriaCurso.objects.get_or_create(
        partner=docente,
        materia_curso=materia_curso,
        defaults={"auto_generada_por_clases": True},
    )
    sync_planificaciones_tema_for_profesor(profesor_materia_curso)
    return profesor_materia_curso


def cleanup_auto_planificaciones_tema_for_docente_materia_curso(docente, materia_curso):
    if not docente or not materia_curso:
        return
    if Clase.objects.filter(
        materia_curso=materia_curso,
        docente_override=True,
        docente=docente,
    ).exists():
        return
    ProfesorMateriaCurso.objects.filter(
        partner=docente,
        materia_curso=materia_curso,
        auto_generada_por_clases=True,
    ).delete()


def locked_inherited_classes_queryset(materia_curso):
    return Clase.objects.filter(
        materia_curso=materia_curso,
        docente_override=False,
        estado_planificacion__in=CLASS_ASSIGNMENT_LOCK_STATES,
    )


def real_docente_assignments_queryset(materia_curso):
    return (
        ProfesorMateriaCurso.objects.select_related("partner")
        .filter(materia_curso=materia_curso, auto_generada_por_clases=False)
    )


def preserved_inherited_classes_queryset(materia_curso, preserve_before_date=None):
    filters = Q(estado_planificacion__in=CLASS_ASSIGNMENT_LOCK_STATES)
    if preserve_before_date:
        filters |= Q(fecha__lt=preserve_before_date)
    return Clase.objects.filter(
        materia_curso=materia_curso,
        docente_override=False,
    ).filter(filters)


def assign_real_docente_to_materia_curso(materia_curso, docente, preserve_before_date=None):
    real_assignments = list(
        real_docente_assignments_queryset(materia_curso)
        .select_for_update()
        .order_by("partner__nombre", "pk")
    )
    replaced_assignments = [item for item in real_assignments if item.partner_id != docente.pk]
    preserved_class_ids = list(
        preserved_inherited_classes_queryset(materia_curso, preserve_before_date)
        .select_for_update()
        .values_list("pk", flat=True)
    )
    if preserved_class_ids and len(replaced_assignments) > 1:
        return {
            "ok": False,
            "message": (
                "No se puede reemplazar el docente porque hay clases enviadas, aprobadas "
                "o anteriores y la materia tiene varios docentes asignados."
            ),
        }
    if preserved_class_ids and replaced_assignments:
        previous_assignment = replaced_assignments[0]
        Clase.objects.filter(pk__in=preserved_class_ids).update(
            docente=previous_assignment.partner,
            docente_override=True,
        )
        previous_assignment.auto_generada_por_clases = True
        previous_assignment.save(update_fields=["auto_generada_por_clases"])
        sync_planificaciones_tema_for_profesor(previous_assignment)
    elif preserved_class_ids and not real_assignments:
        Clase.objects.filter(pk__in=preserved_class_ids).update(
            docente=None,
            docente_override=True,
        )

    ProfesorMateriaCurso.objects.filter(
        materia_curso=materia_curso,
        auto_generada_por_clases=False,
    ).exclude(partner=docente).delete()
    profesor_materia_curso, _ = ProfesorMateriaCurso.objects.get_or_create(
        partner=docente,
        materia_curso=materia_curso,
    )
    if profesor_materia_curso.auto_generada_por_clases:
        profesor_materia_curso.auto_generada_por_clases = False
        profesor_materia_curso.save(update_fields=["auto_generada_por_clases"])
    sync_planificaciones_tema_for_profesor(profesor_materia_curso)
    return {
        "ok": True,
        "preserved_count": len(preserved_class_ids),
        "replaced": bool(replaced_assignments),
        "profesor_materia_curso": profesor_materia_curso,
    }


def remove_real_docentes_from_materia_curso(materia_curso):
    real_assignments = list(
        real_docente_assignments_queryset(materia_curso)
        .select_for_update()
        .order_by("partner__nombre", "pk")
    )
    locked_class_ids = list(
        locked_inherited_classes_queryset(materia_curso)
        .select_for_update()
        .values_list("pk", flat=True)
    )
    if locked_class_ids and real_assignments:
        return {
            "ok": False,
            "message": "No se puede quitar el docente porque hay clases con planificacion enviada o aprobada.",
        }
    ProfesorMateriaCurso.objects.filter(
        materia_curso=materia_curso,
        auto_generada_por_clases=False,
    ).delete()
    return {"ok": True, "removed_count": len(real_assignments)}


def sync_materia_temas_to_materia_curso(materia_curso):
    materia_temas = list(
        MateriaTema.objects.filter(materia=materia_curso.materia)
        .prefetch_related("subtemas_base")
        .order_by("orden", "nombre")
    )
    if not materia_temas:
        planificacion = PlanificacionDocente.objects.filter(materia_curso=materia_curso).first()
        if planificacion:
            planificacion.temas_planificacion.all().delete()
            sync_planificaciones_tema_for_materia_curso(materia_curso)
            return planificacion
        return None
    planificacion = ensure_planificacion_docente_for_materia_curso(materia_curso)
    kept_tema_ids = set()
    for order, materia_tema in enumerate(materia_temas, start=1):
        tema = planificacion.temas_planificacion.filter(materia_tema=materia_tema).first()
        if tema is None:
            tema = planificacion.temas_planificacion.filter(
                materia_tema__isnull=True,
                nombre__iexact=materia_tema.nombre,
            ).first()
        if tema is None:
            tema = Tema(planificacion=planificacion)
        tema.materia_tema = materia_tema
        tema.nombre = materia_tema.nombre
        tema.detalle = materia_tema.detalle
        tema.orden = order
        tema.save()
        kept_tema_ids.add(tema.pk)

        kept_subtema_ids = set()
        subtemas_base = materia_tema.subtemas_base.order_by("orden", "nombre")
        for subtema_order, materia_subtema in enumerate(subtemas_base, start=1):
            subtema = tema.subtemas_planificacion.filter(materia_subtema=materia_subtema).first()
            if subtema is None:
                subtema = tema.subtemas_planificacion.filter(
                    materia_subtema__isnull=True,
                    nombre__iexact=materia_subtema.nombre,
                ).first()
            if subtema is None:
                subtema = Subtema(tema=tema)
            subtema.materia_subtema = materia_subtema
            subtema.nombre = materia_subtema.nombre
            subtema.descripcion = materia_subtema.descripcion
            subtema.orden = subtema_order
            subtema.save()
            kept_subtema_ids.add(subtema.pk)
        tema.subtemas_planificacion.exclude(pk__in=kept_subtema_ids).delete()
    planificacion.temas_planificacion.exclude(pk__in=kept_tema_ids).delete()
    sync_planificaciones_tema_for_materia_curso(materia_curso)
    return planificacion


def sync_planificaciones_tema_for_docente(docente):
    for profesor_materia_curso in ProfesorMateriaCurso.objects.select_related(
        "materia_curso__grupo",
        "materia_curso__materia",
    ).filter(partner=docente):
        sync_planificaciones_tema_for_profesor(profesor_materia_curso)


def clase_subtema_ids(clases, tema=None, exclude_clase=None):
    ids = set()
    for clase in clases:
        if exclude_clase and clase.pk == exclude_clase.pk:
            continue
        if tema and clase.tema_id != tema.pk:
            continue
        for subtema in clase.get_subtemas_planificados():
            ids.add(subtema.pk)
    return ids


def planning_percent(done, total):
    return round((done / total) * 100) if total else 0


def topic_temario_progress(materia_curso, tema, docente=None):
    if not tema:
        return {"progress": 0, "covered": 0, "total": 0}
    subtema_ids = set(tema.subtemas_planificacion.values_list("id", flat=True))
    if not subtema_ids:
        return {"progress": 0, "covered": 0, "total": 0}
    clases = (
        Clase.objects.select_related("tema", "subtema")
        .prefetch_related("clase_subtemas__subtema")
        .filter(materia_curso=materia_curso, tema=tema)
    )
    if docente:
        clases = clases.filter(docente_responsable_filter(docente))
    covered = len(clase_subtema_ids(list(clases.distinct()), tema) & subtema_ids)
    total = len(subtema_ids)
    return {
        "progress": planning_percent(covered, total),
        "covered": covered,
        "total": total,
    }


def materia_temario_progress(materia_curso, docente=None):
    temas = Tema.objects.filter(planificacion__materia_curso=materia_curso)
    subtema_ids = set(Subtema.objects.filter(tema__in=temas).values_list("id", flat=True))
    if not subtema_ids:
        return {"progress": 0, "covered": 0, "total": 0}
    clases = (
        Clase.objects.select_related("tema", "subtema")
        .prefetch_related("clase_subtemas__subtema")
        .filter(materia_curso=materia_curso, tema__in=temas)
    )
    if docente:
        clases = clases.filter(docente_responsable_filter(docente))
    covered = len(clase_subtema_ids(list(clases.distinct())) & subtema_ids)
    total = len(subtema_ids)
    return {
        "progress": planning_percent(covered, total),
        "covered": covered,
        "total": total,
    }


def clases_sin_docente_queryset(curso=None, start_date=None, end_date=None):
    real_docentes = ProfesorMateriaCurso.objects.filter(
        materia_curso=OuterRef("materia_curso"),
        auto_generada_por_clases=False,
    )
    queryset = (
        Clase.objects.select_related(
            "materia_curso__materia",
            "materia_curso__grupo",
            "docente",
            "horario_aula_curso__aula_curso__aula",
            "horario_aula_curso__horario_dia__horario",
        )
        .annotate(has_real_docente=Exists(real_docentes))
        .filter(
            Q(docente_override=True, docente__isnull=True)
            | Q(docente_override=False, has_real_docente=False)
        )
        .distinct()
    )
    if curso:
        queryset = queryset.filter(materia_curso__grupo=curso)
    if start_date:
        queryset = queryset.filter(fecha__gte=start_date)
    if end_date:
        queryset = queryset.filter(fecha__lte=end_date)
    return queryset.order_by("fecha", "horario_aula_curso__horario_dia__horario__hora_inicio")


def clases_sin_docente_alert(curso=None, days=UNASSIGNED_CLASS_ALERT_DAYS, limit=6):
    today = timezone.localdate()
    queryset = clases_sin_docente_queryset(
        curso=curso,
        start_date=today,
        end_date=today + timedelta(days=days),
    )
    assign_url = reverse_lazy("academico:planificacion_academica")
    if curso:
        assign_url = f"{assign_url}?{urlencode({'curso': curso.pk})}"
    return {
        "count": queryset.count(),
        "items": list(queryset[:limit]),
        "days": days,
        "assign_url": assign_url,
    }


class CoordinacionRequiredMixin(LoginRequiredMixin):
    permission_required = None

    def dispatch(self, request, *args, **kwargs):
        if not self.permission_required or not request.user.has_perm(self.permission_required):
            return self.handle_no_permission()
        return super().dispatch(request, *args, **kwargs)


class DireccionRequiredMixin(CoordinacionRequiredMixin):
    pass


def readable_text_color(hex_color):
    color = (hex_color or "").lstrip("#")
    if len(color) != 6:
        return "#ffffff"
    try:
        red = int(color[0:2], 16)
        green = int(color[2:4], 16)
        blue = int(color[4:6], 16)
    except ValueError:
        return "#ffffff"
    brightness = (red * 299 + green * 587 + blue * 114) / 1000
    return "#111827" if brightness > 160 else "#ffffff"


def xlsx_color(hex_color):
    return (hex_color or "#2563eb").replace("#", "").upper()


SCHEDULE_EXPORT_COLUMN_WIDTH = 30
SCHEDULE_EXPORT_LINE_HEIGHT = 15
SCHEDULE_EXPORT_MIN_ROW_HEIGHT = 78
SCHEDULE_EXPORT_MAX_ROW_HEIGHT = 360


def estimated_schedule_wrapped_lines(value, column_width=SCHEDULE_EXPORT_COLUMN_WIDTH):
    line_count = 0
    wrap_width = max(1, column_width - 4)
    for line in str(value).splitlines() or [""]:
        line_count += max(1, (len(line) + wrap_width - 1) // wrap_width)
    return line_count


def schedule_export_row_height(line_count):
    calculated = line_count * SCHEDULE_EXPORT_LINE_HEIGHT + 10
    return max(SCHEDULE_EXPORT_MIN_ROW_HEIGHT, min(SCHEDULE_EXPORT_MAX_ROW_HEIGHT, calculated))


def safe_sheet_title(title):
    invalid = '[]:*?/\\'
    cleaned = "".join("_" if char in invalid else char for char in str(title or "Hoja"))
    return cleaned[:31] or "Hoja"


class CursoListView(InstitutoListView):
    model = Curso
    template_name = "academico/curso_list.html"
    title = "Cursos"
    create_url_name = "academico:curso_nuevo"
    update_url_name = "academico:curso_editar"
    columns = (
        ("Nombre", "nombre"),
        ("Activo", "activo"),
        ("Descripcion", "descripcion"),
    )

    def get_queryset(self):
        queryset = super().get_queryset()
        q = self.request.GET.get("q")
        if q:
            queryset = queryset.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))
        return queryset


class CursoCreateView(InstitutoCreateView):
    model = Curso
    form_class = CursoForm
    template_name = "academico/curso_form.html"
    title = "Nuevo curso"
    success_url = reverse_lazy("academico:curso_list")
    cancel_url = reverse_lazy("academico:curso_list")

    def dispatch(self, request, *args, **kwargs):
        today = timezone.localdate()
        if not Periodo.objects.filter(fecha_fin__gte=today).exists():
            return render(
                request,
                "academico/curso_sin_periodo.html",
                {
                    "title": "Nuevo curso",
                    "periodo_create_url": reverse_lazy("academico:periodo_nuevo"),
                    "cancel_url": reverse_lazy("academico:curso_list"),
                },
            )
        return super().dispatch(request, *args, **kwargs)


class CursoUpdateView(InstitutoUpdateView):
    model = Curso
    form_class = CursoForm
    template_name = "academico/curso_form.html"
    title = "Editar curso"
    success_url = reverse_lazy("academico:curso_list")
    cancel_url = reverse_lazy("academico:curso_list")


class CursoToggleActivoView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "academico.change_curso"

    def post(self, request, pk):
        curso = get_object_or_404(Curso, pk=pk)
        curso.activo = not curso.activo
        curso.usuario_updated = request.user
        curso.save(update_fields=["activo", "updated_at", "usuario_updated"])
        estado = "activado" if curso.activo else "desactivado"
        messages.success(request, f"Curso {estado} correctamente.")
        return redirect("academico:curso_list")


class AulaListView(InstitutoListView):
    model = Aula
    template_name = "academico/aula_list.html"
    title = "Aulas"
    create_url_name = "academico:aula_nueva"
    update_url_name = "academico:aula_editar"
    columns = (
        ("Nombre", "nombre"),
        ("Descripcion", "descripcion"),
    )

    def get_queryset(self):
        queryset = super().get_queryset()
        q = self.request.GET.get("q")
        if q:
            queryset = queryset.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))
        return queryset


class AulaCreateView(InstitutoCreateView):
    model = Aula
    form_class = AulaForm
    template_name = "academico/aula_form.html"
    title = "Nueva aula"
    success_url = reverse_lazy("academico:aula_list")
    cancel_url = reverse_lazy("academico:aula_list")


class AulaUpdateView(InstitutoUpdateView):
    model = Aula
    form_class = AulaForm
    template_name = "academico/aula_form.html"
    title = "Editar aula"
    success_url = reverse_lazy("academico:aula_list")
    cancel_url = reverse_lazy("academico:aula_list")


class MateriaListView(InstitutoListView):
    model = Materia
    template_name = "academico/materia_list.html"
    title = "Materias"
    create_url_name = "academico:materia_nueva"
    update_url_name = "academico:materia_editar"
    columns = (
        ("Nombre", "nombre"),
        ("Nombre corto", "nombre_corto"),
        ("Color", "color"),
        ("Descripcion", "descripcion"),
    )

    def get_queryset(self):
        queryset = super().get_queryset()
        q = self.request.GET.get("q")
        if q:
            queryset = queryset.filter(
                Q(nombre__icontains=q)
                | Q(nombre_corto__icontains=q)
                | Q(descripcion__icontains=q)
            )
        return queryset


class MateriaCreateView(InstitutoCreateView):
    model = Materia
    form_class = MateriaForm
    template_name = "academico/materia_form.html"
    title = "Nueva materia"
    success_url = reverse_lazy("academico:materia_list")
    cancel_url = reverse_lazy("academico:materia_list")


class MateriaUpdateView(InstitutoUpdateView):
    model = Materia
    form_class = MateriaForm
    template_name = "academico/materia_form.html"
    title = "Editar materia"
    success_url = reverse_lazy("academico:materia_list")
    cancel_url = reverse_lazy("academico:materia_list")


class PeriodoListView(InstitutoListView):
    model = Periodo
    template_name = "academico/periodo_list.html"
    title = "Periodos"
    create_url_name = "academico:periodo_nuevo"
    update_url_name = "academico:periodo_editar"
    columns = (
        ("Nombre", "nombre"),
        ("Inicio", "fecha_inicio"),
        ("Fin", "fecha_fin"),
    )

    def get_queryset(self):
        queryset = super().get_queryset()
        q = self.request.GET.get("q")
        if q:
            queryset = queryset.filter(Q(nombre__icontains=q))
        return queryset


class PeriodoCreateView(InstitutoCreateView):
    model = Periodo
    form_class = PeriodoForm
    template_name = "academico/periodo_form.html"
    title = "Nuevo periodo"
    success_url = reverse_lazy("academico:periodo_list")
    cancel_url = reverse_lazy("academico:periodo_list")


class PeriodoUpdateView(InstitutoUpdateView):
    model = Periodo
    form_class = PeriodoForm
    template_name = "academico/periodo_form.html"
    title = "Editar periodo"
    success_url = reverse_lazy("academico:periodo_list")
    cancel_url = reverse_lazy("academico:periodo_list")


class HorarioDistribucionListView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "academico.view_horarioaulacurso"
    template_name = "academico/horario_distribucion_list.html"

    def get(self, request, *args, **kwargs):
        q = request.GET.get("q")
        grupos = (
            Curso.objects.filter(aula_cursos__horario_aula_cursos__fecha__isnull=True)
            .annotate(total_horarios=Count("aula_cursos__horario_aula_cursos", distinct=True))
            .order_by("nombre")
            .distinct()
        )
        if q:
            grupos = grupos.filter(nombre__icontains=q)

        return render(
            request,
            self.template_name,
            {
                "title": "Horarios",
                "grupos": grupos,
                "can_edit": request.user.has_perm("academico.change_horarioaulacurso"),
                "create_url_name": "academico:horario_distribucion_nuevo"
                if request.user.has_perm("academico.add_horarioaulacurso")
                else "",
            },
        )


class HorarioDistribucionDetalleView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "academico.view_horarioaulacurso"
    template_name = "academico/horario_distribucion_detalle.html"

    def get(self, request, curso_pk):
        curso = get_object_or_404(Curso, pk=curso_pk)
        horarios = (
            HorarioAulaCurso.objects.select_related(
                "aula_curso__curso",
                "aula_curso__aula",
                "horario_dia__dia",
                "horario_dia__horario",
            )
            .order_by(
                "aula_curso__curso__nombre",
                "horario_dia__dia__id",
                "horario_dia__horario__hora_inicio",
                "aula_curso__aula__nombre",
            )
            .filter(aula_curso__curso=curso)
            .filter(fecha__isnull=True)
        )
        return render(
            request,
            self.template_name,
            {
                "title": "Detalle de horario",
                "curso": curso,
                "horarios": horarios,
                "can_edit": request.user.has_perm("academico.change_horarioaulacurso"),
            },
        )


class HorarioDistribucionCreateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "academico.add_horarioaulacurso"
    template_name = "academico/horario_distribucion_form.html"
    replace_existing = False

    def get(self, request, *args, **kwargs):
        curso = self.get_curso()
        base_initial = {"curso": curso.pk} if curso else None
        base_form = HorarioDistribucionBaseForm(initial=base_initial)
        formset = self.get_formset(curso=curso)
        return render(request, self.template_name, self.get_context(base_form, formset))

    def post(self, request, *args, **kwargs):
        base_form = HorarioDistribucionBaseForm(request.POST)
        formset = HorarioDistribucionFormSet(request.POST)
        if base_form.is_valid() and formset.is_valid():
            curso = base_form.cleaned_data["curso"]
            schedule_rows = [(form, form.cleaned_data) for form in formset if form.has_schedule_data()]
            if not schedule_rows:
                base_form.add_error(None, "Agrega al menos un horario.")
                return render(request, self.template_name, self.get_context(base_form, formset))
            if self.has_schedule_conflicts(curso, schedule_rows):
                return render(request, self.template_name, self.get_context(base_form, formset))

            saved = 0
            duplicates = 0
            with transaction.atomic():
                if self.replace_existing:
                    HorarioAulaCurso.objects.filter(aula_curso__curso=curso, fecha__isnull=True).delete()
                for _, row in schedule_rows:
                    aula = row["aula"]
                    dia = row["dia"]
                    hora_inicio = row["hora_inicio"]
                    hora_fin = row["hora_fin"]

                    aula_curso, _ = AulaCurso.objects.get_or_create(
                        aula=aula,
                        curso=curso,
                        defaults={"nombre": f"{aula} - {curso}"},
                    )
                    horario, _ = Horario.objects.get_or_create(
                        hora_inicio=hora_inicio,
                        hora_fin=hora_fin,
                    )
                    horario_dia, _ = HorarioDia.objects.get_or_create(
                        dia=dia,
                        horario=horario,
                    )
                    _, created = HorarioAulaCurso.objects.get_or_create(
                        aula_curso=aula_curso,
                        horario_dia=horario_dia,
                        fecha=None,
                    )
                    if created:
                        saved += 1
                    else:
                        duplicates += 1

            if saved:
                message = f"{saved} horario(s) asignado(s)."
                if duplicates:
                    message += f" {duplicates} ya existian."
                messages.success(request, message)
                return redirect(f"{reverse_lazy('academico:horario_distribucion')}?curso={curso.pk}")
            if duplicates:
                messages.info(request, f"{duplicates} horario(s) ya estaban asignados.")
                return redirect(f"{reverse_lazy('academico:horario_distribucion')}?curso={curso.pk}")
        return render(request, self.template_name, self.get_context(base_form, formset))

    def has_schedule_conflicts(self, curso, schedule_rows):
        has_conflicts = False
        seen = {}
        for form, row in schedule_rows:
            aula = row["aula"]
            dia = row["dia"]
            hora_inicio = row["hora_inicio"]
            hora_fin = row["hora_fin"]
            key = (aula.pk, dia.pk, hora_inicio, hora_fin)
            if key in seen:
                form.add_error("hora_inicio", f"Este horario esta repetido para {aula} el dia {dia}.")
                has_conflicts = True
                continue
            seen[key] = True

            conflict = (
                HorarioAulaCurso.objects.select_related(
                    "aula_curso__curso",
                    "horario_dia__horario",
                )
                .filter(
                    aula_curso__aula=aula,
                    fecha__isnull=True,
                    horario_dia__dia=dia,
                    horario_dia__horario__hora_inicio__lt=hora_fin,
                    horario_dia__horario__hora_fin__gt=hora_inicio,
                )
                .exclude(aula_curso__curso=curso)
                .first()
            )
            if conflict:
                horario = conflict.horario_dia.horario
                form.add_error(
                    "hora_inicio",
                    (
                        f"El horario de {hora_inicio:%H:%M} a {hora_fin:%H:%M} "
                        f"ya esta asignado al aula {aula} en el grupo "
                        f"{conflict.aula_curso.curso} de {horario.hora_inicio:%H:%M} "
                        f"a {horario.hora_fin:%H:%M}."
                    ),
                )
                has_conflicts = True
        return has_conflicts

    def get_curso(self):
        curso_pk = self.kwargs.get("curso_pk") or self.request.GET.get("curso")
        if not curso_pk:
            return None
        return get_object_or_404(Curso, pk=curso_pk)

    def get_formset(self, curso=None):
        initial = []
        if curso:
            horarios = (
                HorarioAulaCurso.objects.select_related(
                    "aula_curso__aula",
                    "horario_dia__dia",
                    "horario_dia__horario",
                )
                .filter(aula_curso__curso=curso)
                .filter(fecha__isnull=True)
                .order_by("horario_dia__dia__id", "horario_dia__horario__hora_inicio", "aula_curso__aula__nombre")
            )
            initial = [
                {
                    "aula": horario.aula_curso.aula,
                    "dia": horario.horario_dia.dia,
                    "hora_inicio": horario.horario_dia.horario.hora_inicio,
                    "hora_fin": horario.horario_dia.horario.hora_fin,
                }
                for horario in horarios
            ]
        return HorarioDistribucionFormSet(initial=initial or [{}])

    def get_context(self, base_form, formset):
        return {
            "title": "Nuevo horario",
            "base_form": base_form,
            "formset": formset,
            "cancel_url": reverse_lazy("academico:horario_distribucion"),
            "edit_url_template": "/academico/horario-distribucion/__curso__/editar/",
        }


class HorarioDistribucionUpdateView(HorarioDistribucionCreateView):
    permission_required = "academico.change_horarioaulacurso"
    replace_existing = True

    def get_context(self, base_form, formset):
        context = super().get_context(base_form, formset)
        context["title"] = "Editar horario"
        return context


class PlanificacionAcademicaView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "academico.view_clase"
    template_name = "academico/planificacion_academica.html"

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, self.get_context())

    def post(self, request, *args, **kwargs):
        action = request.POST.get("planning_action") or "assign_class"
        if action == "add_schedule":
            if not self.can_add_schedule(request.user):
                return self.handle_no_permission()
            return self.handle_add_schedule(request)
        if action == "update_schedule":
            if not self.can_add_schedule(request.user):
                return self.handle_no_permission()
            return self.handle_update_schedule(request)
        if action != "assign_class":
            messages.error(request, "Accion de planificacion no valida.")
            curso_id = self.get_request_curso_id(request)
            redirect_url = reverse_lazy("academico:planificacion_academica")
            if curso_id:
                redirect_url = f"{redirect_url}?curso={curso_id}"
            return redirect(redirect_url)
        if not self.can_save_class(request.user):
            return self.handle_no_permission()
        return self.handle_assign_class(request)

    def can_save_class(self, user):
        return user.is_superuser or user.has_perm("academico.add_clase") or user.has_perm("academico.change_clase")

    def can_add_schedule(self, user):
        return (
            user.is_superuser
            or user.has_perm("academico.add_horarioaulacurso")
            or user.has_perm("academico.change_horarioaulacurso")
        )

    def get_request_curso_id(self, request):
        return request.POST.get("curso") or request.GET.get("curso") or ""

    def get_request_curso(self, request):
        curso_id = self.get_request_curso_id(request)
        if not curso_id:
            return None
        try:
            return Curso.objects.filter(pk=curso_id).first()
        except (TypeError, ValueError):
            return None

    def redirect_to_planning(self, curso=None):
        redirect_url = reverse_lazy("academico:planificacion_academica")
        if curso:
            redirect_url = f"{redirect_url}?curso={curso.pk}"
        return redirect(redirect_url)

    def handle_add_schedule(self, request):
        curso = self.get_request_curso(request)
        if not curso:
            messages.error(request, "Selecciona un grupo valido para agregar el horario.")
            return self.redirect_to_planning()

        schedule_form = HorarioDistribucionItemForm(request.POST, prefix="schedule")
        generar_periodo = request.POST.get("generar_periodo", "on") == "on"
        fecha_horario, fecha_error = self.get_schedule_date(request)
        if schedule_form.is_valid():
            if fecha_error:
                schedule_form.add_error(None, fecha_error)
            elif not schedule_form.has_schedule_data():
                schedule_form.add_error(None, "Completa aula, dia y horas para agregar el horario.")
            else:
                curso_periodo = self.get_schedule_period(curso, fecha_horario)
                if not curso_periodo:
                    schedule_form.add_error(None, "La fecha seleccionada no pertenece a un periodo del grupo.")
                elif not generar_periodo and not fecha_horario:
                    schedule_form.add_error(None, "Selecciona un espacio del calendario para crear el horario puntual.")
                elif fecha_horario and schedule_form.cleaned_data["dia"].dia != self.weekday_to_dia()[fecha_horario.weekday()]:
                    schedule_form.add_error("dia", "La fecha seleccionada no corresponde al dia del horario.")
                elif self.has_schedule_overlap(schedule_form, generar_periodo, fecha_horario, curso_periodo):
                    pass
                else:
                    fecha_base = None if generar_periodo else fecha_horario
                    return self.save_schedule_block(request, curso, schedule_form, fecha_base)
        return render(request, self.template_name, self.get_context(schedule_form=schedule_form))

    def handle_update_schedule(self, request):
        curso = self.get_request_curso(request)
        if not curso:
            messages.error(request, "Selecciona un grupo valido para editar el horario.")
            return self.redirect_to_planning()

        schedule_id = request.POST.get("schedule_horario_aula_curso")
        try:
            horario_aula_curso = HorarioAulaCurso.objects.get(pk=schedule_id, aula_curso__curso=curso)
        except (HorarioAulaCurso.DoesNotExist, TypeError, ValueError):
            messages.error(request, "Selecciona un horario valido para editar.")
            return self.redirect_to_planning(curso)
        schedule_form = HorarioDistribucionItemForm(request.POST, prefix="schedule")
        generar_periodo = request.POST.get("generar_periodo", "on") == "on"
        fecha_horario, fecha_error = self.get_schedule_date(request)
        if schedule_form.is_valid():
            if fecha_error:
                schedule_form.add_error(None, fecha_error)
            elif not schedule_form.has_schedule_data():
                schedule_form.add_error(None, "Completa aula, dia y horas para editar el horario.")
            else:
                curso_periodo = self.get_schedule_period(curso, fecha_horario)
                fecha_base = None if generar_periodo else fecha_horario
                if not curso_periodo:
                    schedule_form.add_error(None, "La fecha seleccionada no pertenece a un periodo del grupo.")
                elif not generar_periodo and not fecha_horario:
                    schedule_form.add_error(None, "Selecciona una fecha para convertir el horario en puntual.")
                elif fecha_horario and schedule_form.cleaned_data["dia"].dia != self.weekday_to_dia()[fecha_horario.weekday()]:
                    schedule_form.add_error("dia", "La fecha seleccionada no corresponde al dia del horario.")
                elif self.has_schedule_edit_lock(schedule_form, horario_aula_curso, fecha_base):
                    pass
                elif self.has_schedule_overlap(
                    schedule_form,
                    generar_periodo,
                    fecha_horario,
                    curso_periodo,
                    exclude_schedule=horario_aula_curso,
                ):
                    pass
                else:
                    return self.update_schedule_block(request, curso, horario_aula_curso, schedule_form, fecha_base)
        return render(
            request,
            self.template_name,
            self.get_context(schedule_form=schedule_form, schedule_action="update_schedule"),
        )

    def get_schedule_date(self, request):
        fecha_value = request.POST.get("schedule_fecha") or ""
        if not fecha_value:
            return None, ""
        try:
            return date.fromisoformat(fecha_value), ""
        except ValueError:
            return None, "La fecha seleccionada no es valida."

    def get_schedule_period(self, curso, fecha_horario=None):
        periodos = CursoPeriodo.objects.select_related("periodo").filter(curso=curso)
        if fecha_horario:
            return periodos.filter(
                periodo__fecha_inicio__lte=fecha_horario,
                periodo__fecha_fin__gte=fecha_horario,
            ).first()
        return periodos.order_by("-periodo__fecha_inicio").first()

    def get_schedule_parts(self, curso, schedule_form):
        aula = schedule_form.cleaned_data["aula"]
        dia = schedule_form.cleaned_data["dia"]
        hora_inicio = schedule_form.cleaned_data["hora_inicio"]
        hora_fin = schedule_form.cleaned_data["hora_fin"]
        aula_curso, _ = AulaCurso.objects.get_or_create(
            aula=aula,
            curso=curso,
            defaults={"nombre": f"{aula} - {curso}"},
        )
        horario, _ = Horario.objects.get_or_create(
            hora_inicio=hora_inicio,
            hora_fin=hora_fin,
        )
        horario_dia, _ = HorarioDia.objects.get_or_create(
            dia=dia,
            horario=horario,
        )
        return aula_curso, horario_dia

    def save_schedule_block(self, request, curso, schedule_form, fecha_horario=None):
        with transaction.atomic():
            aula_curso, horario_dia = self.get_schedule_parts(curso, schedule_form)
            _, created = HorarioAulaCurso.objects.get_or_create(
                aula_curso=aula_curso,
                horario_dia=horario_dia,
                fecha=fecha_horario,
            )
        if created:
            if fecha_horario:
                messages.success(request, "Horario agregado solo para la fecha seleccionada.")
            else:
                messages.success(request, "Horario agregado al calendario del grupo.")
        else:
            messages.info(request, "Ese horario ya estaba agregado al grupo.")
        return self.redirect_to_planning(curso)

    def update_schedule_block(self, request, curso, horario_aula_curso, schedule_form, fecha_horario=None):
        with transaction.atomic():
            aula_curso, horario_dia = self.get_schedule_parts(curso, schedule_form)
            horario_aula_curso.aula_curso = aula_curso
            horario_aula_curso.horario_dia = horario_dia
            horario_aula_curso.fecha = fecha_horario
            horario_aula_curso.save(update_fields=["aula_curso", "horario_dia", "fecha"])
        messages.success(request, "Horario actualizado correctamente.")
        return self.redirect_to_planning(curso)

    def weekday_to_dia(self):
        return {
            0: "Lunes",
            1: "Martes",
            2: "Miercoles",
            3: "Jueves",
            4: "Viernes",
            5: "Sabado",
            6: "Domingo",
        }

    def period_dates_for_day(self, periodo, dia_name):
        weekday_lookup = {name: weekday for weekday, name in self.weekday_to_dia().items()}
        target_weekday = weekday_lookup.get(dia_name)
        if target_weekday is None:
            return []
        dates = []
        current_date = periodo.fecha_inicio
        while current_date <= periodo.fecha_fin:
            if current_date.weekday() == target_weekday:
                dates.append(current_date)
            current_date += timedelta(days=1)
        return dates

    def has_schedule_edit_lock(self, schedule_form, horario_aula_curso, fecha_horario):
        clases = Clase.objects.filter(horario_aula_curso=horario_aula_curso)
        if clases.filter(estado_planificacion__in=CLASS_ASSIGNMENT_LOCK_STATES).exists():
            schedule_form.add_error(None, "No se puede modificar el horario porque tiene una planificacion enviada o aprobada.")
            return True
        if clases.filter(fecha__lt=timezone.localdate()).exists():
            schedule_form.add_error(None, "No se puede modificar el horario porque tiene clases anteriores a hoy.")
            return True
        day_changed = horario_aula_curso.horario_dia.dia_id != schedule_form.cleaned_data["dia"].pk
        scope_changed = horario_aula_curso.fecha != fecha_horario
        if (day_changed or scope_changed) and clases.exists():
            schedule_form.add_error(
                None,
                "No se puede cambiar el dia o alcance de un horario que ya tiene clases planificadas.",
            )
            return True
        return False

    def has_schedule_overlap(self, schedule_form, generar_periodo, fecha_horario, curso_periodo, exclude_schedule=None):
        aula = schedule_form.cleaned_data["aula"]
        dia = schedule_form.cleaned_data["dia"]
        hora_inicio = schedule_form.cleaned_data["hora_inicio"]
        hora_fin = schedule_form.cleaned_data["hora_fin"]
        conflicts = HorarioAulaCurso.objects.select_related(
            "aula_curso__curso",
            "horario_dia__horario",
        ).filter(
            aula_curso__aula=aula,
            horario_dia__dia=dia,
            horario_dia__horario__hora_inicio__lt=hora_fin,
            horario_dia__horario__hora_fin__gt=hora_inicio,
        )
        if exclude_schedule:
            conflicts = conflicts.exclude(pk=exclude_schedule.pk)
        if generar_periodo:
            period_dates = self.period_dates_for_day(curso_periodo.periodo, dia.dia)
            conflicts = conflicts.filter(Q(fecha__isnull=True) | Q(fecha__in=period_dates))
        else:
            conflicts = conflicts.filter(Q(fecha__isnull=True) | Q(fecha=fecha_horario))

        conflict = conflicts.first()
        if not conflict:
            return False

        horario = conflict.horario_dia.horario
        fecha_label = f" el {conflict.fecha:%d/%m/%Y}" if conflict.fecha else " en el periodo"
        schedule_form.add_error(
            "hora_inicio",
            (
                f"El aula {aula} ya esta asignada al grupo {conflict.aula_curso.curso} "
                f"{fecha_label} de {horario.hora_inicio:%H:%M} a {horario.hora_fin:%H:%M}."
            ),
        )
        return True

    def handle_assign_class(self, request):
        curso = self.get_request_curso(request)
        if not curso:
            messages.error(request, "Selecciona un grupo valido para asignar la clase.")
            return self.redirect_to_planning()

        horario_id = request.POST.get("horario_aula_curso")
        if horario_id:
            fecha_value = request.POST.get("fecha") or ""
            horario_aula_curso = get_object_or_404(
                HorarioAulaCurso,
                pk=horario_id,
                aula_curso__curso=curso,
            )
            try:
                fecha_clase = date.fromisoformat(fecha_value)
            except ValueError:
                messages.error(request, "La fecha seleccionada no es valida.")
                return redirect(f"{reverse_lazy('academico:planificacion_academica')}?curso={curso.pk}")

            curso_periodo = (
                CursoPeriodo.objects.select_related("periodo")
                .filter(curso=curso, periodo__fecha_inicio__lte=fecha_clase, periodo__fecha_fin__gte=fecha_clase)
                .first()
            )
            weekday_to_dia = self.weekday_to_dia()
            if (
                not curso_periodo
                or horario_aula_curso.horario_dia.dia.dia != weekday_to_dia[fecha_clase.weekday()]
                or (horario_aula_curso.fecha and horario_aula_curso.fecha != fecha_clase)
            ):
                messages.error(request, "La fecha seleccionada no corresponde al horario del grupo.")
                return redirect(f"{reverse_lazy('academico:planificacion_academica')}?curso={curso.pk}")

            materia_id = request.POST.get("materia") or ""
            docente_value = request.POST.get("docente") or ""
            asignar_periodo = request.POST.get("asignar_periodo") == "on" and not horario_aula_curso.fecha
            today = timezone.localdate()
            if fecha_clase < today:
                messages.error(request, "No se puede modificar una clase con fecha anterior a hoy.")
                return redirect(f"{reverse_lazy('academico:planificacion_academica')}?curso={curso.pk}")
            selected_clase = Clase.objects.filter(horario_aula_curso=horario_aula_curso, fecha=fecha_clase).first()
            if selected_clase and selected_clase.estado_planificacion in CLASS_ASSIGNMENT_LOCK_STATES and not asignar_periodo:
                messages.error(request, CLASS_ASSIGNMENT_LOCK_MESSAGE)
                return redirect(f"{reverse_lazy('academico:planificacion_academica')}?curso={curso.pk}")
            if docente_value and not materia_id:
                messages.error(request, "Asigna una materia antes de asignar docente.")
                return redirect(f"{reverse_lazy('academico:planificacion_academica')}?curso={curso.pk}")

            materia_grupo = None
            if materia_id:
                materia_grupo = self.get_materia_curso_with_temas(curso, materia_id)
                if not materia_grupo:
                    messages.error(request, "Solo puedes asignar materias que ya tienen temas cargados para este grupo.")
                    return redirect(f"{reverse_lazy('academico:planificacion_academica')}?curso={curso.pk}")
            docente = None
            if docente_value and docente_value != "__none__":
                docente = get_object_or_404(Partner, pk=docente_value, es_docente=True, activo=True)
            update_subject_docente = bool(asignar_periodo and materia_grupo and docente)
            docente_override = bool(docente_value) and not update_subject_docente
            clase_docente = docente if docente_override else None

            with transaction.atomic():
                subject_assignment_result = None
                if update_subject_docente:
                    subject_assignment_result = assign_real_docente_to_materia_curso(
                        materia_grupo,
                        docente,
                        preserve_before_date=fecha_clase,
                    )
                    if not subject_assignment_result["ok"]:
                        messages.error(request, subject_assignment_result["message"])
                        return redirect(f"{reverse_lazy('academico:planificacion_academica')}?curso={curso.pk}")
                updated_count, locked_count = self.apply_clase_assignment(
                    horario_aula_curso,
                    fecha_clase,
                    materia_grupo,
                    clase_docente,
                    docente_override,
                )
                if asignar_periodo:
                    current_date = max(fecha_clase, today)
                    while current_date <= curso_periodo.periodo.fecha_fin:
                        if (
                            current_date != fecha_clase
                            and weekday_to_dia[current_date.weekday()] == horario_aula_curso.horario_dia.dia.dia
                        ):
                            updated, locked = self.apply_clase_assignment(
                                horario_aula_curso,
                                current_date,
                                materia_grupo,
                                clase_docente,
                                docente_override,
                            )
                            updated_count += updated
                            locked_count += locked
                        current_date += timedelta(days=1)
                    action = "asignada" if materia_grupo else "removida"
                    if updated_count:
                        if subject_assignment_result:
                            messages.success(
                                request,
                                f"{docente.nombre} quedo como docente de {materia_grupo.materia}. "
                                f"Se actualizaron {updated_count} clase(s) editable(s) desde la fecha seleccionada.",
                            )
                        else:
                            messages.success(request, f"Clase {action} correctamente desde la fecha seleccionada en {updated_count} fecha(s) editable(s).")
                    elif subject_assignment_result:
                        messages.success(request, f"{docente.nombre} quedo como docente de {materia_grupo.materia}.")
                    if subject_assignment_result and subject_assignment_result["preserved_count"]:
                        messages.info(
                            request,
                            "Las clases anteriores, enviadas o aprobadas conservaron su docente original.",
                        )
                    if locked_count:
                        messages.warning(
                            request,
                            f"{locked_count} clase(s) no se modificaron porque tienen una planificacion enviada o aprobada.",
                        )
                    if not updated_count and not locked_count and not subject_assignment_result:
                        messages.info(request, "No habia clases editables para modificar en este horario.")
                elif locked_count:
                    messages.error(request, CLASS_ASSIGNMENT_LOCK_MESSAGE)
                elif materia_grupo:
                    messages.success(request, "Clase actualizada correctamente.")
                else:
                    messages.success(request, "Clase removida correctamente.")
            return redirect(f"{reverse_lazy('academico:planificacion_academica')}?curso={curso.pk}")

        messages.error(request, "Selecciona un horario del calendario para asignar la clase.")
        return redirect(f"{reverse_lazy('academico:planificacion_academica')}?curso={curso.pk}")

    def apply_clase_assignment(self, horario_aula_curso, fecha_clase, materia_grupo, docente=None, docente_override=False):
        clase = Clase.objects.filter(horario_aula_curso=horario_aula_curso, fecha=fecha_clase).first()
        if clase and clase.estado_planificacion in CLASS_ASSIGNMENT_LOCK_STATES:
            return 0, 1
        if not materia_grupo:
            if clase:
                self.delete_clase_assignment(clase)
                return 1, 0
            return 0, 0
        if not clase:
            Clase.objects.create(
                horario_aula_curso=horario_aula_curso,
                materia_curso=materia_grupo,
                docente=docente,
                docente_override=docente_override,
                fecha=fecha_clase,
            )
            if docente_override and docente:
                ensure_planificaciones_tema_for_docente_materia_curso(docente, materia_grupo)
            return 1, 0
        previous_docente = clase.docente if clase.docente_override and clase.docente_id else None
        previous_materia_curso = clase.materia_curso if previous_docente else None
        update_fields = []
        if clase.materia_curso_id != materia_grupo.pk:
            reset_fields = self.reset_clase_planificacion(clase)
            clase.materia_curso = materia_grupo
            update_fields.extend(["materia_curso", *reset_fields])
        docente_id = docente.pk if docente else None
        if clase.docente_id != docente_id:
            clase.docente = docente
            update_fields.append("docente")
        if clase.docente_override != docente_override:
            clase.docente_override = docente_override
            update_fields.append("docente_override")
        if update_fields:
            clase.save(update_fields=update_fields)
        if docente_override and docente:
            ensure_planificaciones_tema_for_docente_materia_curso(docente, materia_grupo)
        cleanup_auto_planificaciones_tema_for_docente_materia_curso(previous_docente, previous_materia_curso)
        return 1, 0

    def delete_clase_assignment(self, clase):
        previous_docente = clase.docente if clase.docente_override and clase.docente_id else None
        previous_materia_curso = clase.materia_curso if previous_docente else None
        self.clear_clase_relations(clase)
        clase.delete()
        cleanup_auto_planificaciones_tema_for_docente_materia_curso(previous_docente, previous_materia_curso)

    def reset_clase_planificacion(self, clase):
        self.clear_clase_relations(clase)
        clase.tema = None
        clase.subtema = None
        clase.descripcion = ""
        clase.estado_planificacion = "pendiente"
        clase.notas_revision = ""
        clase.observaciones_revision = {}
        clase.revisado_por = None
        clase.fecha_revision = None
        clase.asistencia_cerrada = False
        clase.asistencia_cerrada_por = None
        clase.fecha_cierre_asistencia = None
        clase.revision_tema_ok = False
        clase.revision_detalle_ok = False
        clase.revision_competencias_ok = False
        clase.revision_estrategias_ok = False
        clase.revision_recursos_ok = False
        return [
            "tema",
            "subtema",
            "descripcion",
            "estado_planificacion",
            "notas_revision",
            "observaciones_revision",
            "revisado_por",
            "fecha_revision",
            "asistencia_cerrada",
            "asistencia_cerrada_por",
            "fecha_cierre_asistencia",
            "revision_tema_ok",
            "revision_detalle_ok",
            "revision_competencias_ok",
            "revision_estrategias_ok",
            "revision_recursos_ok",
        ]

    def clear_clase_relations(self, clase):
        clase.clase_subtemas.all().delete()
        clase.competencias.clear()
        clase.estrategias.clear()
        ClaseRecurso.objects.filter(clase=clase).delete()
        ClaseAsistencia.objects.filter(clase=clase).delete()
        ClaseEstudianteMovimiento.objects.filter(Q(clase_origen=clase) | Q(clase_destino=clase)).delete()
        self.delete_legacy_clase_recursos(clase.pk)

    def delete_legacy_clase_recursos(self, clase_id):
        if connection.vendor != "postgresql":
            return
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT 1
                FROM information_schema.tables
                WHERE table_schema = 'academico'
                  AND table_name = 'clase_recursos'
                """
            )
            if cursor.fetchone():
                cursor.execute("DELETE FROM academico.clase_recursos WHERE clase_id = %s", [clase_id])

    def get_context(self, schedule_form=None, schedule_action=None):
        cursos = Curso.objects.filter(activo=True).order_by("nombre")
        materias = list(Materia.objects.order_by("nombre"))
        materias_asignables = []
        docentes = list(Partner.objects.filter(es_docente=True, activo=True).order_by("nombre"))
        selected_curso_id = self.request.GET.get("curso") or self.request.POST.get("curso") or ""
        selected_curso = None
        selected_curso_periodo = None
        unassigned_alert = clases_sin_docente_alert()
        rows = []
        calendar_events = []
        calendar_default_date = timezone.localdate().isoformat()
        schedule_form = schedule_form or HorarioDistribucionItemForm(prefix="schedule")
        dias_by_name = {item.dia: item.pk for item in Dia.objects.all()}
        weekday_dia_ids = {
            0: dias_by_name.get("Domingo"),
            1: dias_by_name.get("Lunes"),
            2: dias_by_name.get("Martes"),
            3: dias_by_name.get("Miercoles"),
            4: dias_by_name.get("Jueves"),
            5: dias_by_name.get("Viernes"),
            6: dias_by_name.get("Sabado"),
        }
        dia_labels = {str(pk): name for name, pk in dias_by_name.items() if pk}
        schedule_period_enabled = True
        if schedule_form.is_bound:
            schedule_period_enabled = self.request.POST.get("generar_periodo", "on") == "on"
        schedule_action = schedule_action or self.request.POST.get("planning_action") or "add_schedule"
        if schedule_action not in {"add_schedule", "update_schedule"}:
            schedule_action = "add_schedule"

        if selected_curso_id:
            try:
                selected_curso = Curso.objects.filter(pk=selected_curso_id).first()
            except (TypeError, ValueError):
                selected_curso = None
        if selected_curso_id and not selected_curso:
            selected_curso_id = ""
        if selected_curso:
            materias_asignables = list(self.get_assignable_materias(selected_curso))
            unassigned_alert = clases_sin_docente_alert(curso=selected_curso)
            selected_curso_periodo = (
                CursoPeriodo.objects.select_related("periodo")
                .filter(curso=selected_curso)
                .order_by("-periodo__fecha_inicio")
                .first()
            )
            horarios = HorarioAulaCurso.objects.select_related(
                "aula_curso__aula",
                "aula_curso__curso",
                "horario_dia__dia",
                "horario_dia__horario",
            ).filter(aula_curso__curso=selected_curso)
            if selected_curso_periodo:
                periodo = selected_curso_periodo.periodo
                horarios = horarios.filter(Q(fecha__isnull=True) | Q(fecha__gte=periodo.fecha_inicio, fecha__lte=periodo.fecha_fin))
            else:
                horarios = horarios.filter(fecha__isnull=True)
            horarios = list(
                horarios.order_by(
                    "horario_dia__horario__hora_inicio",
                    "horario_dia__horario__hora_fin",
                    "horario_dia__dia__id",
                    "fecha",
                    "aula_curso__aula__nombre",
                )
            )
            clases = {}
            docentes_by_materia_curso = {}
            if selected_curso_periodo:
                clases = {
                    (item.horario_aula_curso_id, item.fecha): item
                    for item in Clase.objects.select_related("materia_curso__materia", "docente").filter(
                        horario_aula_curso__in=horarios,
                        fecha__gte=selected_curso_periodo.periodo.fecha_inicio,
                        fecha__lte=selected_curso_periodo.periodo.fecha_fin,
                    )
                }
                for item in ProfesorMateriaCurso.objects.select_related("partner").filter(
                    materia_curso__grupo=selected_curso,
                    auto_generada_por_clases=False,
                ):
                    docentes_by_materia_curso.setdefault(item.materia_curso_id, []).append(item.partner)

            if selected_curso_periodo:
                weekday_to_dia = self.weekday_to_dia()
                current_date = selected_curso_periodo.periodo.fecha_inicio
                end_date = selected_curso_periodo.periodo.fecha_fin
                today = timezone.localdate()
                if current_date <= today <= end_date:
                    calendar_default_date = today.isoformat()
                else:
                    calendar_default_date = current_date.isoformat()
                while current_date <= end_date:
                    day_name = weekday_to_dia[current_date.weekday()]
                    blocks = []
                    for horario_aula_curso in horarios:
                        if horario_aula_curso.horario_dia.dia.dia != day_name:
                            continue
                        if horario_aula_curso.fecha and horario_aula_curso.fecha != current_date:
                            continue

                        horario = horario_aula_curso.horario_dia.horario
                        clase = clases.get((horario_aula_curso.pk, current_date))
                        materia_id = clase.materia_curso.materia_id if clase else None
                        materia = next((item for item in materias if item.pk == materia_id), None)
                        clase_docentes = []
                        if clase:
                            if clase.docente_override:
                                clase_docentes = [clase.docente] if clase.docente_id else []
                            else:
                                clase_docentes = docentes_by_materia_curso.get(clase.materia_curso_id, [])
                        docente_label = ", ".join(docente.nombre for docente in clase_docentes)
                        sin_docente = bool(clase and materia and not clase_docentes)
                        event_color = "#fef3c7" if sin_docente else (materia.color if materia else "#dff1ff")
                        title_parts = [
                            str(horario_aula_curso.aula_curso.aula),
                            getattr(materia, "nombre_corto", None) or getattr(materia, "nombre", "Sin materia"),
                        ]
                        if docente_label:
                            title_parts.append(docente_label)
                        elif sin_docente:
                            title_parts.append("Sin docente asignado")

                        planning_locked = bool(clase and clase.estado_planificacion in CLASS_ASSIGNMENT_LOCK_STATES)
                        single_date = bool(horario_aula_curso.fecha)
                        class_name = "materia-event sin-docente-event" if sin_docente else ("materia-event" if materia else "sin-materia-event")
                        if single_date:
                            class_name = f"{class_name} single-date-event"
                        is_past_date = current_date < today
                        editable_date = not is_past_date and not planning_locked
                        if is_past_date:
                            class_name = f"{class_name} fecha-pasada-event"
                        if planning_locked:
                            class_name = f"{class_name} approved-locked-event"
                        locked_reason = ""
                        if planning_locked:
                            locked_reason = CLASS_ASSIGNMENT_LOCK_MESSAGE
                        elif is_past_date:
                            locked_reason = "No se puede modificar una clase con fecha anterior a hoy."

                        blocks.append(
                            {
                                "id": horario_aula_curso.pk,
                                "aula": horario_aula_curso.aula_curso.aula,
                                "hora": f"{horario.hora_inicio:%H:%M} - {horario.hora_fin:%H:%M}",
                                "materia_id": materia_id,
                            }
                        )
                        calendar_events.append(
                            {
                                "id": f"{horario_aula_curso.pk}-{current_date.isoformat()}",
                                "horarioId": horario_aula_curso.pk,
                                "fecha": current_date.isoformat(),
                                "title": " · ".join(title_parts),
                                "start": f"{current_date.isoformat()}T{horario.hora_inicio:%H:%M:%S}",
                                "end": f"{current_date.isoformat()}T{horario.hora_fin:%H:%M:%S}",
                                "allDay": False,
                                "className": class_name,
                                "color": event_color,
                                "backgroundColor": event_color,
                                "borderColor": "#f59e0b" if sin_docente else event_color,
                                "textColor": "#92400e" if sin_docente else (readable_text_color(event_color) if materia else "#7c3aed"),
                                "aula": str(horario_aula_curso.aula_curso.aula),
                                "aulaId": horario_aula_curso.aula_curso.aula_id,
                                "diaId": horario_aula_curso.horario_dia.dia_id,
                                "hora": f"{horario.hora_inicio:%H:%M} - {horario.hora_fin:%H:%M}",
                                "horaInicio": f"{horario.hora_inicio:%H:%M}",
                                "horaFin": f"{horario.hora_fin:%H:%M}",
                                "materiaId": materia_id or "",
                                "materia": getattr(materia, "nombre", "") if materia else "",
                                "docenteId": self.get_docente_field_value(clase),
                                "docente": docente_label or ("Sin docente asignado" if sin_docente else ""),
                                "editableDate": editable_date,
                                "lockedReason": locked_reason,
                                "singleDate": single_date,
                                "scope": "Solo esta fecha" if single_date else "Todo el periodo",
                            }
                        )
                    if blocks:
                        rows.append({"fecha": current_date, "dia": day_name, "blocks": blocks})
                    current_date += timedelta(days=1)

        return {
            "title": "Planificacion academica",
            "periodo_create_url": reverse_lazy("academico:periodo_nuevo"),
            "cursos": cursos,
            "materias": materias,
            "materias_asignables": materias_asignables,
            "docentes": docentes,
            "rows": rows,
            "calendar_events_json": json.dumps(calendar_events),
            "calendar_default_date": calendar_default_date,
            "selected_curso": selected_curso,
            "selected_curso_periodo": selected_curso_periodo,
            "selected_curso_id": selected_curso_id,
            "schedule_form": schedule_form,
            "schedule_form_has_errors": schedule_form.is_bound and bool(schedule_form.errors),
            "schedule_action": schedule_action,
            "schedule_edit_id": self.request.POST.get("schedule_horario_aula_curso", ""),
            "schedule_date_value": self.request.POST.get("schedule_fecha", ""),
            "weekday_dia_ids_json": json.dumps(weekday_dia_ids),
            "dia_labels_json": json.dumps(dia_labels),
            "schedule_period_enabled": schedule_period_enabled,
            "unassigned_alert": unassigned_alert,
            "can_save": self.can_save_class(self.request.user),
            "can_add_schedule": self.can_add_schedule(self.request.user),
        }

    def get_assignable_materias(self, curso):
        return (
            Materia.objects.filter(
                Q(temas_base__isnull=False)
                | Q(
                    materia_cursos__grupo=curso,
                    materia_cursos__planificaciones__temas_planificacion__isnull=False,
                )
            )
            .distinct()
            .order_by("nombre")
        )

    def get_materia_curso_with_temas(self, curso, materia_id):
        if not str(materia_id or "").isdigit():
            return None
        materia = Materia.objects.filter(pk=materia_id).first()
        if not materia:
            return None
        materia_curso = MateriaCurso.objects.filter(grupo=curso, materia=materia).first()
        if MateriaTema.objects.filter(materia=materia).exists():
            materia_curso, _ = MateriaCurso.objects.get_or_create(grupo=curso, materia=materia)
            sync_materia_temas_to_materia_curso(materia_curso)
            return materia_curso
        if materia_curso and materia_curso.planificaciones.filter(temas_planificacion__isnull=False).exists():
            return materia_curso
        return None

    def get_docente_field_value(self, clase):
        if not clase or not clase.docente_override:
            return ""
        return clase.docente_id or "__none__"


class PlanificacionAcademicaExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "academico.view_clase"

    weekday_headers = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]

    def get(self, request):
        export_type = request.GET.get("tipo") or "general"
        rows = self.get_rows()
        workbook = Workbook()
        workbook.remove(workbook.active)

        if export_type == "docente":
            grouped = self.group_rows(rows, lambda row: row["docente"] or "Sin docente asignado")
            filename = "horarios_por_docente.xlsx"
        elif export_type == "aula":
            grouped = self.group_rows(rows, lambda row: row["aula"] or "Sin aula")
            filename = "horarios_por_aula.xlsx"
        else:
            grouped = {"General": rows}
            filename = "horarios_general.xlsx"

        for title, sheet_rows in grouped.items():
            self.write_sheet(workbook, title, sheet_rows)

        if not grouped:
            self.write_sheet(workbook, "General", [])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def get_rows(self):
        queryset = (
            Clase.objects.select_related(
                "materia_curso__materia",
                "materia_curso__grupo",
                "docente",
                "horario_aula_curso__aula_curso__aula",
                "horario_aula_curso__horario_dia__dia",
                "horario_aula_curso__horario_dia__horario",
            )
            .order_by(
                "fecha",
                "horario_aula_curso__horario_dia__horario__hora_inicio",
                "materia_curso__grupo__nombre",
            )
        )
        curso_id = self.request.GET.get("curso")
        if curso_id:
            queryset = queryset.filter(materia_curso__grupo_id=curso_id)

        docentes = {}
        for item in ProfesorMateriaCurso.objects.select_related("partner", "materia_curso").filter(
            auto_generada_por_clases=False
        ):
            docentes.setdefault(item.materia_curso_id, []).append(item.partner.nombre)
        rows = []
        for clase in queryset:
            horario = clase.horario_aula_curso.horario_dia.horario
            materia = clase.materia_curso.materia
            if clase.docente_override:
                docente_label = clase.docente.nombre if clase.docente_id else ""
            else:
                docente_label = ", ".join(docentes.get(clase.materia_curso_id, []))
            rows.append(
                {
                    "fecha": clase.fecha,
                    "dia": clase.horario_aula_curso.horario_dia.dia.dia,
                    "hora": f"{horario.hora_inicio:%H:%M} - {horario.hora_fin:%H:%M}",
                    "grupo": clase.materia_curso.grupo.nombre,
                    "aula": str(clase.horario_aula_curso.aula_curso.aula),
                    "materia": materia.nombre,
                    "docente": docente_label or "Sin docente asignado",
                    "color": materia.color,
                }
            )
        return rows

    def group_rows(self, rows, key_func):
        grouped = {}
        for row in rows:
            grouped.setdefault(key_func(row), []).append(row)
        return dict(sorted(grouped.items(), key=lambda item: item[0]))

    def write_sheet(self, workbook, title, rows):
        sheet = workbook.create_sheet(safe_sheet_title(title))
        title_fill = PatternFill("solid", fgColor="DCEBFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        time_fill = PatternFill("solid", fgColor="5B9BD5")
        thin = Side(style="thin", color="D7DEE8")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        schedule_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        title_cell = sheet.cell(row=1, column=1, value=f"Horario - {title}")
        title_cell.font = Font(bold=True, size=14, color="111827")
        title_cell.fill = title_fill
        title_cell.alignment = center

        current_row = 3
        for week_start, week_rows in self.group_rows_by_week(rows).items():
            week_end = week_start + timedelta(days=6)
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            week_cell = sheet.cell(
                row=current_row,
                column=1,
                value=f"Semana {week_start:%d/%m/%Y} - {week_end:%d/%m/%Y}",
            )
            week_cell.font = Font(bold=True, color="111827")
            week_cell.fill = title_fill
            week_cell.alignment = center
            current_row += 1

            sheet.cell(row=current_row, column=1, value="Hora")
            for col, day in enumerate(self.weekday_headers, start=2):
                sheet.cell(row=current_row, column=col, value=day)
            for col in range(1, 9):
                cell = sheet.cell(row=current_row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center
            current_row += 1

            time_slots = sorted({row["hora"] for row in week_rows})
            week_matrix = {}
            for row in week_rows:
                week_matrix.setdefault((row["hora"], row["fecha"].weekday()), []).append(row)
            for time_slot in time_slots:
                time_cell = sheet.cell(row=current_row, column=1, value=time_slot)
                time_cell.font = Font(bold=True, color="FFFFFF")
                time_cell.fill = time_fill
                time_cell.border = border
                time_cell.alignment = center
                max_lines = 1
                for weekday in range(7):
                    cell = sheet.cell(row=current_row, column=weekday + 2)
                    entries = week_matrix.get((time_slot, weekday), [])
                    if entries:
                        cell.value = "\n\n".join(self.schedule_label(item) for item in entries)
                        cell.fill = PatternFill("solid", fgColor=xlsx_color(entries[0]["color"]))
                        cell.font = Font(
                            color=xlsx_color(readable_text_color(entries[0]["color"])),
                            bold=True,
                            size=9,
                        )
                        max_lines = max(max_lines, estimated_schedule_wrapped_lines(cell.value))
                    else:
                        cell.value = ""
                        cell.fill = PatternFill("solid", fgColor="FFFFFF")
                    cell.border = border
                    cell.alignment = schedule_alignment
                sheet.row_dimensions[current_row].height = schedule_export_row_height(max_lines)
                current_row += 1

            current_row += 2

        if not rows:
            sheet.cell(row=3, column=1, value="Sin horarios registrados.")

        widths = [18] + [SCHEDULE_EXPORT_COLUMN_WIDTH] * 7
        for col, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col)].width = width
        sheet.freeze_panes = "B5"

    def group_rows_by_week(self, rows):
        grouped = {}
        for row in rows:
            week_start = row["fecha"] - timedelta(days=row["fecha"].weekday())
            grouped.setdefault(week_start, []).append(row)
        return dict(sorted(grouped.items(), key=lambda item: item[0]))

    def schedule_label(self, row):
        parts = [row["materia"], row["grupo"], row["aula"]]
        if row["docente"]:
            parts.append(row["docente"])
        return "\n".join(parts)


class PlanificacionDocenteListView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "academico.view_profesormateriacurso"
    template_name = "academico/planificacion_docente_list.html"

    def get(self, request):
        cursos = list(Curso.objects.filter(activo=True).order_by("nombre"))
        selected_group = self.get_selected_group(cursos)
        q = (request.GET.get("q") or "").strip()
        asignaciones = self.get_asignaciones()
        if selected_group:
            asignaciones = asignaciones.filter(grupo=selected_group)
        else:
            asignaciones = asignaciones.none()
        if q:
            asignaciones = asignaciones.filter(
                Q(grupo__nombre__icontains=q)
                | Q(materia__nombre__icontains=q)
                | Q(materia__nombre_corto__icontains=q)
                | Q(
                    profesor_materia_cursos__auto_generada_por_clases=False,
                    profesor_materia_cursos__partner__nombre__icontains=q,
                )
                | Q(
                    profesor_materia_cursos__auto_generada_por_clases=False,
                    profesor_materia_cursos__partner__identificacion__icontains=q,
                )
            )
        asignaciones = list(asignaciones.distinct())
        grupos = self.group_asignaciones(asignaciones)
        active_group_summary = grupos[0] if grupos else self.empty_group_summary(selected_group)
        stats = self.get_stats(asignaciones)
        return render(
            request,
            self.template_name,
            {
                "title": "Planificacion docente",
                "docentes_disponibles": Partner.objects.filter(es_docente=True, activo=True).order_by("nombre"),
                "group_tabs": self.get_group_tabs(cursos, selected_group, q),
                "selected_group": selected_group,
                "active_group_summary": active_group_summary,
                "assignment_groups": grupos,
                "stats": stats,
                "unassigned_alert": clases_sin_docente_alert(curso=selected_group) if selected_group else clases_sin_docente_alert(),
                "can_assign_docente": can_assign_docente(request.user),
            },
        )

    def post(self, request):
        if not can_assign_docente(request.user):
            return self.handle_no_permission()

        materia_curso = get_object_or_404(
            MateriaCurso.objects.select_related("grupo", "materia"),
            pk=request.POST.get("materia_curso"),
        )
        docente_id = request.POST.get("docente") or ""
        with transaction.atomic():
            if docente_id:
                docente = get_object_or_404(Partner, pk=docente_id, es_docente=True, activo=True)
                assignment_result = assign_real_docente_to_materia_curso(materia_curso, docente)
                if not assignment_result["ok"]:
                    messages.error(request, assignment_result["message"])
                    return self.redirect_after_post(request, materia_curso)
                if assignment_result["preserved_count"] and assignment_result["replaced"]:
                    messages.success(
                        request,
                        f"{docente.nombre} tomara las clases pendientes de {materia_curso.grupo} - {materia_curso.materia}. "
                        "Las clases enviadas o aprobadas quedaron con el docente anterior.",
                    )
                else:
                    messages.success(
                        request,
                        f"{docente.nombre} asignado a {materia_curso.grupo} - {materia_curso.materia}.",
                    )
            else:
                assignment_result = remove_real_docentes_from_materia_curso(materia_curso)
                if not assignment_result["ok"]:
                    messages.error(request, assignment_result["message"])
                    return self.redirect_after_post(request, materia_curso)
                messages.success(request, f"{materia_curso.grupo} - {materia_curso.materia} quedo sin docente asignado.")

        return self.redirect_after_post(request, materia_curso)

    def redirect_after_post(self, request, materia_curso):
        redirect_url = reverse_lazy("academico:planificacion_docente")
        redirect_params = {"grupo": request.POST.get("grupo") or materia_curso.grupo_id}
        q = request.POST.get("q") or ""
        if q:
            redirect_params["q"] = q
        if redirect_params:
            redirect_url = f"{redirect_url}?{urlencode(redirect_params)}"
        return redirect(redirect_url)

    def get_selected_group(self, cursos):
        grupo_id = self.request.GET.get("grupo") or ""
        if grupo_id:
            selected = next((curso for curso in cursos if str(curso.pk) == str(grupo_id)), None)
            if selected:
                return selected

        pending_group_id = (
            MateriaCurso.objects.filter(grupo__in=cursos)
            .annotate(
                total_docentes=Count(
                    "profesor_materia_cursos",
                    filter=Q(profesor_materia_cursos__auto_generada_por_clases=False),
                    distinct=True,
                )
            )
            .filter(total_docentes=0)
            .order_by("grupo__nombre")
            .values_list("grupo_id", flat=True)
            .first()
        )
        if pending_group_id:
            return next((curso for curso in cursos if curso.pk == pending_group_id), None)
        return cursos[0] if cursos else None

    def get_asignaciones(self):
        today = timezone.localdate()
        return (
            MateriaCurso.objects.select_related("grupo", "materia")
            .prefetch_related(
                Prefetch(
                    "profesor_materia_cursos",
                    queryset=ProfesorMateriaCurso.objects.select_related("partner")
                    .filter(auto_generada_por_clases=False)
                    .order_by("partner__nombre"),
                )
            )
            .annotate(
                total_docentes=Count(
                    "profesor_materia_cursos",
                    filter=Q(profesor_materia_cursos__auto_generada_por_clases=False),
                    distinct=True,
                ),
                total_clases=Count("clases", distinct=True),
                proximas_clases=Count(
                    "clases",
                    filter=Q(clases__fecha__gte=today, clases__fecha__lte=today + timedelta(days=UNASSIGNED_CLASS_ALERT_DAYS)),
                    distinct=True,
                ),
                proxima_clase=Min("clases__fecha", filter=Q(clases__fecha__gte=today)),
            )
            .order_by("grupo__nombre", "materia__nombre")
        )

    def get_group_tabs(self, cursos, selected_group, q=""):
        rows = (
            MateriaCurso.objects.filter(grupo__in=cursos)
            .values("grupo_id")
            .annotate(
                total=Count("id", distinct=True),
                asignadas=Count(
                    "id",
                    filter=Q(profesor_materia_cursos__auto_generada_por_clases=False),
                    distinct=True,
                ),
            )
        )
        counters = {
            row["grupo_id"]: {
                "total": row["total"],
                "asignadas": row["asignadas"],
                "pendientes": row["total"] - row["asignadas"],
            }
            for row in rows
        }
        tabs = []
        base_url = reverse_lazy("academico:planificacion_docente")
        for curso in cursos:
            params = {"grupo": curso.pk}
            if q:
                params["q"] = q
            stats = counters.get(curso.pk, {"total": 0, "asignadas": 0, "pendientes": 0})
            tabs.append(
                {
                    "grupo": curso,
                    "url": f"{base_url}?{urlencode(params)}",
                    "is_active": bool(selected_group and selected_group.pk == curso.pk),
                    **stats,
                }
            )
        return tabs

    def group_asignaciones(self, asignaciones):
        grupos = []
        grouped = {}
        for materia_curso in asignaciones:
            docentes = [item.partner for item in materia_curso.profesor_materia_cursos.all()]
            item = {
                "materia_curso": materia_curso,
                "docentes": docentes,
                "docente_principal_id": docentes[0].pk if len(docentes) == 1 else "",
                "has_docente": bool(docentes),
                "total_clases": materia_curso.total_clases,
                "proximas_clases": materia_curso.proximas_clases,
                "proxima_clase": materia_curso.proxima_clase,
            }
            grouped.setdefault(materia_curso.grupo, []).append(item)

        for grupo, items in grouped.items():
            pendientes = sum(1 for item in items if not item["has_docente"])
            grupos.append(
                {
                    "grupo": grupo,
                    "items": items,
                    "total": len(items),
                    "pendientes": pendientes,
                    "asignadas": len(items) - pendientes,
                }
            )
        return grupos

    def empty_group_summary(self, selected_group):
        return {
            "grupo": selected_group,
            "items": [],
            "total": 0,
            "pendientes": 0,
            "asignadas": 0,
        }

    def get_stats(self, asignaciones):
        total = len(asignaciones)
        asignadas = sum(1 for item in asignaciones if item.total_docentes)
        return {
            "total": total,
            "asignadas": asignadas,
            "pendientes": total - asignadas,
            "docentes": Partner.objects.filter(es_docente=True, activo=True).count(),
        }


class GrupoEstudianteListView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "academico.view_grupoestudiante"
    template_name = "academico/grupo_estudiantes.html"

    def get(self, request):
        selected_group = self.get_selected_group()
        return render(request, self.template_name, self.get_context(selected_group=selected_group))

    def post(self, request):
        if not self.can_manage(request.user):
            return self.handle_no_permission()

        selected_group = self.get_selected_group()
        action = request.POST.get("assignment_action") or "assign_students"
        if action == "sync_students":
            return self.handle_sync_students(request, selected_group)
        if action == "assign_students":
            return self.handle_assign_students(request, selected_group)
        if action == "move_student":
            return self.handle_move_student(request, selected_group)
        if action == "cancel_move":
            return self.handle_cancel_move(request, selected_group)
        messages.error(request, "Accion no valida.")
        return self.redirect_to_group(selected_group)

    def can_manage(self, user):
        return (
            user.is_superuser
            or user.has_perm("academico.add_grupoestudiante")
            or user.has_perm("academico.change_grupoestudiante")
        )

    def get_selected_group(self):
        grupo_id = self.request.POST.get("grupo") or self.request.GET.get("grupo") or ""
        if grupo_id:
            try:
                return Curso.objects.filter(pk=grupo_id, activo=True).first()
            except (TypeError, ValueError):
                return None
        return Curso.objects.filter(activo=True).order_by("nombre").first()

    def redirect_to_group(self, group=None):
        url = reverse_lazy("academico:grupo_estudiantes")
        if group:
            url = f"{url}?{urlencode({'grupo': group.pk})}"
        return redirect(url)

    def handle_assign_students(self, request, selected_group):
        form = GrupoEstudianteBulkForm(request.POST, selected_group=selected_group)
        if form.is_valid():
            grupo = form.cleaned_data["grupo"]
            saved = 0
            with transaction.atomic():
                for ficha in form.cleaned_data["fichas"]:
                    asignacion = GrupoEstudiante(
                        ficha_inscripcion=ficha,
                        estudiante=ficha.estudiante,
                        grupo=grupo,
                        fecha_asignacion=form.cleaned_data["fecha_asignacion"],
                        estado="activo",
                        usuario_updated=request.user,
                    )
                    asignacion.full_clean()
                    asignacion.save()
                    saved += 1
            if saved:
                messages.success(request, f"{saved} estudiante(s) asignado(s) al grupo.")
            else:
                messages.info(request, "Selecciona al menos un estudiante sin grupo.")
            return self.redirect_to_group(grupo)
        return render(request, self.template_name, self.get_context(selected_group=selected_group, bulk_form=form))

    def handle_sync_students(self, request, selected_group):
        form = GrupoEstudianteBulkForm(request.POST, selected_group=selected_group)
        if form.is_valid():
            grupo = form.cleaned_data["grupo"]
            assignments_to_remove = GrupoEstudiante.objects.filter(
                pk__in=request.POST.getlist("asignaciones_remover"),
                grupo=grupo,
            )
            removed = assignments_to_remove.count()
            saved = 0
            with transaction.atomic():
                assignments_to_remove.delete()
                for ficha in form.cleaned_data["fichas"]:
                    if GrupoEstudiante.objects.filter(ficha_inscripcion=ficha).exists():
                        continue
                    asignacion = GrupoEstudiante(
                        ficha_inscripcion=ficha,
                        estudiante=ficha.estudiante,
                        grupo=grupo,
                        fecha_asignacion=form.cleaned_data["fecha_asignacion"],
                        estado="activo",
                        usuario_updated=request.user,
                    )
                    asignacion.full_clean()
                    asignacion.save()
                    saved += 1
            if saved or removed:
                messages.success(
                    request,
                    f"Grupo actualizado: {saved} asignado(s) a todas sus clases, {removed} removido(s).",
                )
            else:
                messages.info(request, "No hubo cambios para guardar.")
            return self.redirect_to_group(grupo)
        return render(request, self.template_name, self.get_context(selected_group=selected_group, bulk_form=form))

    def handle_move_student(self, request, selected_group):
        form = ClaseEstudianteMovimientoForm(request.POST, grupo=selected_group)
        if form.is_valid():
            asignacion = form.cleaned_data["asignacion"]
            materia_origen = form.cleaned_data["materia_origen"]
            clase_origen = form.cleaned_data["clase_origen"]
            with transaction.atomic():
                active_movement = ClaseEstudianteMovimiento.objects.filter(
                    asignacion=asignacion,
                    clase_origen__materia_curso=materia_origen,
                    activo=True,
                ).first()
                exact_movement = ClaseEstudianteMovimiento.objects.filter(
                    asignacion=asignacion,
                    clase_origen=clase_origen,
                ).first()
                if active_movement and exact_movement and active_movement.pk != exact_movement.pk:
                    active_movement.activo = False
                    active_movement.usuario_updated = request.user
                    active_movement.save(update_fields=["activo", "usuario_updated", "updated_at"])
                movimiento = exact_movement or active_movement or ClaseEstudianteMovimiento(asignacion=asignacion)
                movimiento.asignacion = asignacion
                movimiento.clase_origen = clase_origen
                movimiento.clase_destino = form.cleaned_data["clase_destino"]
                movimiento.fecha_inicio = form.cleaned_data["fecha_inicio"]
                movimiento.motivo = form.cleaned_data.get("motivo") or ""
                movimiento.usuario_updated = request.user
                movimiento.activo = True
                movimiento.full_clean()
                movimiento.save()
            messages.success(
                request,
                "Cambio guardado. El estudiante conserva su grupo base y cursara solo esta materia en el horario destino.",
            )
            return self.redirect_to_group(movimiento.asignacion.grupo)
        return render(request, self.template_name, self.get_context(selected_group=selected_group, movement_form=form))

    def handle_cancel_move(self, request, selected_group):
        movimiento_id = request.POST.get("movimiento_id")
        movimiento = get_object_or_404(
            ClaseEstudianteMovimiento,
            pk=movimiento_id,
            clase_origen__materia_curso__grupo=selected_group,
        )
        movimiento.activo = False
        movimiento.usuario_updated = request.user
        movimiento.save(update_fields=["activo", "usuario_updated", "updated_at"])
        messages.success(request, "Movimiento anulado correctamente.")
        return self.redirect_to_group(selected_group)

    def get_context(self, selected_group=None, bulk_form=None, movement_form=None):
        cursos = list(
            Curso.objects.filter(activo=True)
            .annotate(total_estudiantes=Count(
                "estudiantes_asignados",
                filter=Q(estudiantes_asignados__estado="activo"),
                distinct=True,
            ))
            .order_by("nombre")
        )
        if selected_group is None and cursos:
            selected_group = cursos[0]
        selected_group_id = selected_group.pk if selected_group else None
        q = ""
        asignaciones = GrupoEstudiante.objects.select_related(
            "ficha_inscripcion",
            "estudiante",
            "grupo",
        ).filter(grupo=selected_group) if selected_group else GrupoEstudiante.objects.none()
        asignaciones = asignaciones.order_by("estudiante__nombre")
        bulk_form = bulk_form or GrupoEstudianteBulkForm(selected_group=selected_group)
        clases = list(self.get_group_classes(selected_group))
        movement_materia_cursos = list(self.get_movement_materia_cursos(selected_group))
        movimientos = self.get_group_movements(selected_group)
        return {
            "title": "Estudiantes por grupo",
            "group_tabs": self.get_group_tabs(cursos, selected_group, q),
            "selected_group": selected_group,
            "selected_group_id": selected_group_id,
            "asignaciones": asignaciones,
            "available_fichas": list(bulk_form.fields["fichas"].queryset),
            "clases": clases,
            "movement_materia_cursos": movement_materia_cursos,
            "movimientos": movimientos,
            "bulk_form": bulk_form,
            "movement_form": movement_form or ClaseEstudianteMovimientoForm(grupo=selected_group),
            "movement_materia_map_json": json.dumps(self.get_movement_materia_map(movement_materia_cursos)),
            "can_manage": self.can_manage(self.request.user),
            "q": q,
            "stats": self.get_stats(selected_group),
        }

    def get_group_tabs(self, cursos, selected_group, q=""):
        base_url = reverse_lazy("academico:grupo_estudiantes")
        tabs = []
        for curso in cursos:
            params = {"grupo": curso.pk}
            if q:
                params["q"] = q
            tabs.append(
                {
                    "grupo": curso,
                    "url": f"{base_url}?{urlencode(params)}",
                    "is_active": bool(selected_group and curso.pk == selected_group.pk),
                    "total_estudiantes": curso.total_estudiantes,
                }
            )
        return tabs

    def get_group_classes(self, selected_group):
        if not selected_group:
            return Clase.objects.none()
        return (
            Clase.objects.select_related(
                "materia_curso__materia",
                "materia_curso__grupo",
                "horario_aula_curso__aula_curso__aula",
                "horario_aula_curso__horario_dia__horario",
            )
            .filter(materia_curso__grupo=selected_group)
            .order_by("fecha", "materia_curso__materia__nombre", "horario_aula_curso__horario_dia__horario__hora_inicio")
        )

    def get_movement_materia_cursos(self, selected_group):
        if not selected_group:
            return []
        materias = MateriaCurso.objects.filter(grupo=selected_group).values_list("materia_id", flat=True)
        return (
            MateriaCurso.objects.select_related("materia", "grupo")
            .filter(
                Q(grupo=selected_group) | Q(materia_id__in=materias),
            )
            .filter(clases__isnull=False)
            .distinct()
            .order_by("materia__nombre", "grupo__nombre")
        )

    def get_movement_materia_map(self, materias_curso):
        return {
            str(materia_curso.pk): {
                "materiaId": materia_curso.materia_id,
                "grupoId": materia_curso.grupo_id,
            }
            for materia_curso in materias_curso
        }

    def get_group_movements(self, selected_group):
        if not selected_group:
            return ClaseEstudianteMovimiento.objects.none()
        return (
            ClaseEstudianteMovimiento.objects.select_related(
                "asignacion__estudiante",
                "clase_origen__materia_curso__materia",
                "clase_origen__materia_curso__grupo",
                "clase_origen__horario_aula_curso__aula_curso__aula",
                "clase_origen__horario_aula_curso__horario_dia__horario",
                "clase_destino__materia_curso__grupo",
                "clase_destino__materia_curso__materia",
                "clase_destino__horario_aula_curso__aula_curso__aula",
                "clase_destino__horario_aula_curso__horario_dia__horario",
            )
            .filter(clase_origen__materia_curso__grupo=selected_group)
            .order_by("-fecha_inicio", "-created_at")[:20]
        )

    def get_stats(self, selected_group):
        if not selected_group:
            return {"asignados": 0, "sin_grupo": 0, "clases": 0, "movimientos": 0}
        sin_grupo = (
            FichaInscripcion.objects.filter(estudiante__es_estudiante=True, activo=True)
            .exclude(estado="anulada")
            .filter(asignacion_grupo__isnull=True)
            .count()
        )
        return {
            "asignados": GrupoEstudiante.objects.filter(grupo=selected_group, estado="activo").count(),
            "sin_grupo": sin_grupo,
            "clases": Clase.objects.filter(materia_curso__grupo=selected_group).count(),
            "movimientos": ClaseEstudianteMovimiento.objects.filter(
                clase_origen__materia_curso__grupo=selected_group,
                activo=True,
            ).count(),
        }


class PlanificacionDocenteAsignadorView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "academico.view_profesormateriacurso"
    template_name = "academico/planificacion_docente.html"

    def get(self, request, *args, **kwargs):
        docente = self.get_docente()
        base_form = PlanificacionDocenteBaseForm(initial={"docente": docente})
        formset = self.get_formset(docente=docente)
        return render(request, self.template_name, self.get_context(base_form, formset, docente))

    def post(self, request, *args, **kwargs):
        if not (
            request.user.has_perm("academico.add_profesormateriacurso")
            or request.user.has_perm("academico.change_profesormateriacurso")
            or request.user.is_superuser
        ):
            return self.handle_no_permission()

        base_form = PlanificacionDocenteBaseForm(request.POST)
        docente = None
        if base_form.is_valid():
            docente = base_form.cleaned_data["docente"]
        formset = PlanificacionDocenteFormSet(request.POST, form_kwargs={"docente": docente})

        if base_form.is_valid() and formset.is_valid():
            rows = []
            materia_ids = set()
            duplicate = False
            for form in formset:
                if form.cleaned_data.get("DELETE") or not form.has_assignment_data():
                    continue
                materia_curso = form.cleaned_data["materia_curso"]
                if materia_curso.pk in materia_ids:
                    form.add_error("materia_curso", "Esta materia ya esta repetida en el formulario.")
                    duplicate = True
                    continue
                materia_ids.add(materia_curso.pk)
                rows.append(materia_curso)

            if not duplicate:
                with transaction.atomic():
                    removed_assignments = list(
                        ProfesorMateriaCurso.objects.select_for_update()
                        .select_related("materia_curso__grupo", "materia_curso__materia")
                        .filter(partner=docente, auto_generada_por_clases=False)
                        .exclude(materia_curso_id__in=materia_ids)
                    )
                    locked_removed = [
                        item
                        for item in removed_assignments
                        if locked_inherited_classes_queryset(item.materia_curso).exists()
                    ]
                    if locked_removed:
                        materias = ", ".join(
                            f"{item.materia_curso.grupo} - {item.materia_curso.materia}"
                            for item in locked_removed
                        )
                        messages.error(
                            request,
                            "No se puede quitar el docente porque hay clases con planificacion enviada o aprobada: "
                            f"{materias}.",
                        )
                        return render(request, self.template_name, self.get_context(base_form, formset, docente))
                    ProfesorMateriaCurso.objects.filter(
                        partner=docente,
                        auto_generada_por_clases=False,
                    ).exclude(materia_curso_id__in=materia_ids).delete()
                    for materia_curso in rows:
                        profesor_materia_curso, _ = ProfesorMateriaCurso.objects.get_or_create(
                            partner=docente,
                            materia_curso=materia_curso,
                        )
                        if profesor_materia_curso.auto_generada_por_clases:
                            profesor_materia_curso.auto_generada_por_clases = False
                            profesor_materia_curso.save(update_fields=["auto_generada_por_clases"])
                        sync_planificaciones_tema_for_profesor(profesor_materia_curso)
                messages.success(request, "Planificacion docente guardada correctamente.")
                return redirect("academico:planificacion_docente")

        return render(request, self.template_name, self.get_context(base_form, formset, docente))

    def get_docente(self):
        docente_id = self.kwargs.get("docente_pk") or self.request.GET.get("docente")
        if not docente_id:
            return None
        return get_object_or_404(Partner, pk=docente_id, es_docente=True, activo=True)

    def get_formset(self, docente=None):
        initial = []
        if docente:
            initial = [
                {
                    "grupo": item.materia_curso.grupo,
                    "materia_curso": item.materia_curso,
                }
                for item in ProfesorMateriaCurso.objects.select_related(
                    "materia_curso__grupo",
                    "materia_curso__materia",
                )
                .filter(partner=docente, auto_generada_por_clases=False)
                .order_by("materia_curso__grupo__nombre", "materia_curso__materia__nombre")
            ]
        if not initial:
            initial = [{}]
        return PlanificacionDocenteFormSet(initial=initial, form_kwargs={"docente": docente})

    def get_context(self, base_form, formset, docente):
        materia_options = [
            {
                "id": item.pk,
                "grupo_id": item.grupo_id,
                "nombre": item.materia.nombre,
            }
            for item in MateriaCurso.objects.select_related("materia", "grupo").order_by("grupo__nombre", "materia__nombre")
        ]
        return {
            "title": "Planificacion docente",
            "base_form": base_form,
            "formset": formset,
            "selected_docente": docente,
            "materia_options_json": json.dumps(materia_options),
            "list_url": reverse_lazy("academico:planificacion_docente"),
            "assign_url": reverse_lazy("academico:planificacion_docente_asignar"),
        }


class CoordinacionPlanificacionListView(CoordinacionRequiredMixin, View):
    permission_required = "academico.view_tema"
    template_name = "academico/coordinacion_planificacion_list.html"

    def get(self, request):
        q = request.GET.get("q")
        asignaciones = (
            MateriaCurso.objects.select_related(
                "materia",
                "grupo",
            )
            .prefetch_related(
                Prefetch(
                    "profesor_materia_cursos",
                    queryset=ProfesorMateriaCurso.objects.select_related("partner")
                    .filter(auto_generada_por_clases=False)
                    .order_by("partner__nombre"),
                )
            )
            .annotate(total_temas=Count("planificaciones__temas_planificacion", distinct=True))
            .order_by("grupo__nombre", "materia__nombre")
        )
        if q:
            asignaciones = asignaciones.filter(
                Q(materia__nombre__icontains=q)
                | Q(grupo__nombre__icontains=q)
                | Q(
                    profesor_materia_cursos__auto_generada_por_clases=False,
                    profesor_materia_cursos__partner__nombre__icontains=q,
                )
            )
        return render(
            request,
            self.template_name,
            {
                "title": "Temas y subtemas",
                "asignaciones": asignaciones,
            },
        )


class CoordinacionPlanificacionEditorView(CoordinacionRequiredMixin, View):
    permission_required = "academico.change_tema"
    template_name = "academico/coordinacion_planificacion_form.html"

    def get_materia_curso(self):
        pk = self.kwargs.get("materia_curso_pk")
        if pk:
            return get_object_or_404(
                MateriaCurso.objects.select_related("materia", "grupo").prefetch_related(
                    Prefetch(
                        "profesor_materia_cursos",
                        queryset=ProfesorMateriaCurso.objects.select_related("partner")
                        .filter(auto_generada_por_clases=False)
                        .order_by("partner__nombre"),
                    )
                ),
                pk=pk,
            )
        return None

    def get_planificacion(self, materia_curso):
        if not materia_curso:
            return None
        return (
            PlanificacionDocente.objects.filter(materia_curso=materia_curso)
            .order_by("id")
            .first()
        )

    def get_materia(self, materia_curso=None):
        pk = self.kwargs.get("materia_pk")
        if pk:
            return get_object_or_404(Materia, pk=pk)
        if materia_curso:
            return materia_curso.materia
        return None

    def ensure_planificacion(self, materia_curso):
        return ensure_planificacion_docente_for_materia_curso(materia_curso)

    def get(self, request, *args, **kwargs):
        materia_curso = self.get_materia_curso()
        materia = self.get_materia(materia_curso)
        planificacion = self.get_planificacion(materia_curso)
        form = CoordinacionPlanificacionForm(materia=materia, materia_curso=materia_curso)
        formset = CoordinacionTemaFormSet(initial=self.get_tema_initial(materia, planificacion))
        return render(request, self.template_name, self.get_context(form, formset, materia, materia_curso, planificacion))

    def post(self, request, *args, **kwargs):
        materia_curso = self.get_materia_curso()
        materia = self.get_materia(materia_curso)
        form = CoordinacionPlanificacionForm(request.POST, materia=materia, materia_curso=materia_curso)
        formset = CoordinacionTemaFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            materia = materia or form.cleaned_data["materia"]
            with transaction.atomic():
                self.save_materia_topics(materia, formset)
                target_materia_cursos = list(
                    MateriaCurso.objects.select_related("materia", "grupo")
                    .filter(materia=materia)
                    .order_by("grupo__nombre")
                )
                for target_materia_curso in target_materia_cursos:
                    sync_materia_temas_to_materia_curso(target_materia_curso)
            if target_materia_cursos:
                messages.success(request, f"Temas y subtemas guardados y aplicados en {len(target_materia_cursos)} grupo(s).")
            else:
                messages.success(request, "Temas y subtemas guardados para la materia.")
            return redirect("academico:coordinacion_planificacion_materia_editar", materia_pk=materia.pk)
        return render(request, self.template_name, self.get_context(form, formset, materia, materia_curso, self.get_planificacion(materia_curso)))

    def save_materia_topics(self, materia, formset):
        kept_tema_ids = set()
        for tema_index, tema_form in enumerate(formset):
            if tema_form.cleaned_data.get("DELETE") or not tema_form.has_topic_data():
                continue
            tema_id = tema_form.cleaned_data.get("tema_id")
            tema_nombre = tema_form.cleaned_data["nombre"]
            materia_tema = None
            if tema_id:
                materia_tema = MateriaTema.objects.filter(pk=tema_id, materia=materia).first()
            if materia_tema is None:
                materia_tema = MateriaTema.objects.filter(materia=materia, nombre__iexact=tema_nombre).first()
            if materia_tema is None:
                materia_tema = MateriaTema(materia=materia)
            materia_tema.nombre = tema_nombre
            materia_tema.detalle = tema_form.cleaned_data.get("detalle") or None
            materia_tema.orden = len(kept_tema_ids) + 1
            materia_tema.save()
            kept_tema_ids.add(materia_tema.pk)

            submitted_subtemas = self.get_subtemas_from_post(tema_index)
            kept_subtema_ids = set()
            for subtema_order, subtema_data in enumerate(submitted_subtemas, start=1):
                subtema_nombre = subtema_data["nombre"]
                materia_subtema = None
                if subtema_data["id"]:
                    materia_subtema = materia_tema.subtemas_base.filter(pk=subtema_data["id"]).first()
                if subtema_data["delete"] or not subtema_nombre:
                    if materia_subtema:
                        materia_subtema.delete()
                    continue
                if materia_subtema is None:
                    materia_subtema = materia_tema.subtemas_base.filter(nombre__iexact=subtema_nombre).first()
                if materia_subtema is None:
                    materia_subtema = MateriaSubtema(tema=materia_tema)
                materia_subtema.nombre = subtema_nombre
                materia_subtema.descripcion = None
                materia_subtema.orden = subtema_order
                materia_subtema.save()
                kept_subtema_ids.add(materia_subtema.pk)
            materia_tema.subtemas_base.exclude(pk__in=kept_subtema_ids).delete()

        MateriaTema.objects.filter(materia=materia).exclude(pk__in=kept_tema_ids).delete()

    def get_subtemas_from_post(self, tema_index):
        prefix = f"form-{tema_index}-subtemas"
        try:
            total = int(self.request.POST.get(f"{prefix}-TOTAL_FORMS", 0))
        except (TypeError, ValueError):
            total = 0
        subtemas = []
        for index in range(total):
            row_prefix = f"{prefix}-{index}"
            subtemas.append(
                {
                    "id": self.request.POST.get(f"{row_prefix}-id") or "",
                    "nombre": (self.request.POST.get(f"{row_prefix}-nombre") or "").strip(),
                    "delete": self.request.POST.get(f"{row_prefix}-DELETE") == "on",
                }
            )
        return subtemas

    def get_tema_initial(self, materia=None, planificacion=None):
        if materia:
            temas_base = MateriaTema.objects.filter(materia=materia).prefetch_related("subtemas_base").order_by("orden", "nombre")
            if temas_base.exists():
                initial = []
                for tema in temas_base:
                    initial.append(
                        {
                            "tema_id": tema.pk,
                            "nombre": tema.nombre,
                            "detalle": tema.detalle,
                            "orden": tema.orden,
                            "subtemas": [
                                {"id": subtema.pk, "nombre": subtema.nombre}
                                for subtema in tema.subtemas_base.order_by("orden", "nombre")
                            ],
                        }
                    )
                return initial or [{}]
        if not planificacion:
            return [{}]
        initial = []
        temas = planificacion.temas_planificacion.prefetch_related("subtemas_planificacion").order_by("orden", "nombre")
        for tema in temas:
            initial.append(
                {
                    "tema_id": tema.materia_tema_id or "",
                    "nombre": tema.nombre,
                    "detalle": tema.detalle,
                    "orden": tema.orden,
                    "subtemas": [
                        {"id": subtema.materia_subtema_id or "", "nombre": subtema.nombre}
                        for subtema in tema.subtemas_planificacion.order_by("orden", "nombre")
                    ],
                }
            )
        return initial or [{}]

    def get_context(self, form, formset, materia, materia_curso, planificacion):
        docentes = []
        if materia_curso:
            docentes = [
                item.partner
                for item in materia_curso.profesor_materia_cursos.all()
                if not item.auto_generada_por_clases
            ]
        tema_count = 0
        subtema_count = 0
        applied_group_count = 0
        if materia:
            tema_count = MateriaTema.objects.filter(materia=materia).count()
            subtema_count = MateriaSubtema.objects.filter(tema__materia=materia).count()
            applied_group_count = MateriaCurso.objects.filter(materia=materia).count()
        return {
            "title": "Temas y subtemas",
            "form": form,
            "formset": formset,
            "materia": materia,
            "materia_curso": materia_curso,
            "docentes": docentes,
            "planificacion": planificacion,
            "tema_count": tema_count,
            "subtema_count": subtema_count,
            "applied_group_count": applied_group_count,
            "tema_suggestions": MateriaTema.objects.order_by("nombre").values_list("nombre", flat=True).distinct(),
            "list_url": reverse_lazy("academico:coordinacion_planificacion_list"),
        }


class DocenteHorariosView(LoginRequiredMixin, View):
    template_name = "academico/docente_horarios.html"
    status_filters = (
        {
            "key": "trabajo",
            "label": "Por atender",
            "states": ("pendiente", "rechazada"),
            "icon": "ri-inbox-unarchive-line",
        },
        {
            "key": "revision",
            "label": "En revision",
            "states": ("revision",),
            "icon": "ri-search-eye-line",
        },
        {
            "key": "rechazada",
            "label": "Observadas",
            "states": ("rechazada",),
            "icon": "ri-error-warning-line",
        },
        {
            "key": "aprobada",
            "label": "Aprobadas",
            "states": ("aprobada",),
            "icon": "ri-checkbox-circle-line",
        },
        {
            "key": "todas",
            "label": "Todas",
            "states": None,
            "icon": "ri-calendar-check-line",
        },
    )

    def get(self, request):
        docente = getattr(request.user, "partner", None)
        if not docente or not docente.es_docente:
            return render(
                request,
                self.template_name,
                {
                    "title": "Mis planificaciones",
                    "docente": None,
                    "calendar_events_json": json.dumps([]),
                    "calendar_default_date": timezone.localdate().isoformat(),
                    "has_events": False,
                    "planificacion_stats": self.empty_stats(),
                    "status_tabs": [],
                    "selected_filter": "trabajo",
                    "selected_filter_label": "",
                    "tema_cards": [],
                    "planificacion_cards": [],
                },
            )

        clases = list(
            Clase.objects.select_related(
                "materia_curso__materia",
                "materia_curso__grupo",
                "docente",
                "horario_aula_curso__aula_curso__aula",
                "horario_aula_curso__aula_curso__curso",
                "horario_aula_curso__horario_dia__horario",
            )
            .prefetch_related("competencias", "estrategias", "recursos", "clase_subtemas__subtema")
            .filter(docente_responsable_filter(docente))
            .distinct()
            .order_by("fecha", "horario_aula_curso__horario_dia__horario__hora_inicio")
        )
        stats = self.get_planificacion_stats(clases)
        selected_filter = self.get_selected_filter(request.GET.get("estado"), stats)
        status_tabs = self.get_status_tabs(selected_filter, stats)
        planificaciones_tema = list(self.get_docente_planificaciones_tema_queryset(docente))
        tema_cards = [
            self.build_tema_card(planificacion_tema, clases)
            for planificacion_tema in planificaciones_tema
        ]
        tema_cards = [
            card
            for card in tema_cards
            if self.topic_matches_filter(card, selected_filter)
        ]
        planificacion_cards = [
            self.build_planificacion_card(clase)
            for clase in self.filter_clases(clases, selected_filter)
        ]
        calendar_events = []
        calendar_default_date = timezone.localdate().isoformat()
        first_clase = clases[0] if clases else None
        if first_clase:
            calendar_default_date = first_clase.fecha.isoformat()

        for clase in clases:
            horario = clase.horario_aula_curso.horario_dia.horario
            materia = clase.materia_curso.materia
            aula = clase.horario_aula_curso.aula_curso.aula
            grupo = clase.materia_curso.grupo
            event_color = materia.color
            calendar_events.append(
                {
                    "id": clase.pk,
                    "title": f"{grupo.nombre} · {aula} · {materia.nombre}",
                    "start": f"{clase.fecha.isoformat()}T{horario.hora_inicio:%H:%M:%S}",
                    "end": f"{clase.fecha.isoformat()}T{horario.hora_fin:%H:%M:%S}",
                    "allDay": False,
                    "className": f"materia-event docente-event estado-{clase.estado_planificacion}",
                    "backgroundColor": event_color,
                    "borderColor": event_color,
                    "textColor": readable_text_color(event_color),
                    "grupo": grupo.nombre,
                    "aula": str(aula),
                    "materia": materia.nombre,
                    "hora": f"{horario.hora_inicio:%H:%M} - {horario.hora_fin:%H:%M}",
                    "estado": clase.get_estado_planificacion_display(),
                    "url": str(reverse_lazy("academico:docente_clase_planificar", kwargs={"pk": clase.pk})),
                }
            )

        return render(
            request,
            self.template_name,
            {
                "title": "Mis planificaciones",
                "docente": docente,
                "calendar_events_json": json.dumps(calendar_events),
                "calendar_default_date": calendar_default_date,
                "has_events": bool(calendar_events),
                "planificacion_stats": stats,
                "status_tabs": status_tabs,
                "selected_filter": selected_filter,
                "selected_filter_label": self.get_filter_label(selected_filter),
                "tema_cards": tema_cards,
                "planificacion_cards": planificacion_cards,
            },
        )

    def empty_stats(self):
        return {
            "total": 0,
            "pendiente": 0,
            "revision": 0,
            "rechazada": 0,
            "aprobada": 0,
            "por_atender": 0,
            "avance": 0,
        }

    def get_planificacion_stats(self, clases):
        stats = self.empty_stats()
        stats["total"] = len(clases)
        for clase in clases:
            if clase.estado_planificacion in stats:
                stats[clase.estado_planificacion] += 1
        stats["por_atender"] = stats["pendiente"] + stats["rechazada"]
        stats["avance"] = round((stats["aprobada"] / stats["total"]) * 100) if stats["total"] else 0
        return stats

    def get_selected_filter(self, requested_filter, stats):
        valid_filters = {item["key"] for item in self.status_filters}
        if requested_filter in valid_filters:
            return requested_filter
        if stats["por_atender"]:
            return "trabajo"
        if stats["revision"]:
            return "revision"
        return "todas"

    def get_filter_label(self, selected_filter):
        filter_item = next((item for item in self.status_filters if item["key"] == selected_filter), None)
        return filter_item["label"] if filter_item else ""

    def get_status_tabs(self, selected_filter, stats):
        base_url = reverse_lazy("academico:docente_horarios")
        tabs = []
        for item in self.status_filters:
            count = stats["total"] if item["states"] is None else sum(stats[state] for state in item["states"])
            tabs.append(
                {
                    **item,
                    "count": count,
                    "url": f"{base_url}?{urlencode({'estado': item['key']})}",
                    "is_active": item["key"] == selected_filter,
                }
            )
        return tabs

    def filter_clases(self, clases, selected_filter):
        filter_item = next((item for item in self.status_filters if item["key"] == selected_filter), None)
        if not filter_item or filter_item["states"] is None:
            return clases
        allowed_states = set(filter_item["states"])
        return [clase for clase in clases if clase.estado_planificacion in allowed_states]

    def get_docente_planificaciones_tema_queryset(self, docente):
        return (
            PlanificacionTema.objects.select_related(
                "profesor_materia_curso__materia_curso__materia",
                "profesor_materia_curso__materia_curso__grupo",
                "tema",
                "tema__planificacion",
            )
            .prefetch_related("tema__subtemas_planificacion")
            .filter(profesor_materia_curso__partner=docente)
            .order_by(
                "profesor_materia_curso__materia_curso__grupo__nombre",
                "profesor_materia_curso__materia_curso__materia__nombre",
                "tema__orden",
                "tema__nombre",
            )
        )

    def topic_matches_filter(self, card, selected_filter):
        if selected_filter == "todas":
            return True
        if selected_filter == "trabajo":
            return card["work_count"] > 0
        return card["stats"].get(selected_filter, 0) > 0

    def build_tema_card(self, planificacion_tema, clases):
        tema = planificacion_tema.tema
        materia_curso = planificacion_tema.profesor_materia_curso.materia_curso
        subtemas = list(tema.subtemas_planificacion.all())
        topic_classes = [
            clase
            for clase in clases
            if clase.materia_curso_id == materia_curso.pk and clase.tema_id == tema.pk
        ]
        covered_subtema_ids = clase_subtema_ids(topic_classes, tema)
        topic_subtema_ids = {subtema.pk for subtema in subtemas}
        covered_subtema_count = len(covered_subtema_ids & topic_subtema_ids)
        pending_subtema_count = max(len(subtemas) - covered_subtema_count, 0)
        available_classes = [
            clase
            for clase in clases
            if (
                clase.materia_curso_id == materia_curso.pk
                and not clase.tema_id
                and clase.estado_planificacion not in CLASS_ASSIGNMENT_LOCK_STATES
            )
        ]
        assignable_classes = available_classes if not subtemas or pending_subtema_count else []
        stats = self.empty_topic_stats()
        stats["total"] = len(topic_classes)
        for clase in topic_classes:
            if clase.estado_planificacion in stats:
                stats[clase.estado_planificacion] += 1
        work_count = stats["pendiente"] + stats["rechazada"] + len(assignable_classes)
        status = self.get_topic_status(stats, assignable_classes)
        upcoming_classes = sorted(
            assignable_classes + [clase for clase in topic_classes if clase.estado_planificacion in {"pendiente", "rechazada"}],
            key=lambda item: (item.fecha, item.horario_aula_curso.horario_dia.horario.hora_inicio),
        )[:3]
        return {
            "tema": tema,
            "planificacion_tema": planificacion_tema,
            "materia": materia_curso.materia,
            "grupo": materia_curso.grupo,
            "detalle": tema.detalle,
            "subtemas": subtemas,
            "covered_subtema_count": covered_subtema_count,
            "pending_subtema_count": pending_subtema_count,
            "temario_progress": round((covered_subtema_count / len(subtemas)) * 100) if subtemas else 0,
            "available_count": len(assignable_classes),
            "assigned_count": len(topic_classes),
            "work_count": work_count,
            "stats": stats,
            "status": status,
            "status_label": self.get_topic_status_label(status),
            "progress": round((stats["aprobada"] / stats["total"]) * 100) if stats["total"] else 0,
            "upcoming_classes": [
                {
                    "clase": clase,
                    "horario": clase.horario_aula_curso.horario_dia.horario,
                    "aula": clase.horario_aula_curso.aula_curso.aula,
                    "is_available": not clase.tema_id,
                }
                for clase in upcoming_classes
            ],
            "url": reverse_lazy("academico:docente_tema_planificar", kwargs={"pk": planificacion_tema.pk}),
        }

    def empty_topic_stats(self):
        return {
            "total": 0,
            "pendiente": 0,
            "revision": 0,
            "rechazada": 0,
            "aprobada": 0,
        }

    def get_topic_status(self, stats, available_classes):
        if stats["rechazada"]:
            return "rechazada"
        if stats["revision"]:
            return "revision"
        if stats["pendiente"] or available_classes:
            return "pendiente"
        if stats["aprobada"]:
            return "aprobada"
        return "pendiente"

    def get_topic_status_label(self, status):
        labels = dict(Clase.ESTADO_PLANIFICACION_CHOICES)
        if status == "pendiente":
            return "Por planificar"
        return labels.get(status, status)

    def build_planificacion_card(self, clase):
        horario = clase.horario_aula_curso.horario_dia.horario
        competencias = list(clase.competencias.all())
        estrategias = list(clase.estrategias.all())
        recursos = list(clase.recursos.all())
        subtemas = clase.get_subtemas_planificados()
        observaciones = clase.observaciones_revision or {}
        steps = [
            {"label": "Tema", "done": bool(clase.tema_id and (subtemas or not clase.tema.subtemas_planificacion.exists())), "note": observaciones.get("tema", "")},
            {"label": "Competencias", "done": bool(competencias), "note": observaciones.get("competencias", "")},
            {"label": "Estrategias", "done": bool(estrategias), "note": observaciones.get("estrategias", "")},
            {"label": "Recursos", "done": bool(recursos), "note": observaciones.get("recursos", "")},
        ]
        completed_steps = sum(1 for item in steps if item["done"])
        action_labels = {
            "pendiente": "Planificar",
            "revision": "Ver envio",
            "rechazada": "Corregir",
            "aprobada": "Ver",
        }
        return {
            "clase": clase,
            "horario": horario,
            "estado": clase.estado_planificacion,
            "estado_label": clase.get_estado_planificacion_display(),
            "aula": clase.horario_aula_curso.aula_curso.aula,
            "grupo": clase.materia_curso.grupo,
            "materia": clase.materia_curso.materia,
            "tema": clase.tema,
            "subtema": clase.subtema,
            "subtemas": subtemas,
            "steps": steps,
            "completed_steps": completed_steps,
            "progress": round((completed_steps / len(steps)) * 100),
            "revision_note": clase.notas_revision if clase.estado_planificacion == "rechazada" else "",
            "is_late": clase.fecha < timezone.localdate() and clase.estado_planificacion in {"pendiente", "rechazada"},
            "action_label": action_labels.get(clase.estado_planificacion, "Abrir"),
            "url": reverse_lazy("academico:docente_clase_planificar", kwargs={"pk": clase.pk}),
            "can_take_attendance": clase.fecha == timezone.localdate() and not clase.asistencia_cerrada,
            "attendance_closed": clase.asistencia_cerrada,
            "attendance_url": reverse_lazy("academico:docente_clase_asistencia", kwargs={"pk": clase.pk}),
        }


class DocenteCalendarioMixin:
    weekday_headers = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]

    def get_docente(self):
        docente = getattr(self.request.user, "partner", None)
        if docente and docente.es_docente:
            return docente
        return None

    def get_selected_date(self):
        selected = (self.request.GET.get("fecha") or "").strip()
        if selected:
            try:
                return date.fromisoformat(selected)
            except ValueError:
                pass
        return timezone.localdate()

    def get_week_bounds(self, selected_date):
        week_start = selected_date - timedelta(days=selected_date.weekday())
        return week_start, week_start + timedelta(days=6)

    def get_docente_clases_queryset(self, docente):
        return (
            Clase.objects.select_related(
                "materia_curso__materia",
                "materia_curso__grupo",
                "docente",
                "tema",
                "subtema",
                "horario_aula_curso__aula_curso__aula",
                "horario_aula_curso__horario_dia__dia",
                "horario_aula_curso__horario_dia__horario",
            )
            .prefetch_related("clase_subtemas__subtema")
            .filter(docente_responsable_filter(docente))
            .distinct()
            .order_by(
                "fecha",
                "horario_aula_curso__horario_dia__horario__hora_inicio",
                "materia_curso__grupo__nombre",
                "materia_curso__materia__nombre",
            )
        )

    def build_schedule_row(self, clase):
        horario = clase.horario_aula_curso.horario_dia.horario
        materia = clase.materia_curso.materia
        grupo = clase.materia_curso.grupo
        aula = clase.horario_aula_curso.aula_curso.aula
        return {
            "clase": clase,
            "fecha": clase.fecha,
            "weekday": clase.fecha.weekday(),
            "time_key": (horario.hora_inicio, horario.hora_fin),
            "hora": f"{horario.hora_inicio:%H:%M} - {horario.hora_fin:%H:%M}",
            "materia": materia.nombre,
            "materia_corta": materia.nombre_corto or materia.nombre,
            "grupo": grupo.nombre,
            "aula": str(aula),
            "tema": str(clase.tema) if clase.tema else "",
            "subtema": clase.get_subtemas_label(),
            "estado": clase.get_estado_planificacion_display(),
            "color": materia.color,
            "url": str(reverse_lazy("academico:docente_clase_planificar", kwargs={"pk": clase.pk})),
        }

    def build_calendar_event(self, clase):
        row = self.build_schedule_row(clase)
        start_time, end_time = row["time_key"]
        event_color = row["color"]
        title = f"{row['materia_corta']} · {row['grupo']}"
        return {
            "id": clase.pk,
            "title": title,
            "start": f"{row['fecha'].isoformat()}T{start_time:%H:%M:%S}",
            "end": f"{row['fecha'].isoformat()}T{end_time:%H:%M:%S}",
            "allDay": False,
            "className": f"materia-event docente-schedule-event estado-{clase.estado_planificacion}",
            "backgroundColor": event_color,
            "borderColor": event_color,
            "textColor": readable_text_color(event_color),
            "materia": row["materia"],
            "grupo": row["grupo"],
            "aula": row["aula"],
            "hora": row["hora"],
            "tema": row["tema"],
            "subtema": row["subtema"],
            "estado": row["estado"],
            "url": row["url"],
        }


class DocenteCalendarioView(LoginRequiredMixin, DocenteCalendarioMixin, View):
    template_name = "academico/docente_calendario.html"

    def get(self, request):
        docente = self.get_docente()
        selected_date = self.get_selected_date()
        week_start, week_end = self.get_week_bounds(selected_date)
        export_base_url = reverse_lazy("academico:docente_calendario_exportar")
        export_url = f"{export_base_url}?{urlencode({'fecha': selected_date.isoformat()})}"
        context = {
            "title": "Mi calendario",
            "docente": docente,
            "selected_date": selected_date,
            "week_start": week_start,
            "week_end": week_end,
            "calendar_events_json": json.dumps([]),
            "calendar_default_date": selected_date.isoformat(),
            "has_events": False,
            "week_count": 0,
            "total_count": 0,
            "materia_count": 0,
            "grupo_count": 0,
            "export_base_url": export_base_url,
            "export_url": export_url,
        }
        if not docente:
            return render(request, self.template_name, context)

        clases = list(self.get_docente_clases_queryset(docente))
        week_rows = [self.build_schedule_row(clase) for clase in clases if week_start <= clase.fecha <= week_end]
        context.update(
            {
                "calendar_events_json": json.dumps([self.build_calendar_event(clase) for clase in clases]),
                "has_events": bool(clases),
                "week_count": len(week_rows),
                "total_count": len(clases),
                "materia_count": len({row["materia"] for row in week_rows}),
                "grupo_count": len({row["grupo"] for row in week_rows}),
            }
        )
        return render(request, self.template_name, context)


class DocenteCalendarioExportView(LoginRequiredMixin, DocenteCalendarioMixin, View):
    def get(self, request):
        docente = self.get_docente()
        if not docente:
            return HttpResponse("Tu usuario no esta vinculado a un docente activo.", status=403)

        selected_date = self.get_selected_date()
        week_start, week_end = self.get_week_bounds(selected_date)
        rows = [
            self.build_schedule_row(clase)
            for clase in self.get_docente_clases_queryset(docente).filter(fecha__range=(week_start, week_end))
        ]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = safe_sheet_title("Mi horario")
        self.write_week_sheet(sheet, docente, week_start, week_end, rows)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        filename = f"mi_horario_{week_start:%Y%m%d}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def write_week_sheet(self, sheet, docente, week_start, week_end, rows):
        title_fill = PatternFill("solid", fgColor="DCEBFF")
        header_fill = PatternFill("solid", fgColor="0F766E")
        time_fill = PatternFill("solid", fgColor="134E4A")
        empty_fill = PatternFill("solid", fgColor="FFFFFF")
        thin = Side(style="thin", color="D7DEE8")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        schedule_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        title_cell = sheet.cell(row=1, column=1, value=f"Horario semanal - {docente.nombre}")
        title_cell.font = Font(bold=True, size=14, color="111827")
        title_cell.fill = title_fill
        title_cell.alignment = center

        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        week_cell = sheet.cell(row=2, column=1, value=f"Semana {week_start:%d/%m/%Y} - {week_end:%d/%m/%Y}")
        week_cell.font = Font(bold=True, color="111827")
        week_cell.fill = title_fill
        week_cell.alignment = center

        sheet.cell(row=4, column=1, value="Hora")
        for col, day in enumerate(self.weekday_headers, start=2):
            day_date = week_start + timedelta(days=col - 2)
            sheet.cell(row=4, column=col, value=f"{day}\n{day_date:%d/%m}")
        for col in range(1, 9):
            cell = sheet.cell(row=4, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center

        time_slots = sorted({row["time_key"] for row in rows})
        week_matrix = {}
        for row in rows:
            week_matrix.setdefault((row["time_key"], row["weekday"]), []).append(row)

        current_row = 5
        if not time_slots:
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            cell = sheet.cell(row=current_row, column=1, value="Sin clases asignadas en esta semana.")
            cell.alignment = center
            cell.border = border
        for time_slot in time_slots:
            start_time, end_time = time_slot
            time_cell = sheet.cell(row=current_row, column=1, value=f"{start_time:%H:%M} - {end_time:%H:%M}")
            time_cell.font = Font(bold=True, color="FFFFFF")
            time_cell.fill = time_fill
            time_cell.border = border
            time_cell.alignment = center
            max_lines = 1
            for weekday in range(7):
                cell = sheet.cell(row=current_row, column=weekday + 2)
                entries = week_matrix.get((time_slot, weekday), [])
                if entries:
                    cell.value = "\n\n".join(self.schedule_label(row) for row in entries)
                    cell.fill = PatternFill("solid", fgColor=xlsx_color(entries[0]["color"]))
                    cell.font = Font(color=xlsx_color(readable_text_color(entries[0]["color"])), bold=True, size=9)
                    max_lines = max(max_lines, estimated_schedule_wrapped_lines(cell.value))
                else:
                    cell.value = ""
                    cell.fill = empty_fill
                cell.border = border
                cell.alignment = schedule_alignment
            sheet.row_dimensions[current_row].height = schedule_export_row_height(max_lines)
            current_row += 1

        widths = [18] + [SCHEDULE_EXPORT_COLUMN_WIDTH] * 7
        for col, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col)].width = width
        sheet.freeze_panes = "B5"

    def schedule_label(self, row):
        parts = [row["materia"], row["grupo"], row["aula"]]
        if row["tema"]:
            parts.append(row["tema"])
        if row["subtema"]:
            parts.append(row["subtema"])
        return "\n".join(parts)


class DocenteTemaPlanificacionView(LoginRequiredMixin, View):
    template_name = "academico/docente_tema_planificacion.html"

    def get_docente(self):
        docente = getattr(self.request.user, "partner", None)
        if docente and docente.es_docente:
            return docente
        return None

    def get_planificacion_tema(self):
        docente = self.get_docente()
        if not docente:
            return get_object_or_404(PlanificacionTema.objects.none(), pk=self.kwargs["pk"])
        return get_object_or_404(
            PlanificacionTema.objects.select_related(
                "profesor_materia_curso__materia_curso__materia",
                "profesor_materia_curso__materia_curso__grupo",
                "tema",
                "tema__planificacion",
            ).prefetch_related("tema__subtemas_planificacion"),
            pk=self.kwargs["pk"],
            profesor_materia_curso__partner=docente,
        )

    def get(self, request, pk):
        docente = self.get_docente()
        planificacion_tema = self.get_planificacion_tema()
        return render(request, self.template_name, self.get_context(docente, planificacion_tema))

    def post(self, request, pk):
        docente = self.get_docente()
        planificacion_tema = self.get_planificacion_tema()
        action = request.POST.get("tema_action")
        if action == "assign":
            return self.assign_class(request, docente, planificacion_tema)
        if action == "unassign":
            return self.unassign_class(request, docente, planificacion_tema)
        if action in {"save_class", "send_class"}:
            return self.save_class_plan(request, docente, planificacion_tema, send=action == "send_class")
        messages.error(request, "Accion no valida.")
        return self.redirect_to_topic(planificacion_tema)

    def redirect_to_topic(self, planificacion_tema, clase=None):
        url = str(reverse_lazy("academico:docente_tema_planificar", kwargs={"pk": planificacion_tema.pk}))
        if clase:
            url = f"{url}#clase-{clase.pk}"
        return redirect(url)

    def assign_class(self, request, docente, planificacion_tema):
        tema = planificacion_tema.tema
        clase = self.get_topic_class(request.POST.get("clase_id"), docente, planificacion_tema)
        if not clase:
            messages.error(request, "La clase seleccionada no esta disponible para este tema.")
            return self.redirect_to_topic(planificacion_tema)
        if clase.estado_planificacion in CLASS_ASSIGNMENT_LOCK_STATES:
            messages.error(request, "La clase ya fue enviada o aprobada y no se puede reasignar.")
            return self.redirect_to_topic(planificacion_tema)
        if clase.tema_id and clase.tema_id != tema.pk:
            messages.error(request, "La clase ya esta asignada a otro tema.")
            return self.redirect_to_topic(planificacion_tema)

        current_assigned_classes = [
            item
            for item in self.get_clases_queryset(docente, planificacion_tema)
            if item.tema_id == tema.pk and item.pk != clase.pk
        ]
        used_subtema_ids = self.get_used_subtema_ids(docente, planificacion_tema, exclude_clase=clase)
        assignment_blocker = self.get_assignment_blocker(tema, current_assigned_classes, used_subtema_ids)
        if assignment_blocker:
            messages.info(request, assignment_blocker)
            return self.redirect_to_topic(planificacion_tema)

        selected_subtemas, invalid_subtemas = self.get_subtemas_from_request(request, tema)
        if invalid_subtemas:
            messages.error(request, "Uno o mas subtemas seleccionados no pertenecen al tema.")
            return self.redirect_to_topic(planificacion_tema)
        if any(subtema.pk in used_subtema_ids for subtema in selected_subtemas):
            messages.error(request, "Uno o mas subtemas ya estan asignados a otra clase del tema.")
            return self.redirect_to_topic(planificacion_tema)

        clase.tema = tema
        clase.save(update_fields=["tema"])
        clase.sync_subtemas_planificados(selected_subtemas)
        messages.success(request, "Clase agregada al tema. Ahora completa su planificacion.")
        return self.redirect_to_topic(planificacion_tema, clase)

    def unassign_class(self, request, docente, planificacion_tema):
        tema = planificacion_tema.tema
        clase = self.get_topic_class(request.POST.get("clase_id"), docente, planificacion_tema)
        if not clase or clase.tema_id != tema.pk:
            messages.error(request, "La clase no pertenece a este tema.")
            return self.redirect_to_topic(planificacion_tema)
        if clase.estado_planificacion in CLASS_ASSIGNMENT_LOCK_STATES:
            messages.error(request, "La clase ya fue enviada o aprobada y no se puede liberar.")
            return self.redirect_to_topic(planificacion_tema, clase)
        clase.tema = None
        clase.save(update_fields=["tema"])
        clase.sync_subtemas_planificados([])
        messages.success(request, "La clase quedo disponible para otro tema.")
        return self.redirect_to_topic(planificacion_tema)

    def save_class_plan(self, request, docente, planificacion_tema, send=False):
        tema = planificacion_tema.tema
        clase = self.get_topic_class(request.POST.get("clase_id"), docente, planificacion_tema)
        if not clase or clase.tema_id != tema.pk:
            messages.error(request, "La clase no pertenece a este tema.")
            return self.redirect_to_topic(planificacion_tema)
        if clase.estado_planificacion == "aprobada":
            messages.error(request, "La planificacion aprobada no se puede editar desde el perfil docente.")
            return self.redirect_to_topic(planificacion_tema, clase)

        selected_subtemas, invalid_subtemas = self.get_subtemas_from_request(request, tema)
        if invalid_subtemas:
            messages.error(request, "Uno o mas subtemas seleccionados no pertenecen al tema.")
            return self.redirect_to_topic(planificacion_tema, clase)
        used_subtema_ids = self.get_used_subtema_ids(docente, planificacion_tema, exclude_clase=clase)
        if any(subtema.pk in used_subtema_ids for subtema in selected_subtemas):
            messages.error(request, "Uno o mas subtemas ya estan asignados a otra clase del tema.")
            return self.redirect_to_topic(planificacion_tema, clase)

        selected_subtemas, new_subtema_names = self.resolve_inline_subtemas(
            tema,
            selected_subtemas,
            self.get_inline_new_subtema_names(request),
        )
        if any(subtema.pk in used_subtema_ids for subtema in selected_subtemas):
            messages.error(request, "Uno o mas subtemas ya estan asignados a otra clase del tema.")
            return self.redirect_to_topic(planificacion_tema, clase)

        ready_errors = self.get_inline_ready_errors(request, tema, selected_subtemas, new_subtema_names) if send else []
        if ready_errors:
            for error in ready_errors:
                messages.error(request, error)
            return self.redirect_to_topic(planificacion_tema, clase)

        with transaction.atomic():
            created_subtemas = self.create_inline_subtemas(tema, new_subtema_names)
            clase.tema = tema
            clase.save(update_fields=["tema"])
            clase.sync_subtemas_planificados([*selected_subtemas, *created_subtemas])
            self.sync_inline_tags(clase.competencias, Competencia, "competencias")
            self.sync_inline_tags(clase.estrategias, Estrategia, "estrategias")
            self.sync_inline_tags(clase.recursos, Recurso, "recursos")
            if send:
                clase.estado_planificacion = "revision"
                clase.notas_revision = ""
                clase.observaciones_revision = {}
                clase.revisado_por = None
                clase.fecha_revision = None
                clase.revision_tema_ok = False
                clase.revision_detalle_ok = False
                clase.revision_competencias_ok = False
                clase.revision_estrategias_ok = False
                clase.revision_recursos_ok = False
                clase.save(update_fields=[
                    "estado_planificacion",
                    "notas_revision",
                    "observaciones_revision",
                    "revisado_por",
                    "fecha_revision",
                    "revision_tema_ok",
                    "revision_detalle_ok",
                    "revision_competencias_ok",
                    "revision_estrategias_ok",
                    "revision_recursos_ok",
                ])
            elif clase.estado_planificacion == "revision":
                clase.estado_planificacion = "pendiente"
                clase.save(update_fields=["estado_planificacion"])

        if send:
            messages.success(request, "Planificacion de la clase enviada a revision.")
        else:
            messages.success(request, "Planificacion de la clase guardada.")
        return self.redirect_to_topic(planificacion_tema, clase)

    def get_inline_ready_errors(self, request, tema, selected_subtemas, new_subtema_names=None):
        new_subtema_names = new_subtema_names or []
        errors = []
        if tema.subtemas_planificacion.exists() and not selected_subtemas and not new_subtema_names:
            errors.append("Selecciona al menos un subtema de la clase.")
        if not self.has_inline_tag_items(request, "competencias"):
            errors.append("Selecciona o agrega al menos una competencia.")
        if not self.has_inline_tag_items(request, "estrategias"):
            errors.append("Selecciona o agrega al menos una estrategia.")
        if not self.has_inline_tag_items(request, "recursos"):
            errors.append("Selecciona o agrega al menos un recurso.")
        return errors

    def has_inline_tag_items(self, request, prefix):
        return bool(self.get_inline_selected_ids(request, prefix) or self.get_inline_new_names(request, prefix))

    def get_inline_selected_ids(self, request, prefix):
        return {
            int(value)
            for value in request.POST.getlist(f"{prefix}_existentes")
            if str(value).isdigit()
        }

    def get_inline_new_names(self, request, prefix):
        raw_value = request.POST.get(f"{prefix}_nuevos") or ""
        names = []
        for value in raw_value.replace("\n", ",").split(","):
            name = value.strip()
            if name and name not in names:
                names.append(name)
        return names

    def get_inline_new_subtema_names(self, request):
        raw_value = request.POST.get("subtemas_nuevos") or ""
        names = []
        for value in raw_value.replace("\n", ",").split(","):
            name = value.strip()
            if name and name.lower() not in [item.lower() for item in names]:
                names.append(name)
        return names

    def resolve_inline_subtemas(self, tema, selected_subtemas, new_subtema_names):
        resolved = []
        resolved_ids = set()
        pending_names = []
        for subtema in selected_subtemas:
            if subtema.pk not in resolved_ids:
                resolved.append(subtema)
                resolved_ids.add(subtema.pk)
        for name in new_subtema_names:
            existing = Subtema.objects.filter(tema=tema, nombre__iexact=name).first()
            if existing:
                if existing.pk not in resolved_ids:
                    resolved.append(existing)
                    resolved_ids.add(existing.pk)
                continue
            if name.lower() not in [item.lower() for item in pending_names]:
                pending_names.append(name)
        return resolved, pending_names

    def create_inline_subtemas(self, tema, names):
        if not names:
            return []
        next_order = (Subtema.objects.filter(tema=tema).aggregate(max_order=Max("orden"))["max_order"] or 0) + 1
        created = []
        for offset, name in enumerate(names):
            created.append(Subtema.objects.create(tema=tema, nombre=name, orden=next_order + offset))
        return created

    def sync_inline_tags(self, relation, model, prefix):
        selected_ids = self.get_inline_selected_ids(self.request, prefix)
        for name in self.get_inline_new_names(self.request, prefix):
            obj, _ = model.objects.get_or_create(nombre=name)
            selected_ids.add(obj.pk)
        relation.set(selected_ids)

    def get_topic_class(self, clase_id, docente, planificacion_tema):
        if not str(clase_id or "").isdigit():
            return None
        return self.get_clases_queryset(docente, planificacion_tema).filter(pk=clase_id).first()

    def get_subtemas_from_request(self, request, tema):
        subtema_ids = request.POST.getlist("subtema_ids")
        legacy_subtema_id = request.POST.get("subtema_id")
        if not subtema_ids and legacy_subtema_id:
            subtema_ids = [legacy_subtema_id]
        parsed_ids = []
        for value in subtema_ids:
            if not str(value or "").isdigit():
                return [], True
            subtema_id = int(value)
            if subtema_id not in parsed_ids:
                parsed_ids.append(subtema_id)
        if not parsed_ids:
            return [], False
        subtemas = Subtema.objects.filter(pk__in=parsed_ids, tema=tema)
        subtemas_by_id = {subtema.pk: subtema for subtema in subtemas}
        if len(subtemas_by_id) != len(parsed_ids):
            return [], True
        return [subtemas_by_id[subtema_id] for subtema_id in parsed_ids], False

    def get_used_subtema_ids(self, docente, planificacion_tema, exclude_clase=None):
        clases = self.get_clases_queryset(docente, planificacion_tema)
        return clase_subtema_ids(clases, planificacion_tema.tema, exclude_clase=exclude_clase)

    def get_clases_queryset(self, docente, planificacion_tema):
        materia_curso = planificacion_tema.profesor_materia_curso.materia_curso
        return (
            Clase.objects.select_related(
                "materia_curso__materia",
                "materia_curso__grupo",
                "docente",
                "tema",
                "subtema",
                "horario_aula_curso__aula_curso__aula",
                "horario_aula_curso__horario_dia__horario",
            )
            .prefetch_related("competencias", "estrategias", "recursos", "clase_subtemas__subtema")
            .filter(materia_curso=materia_curso)
            .filter(docente_responsable_filter(docente))
            .distinct()
            .order_by("fecha", "horario_aula_curso__horario_dia__horario__hora_inicio")
        )

    def get_context(self, docente, planificacion_tema):
        tema = planificacion_tema.tema
        materia_curso = planificacion_tema.profesor_materia_curso.materia_curso
        periodo = self.get_planificacion_periodo(materia_curso)
        clases = list(self.get_clases_queryset(docente, planificacion_tema)) if docente else []
        assigned_classes = [clase for clase in clases if clase.tema_id == tema.pk]
        available_classes = [
            clase
            for clase in clases
            if not clase.tema_id and clase.estado_planificacion not in CLASS_ASSIGNMENT_LOCK_STATES
        ]
        occupied_classes = [
            clase
            for clase in clases
            if clase.tema_id and clase.tema_id != tema.pk
        ]
        status_counts = {
            "pendiente": 0,
            "revision": 0,
            "rechazada": 0,
            "aprobada": 0,
        }
        for clase in assigned_classes:
            if clase.estado_planificacion in status_counts:
                status_counts[clase.estado_planificacion] += 1
        subtemas = list(tema.subtemas_planificacion.all().order_by("orden", "nombre"))
        subtema_ids = {subtema.pk for subtema in subtemas}
        covered_subtema_ids = clase_subtema_ids(assigned_classes, tema) & subtema_ids
        available_subtemas = [subtema for subtema in subtemas if subtema.pk not in covered_subtema_ids]
        pending_subtema_count = max(len(subtemas) - len(covered_subtema_ids), 0)
        assignment_blocker = self.get_assignment_blocker(tema, assigned_classes, covered_subtema_ids)
        assignable_classes = available_classes if not assignment_blocker else []
        tag_catalogs = self.get_inline_tag_catalogs()
        return {
            "title": "Planificar tema",
            "docente": docente,
            "planificacion_tema": planificacion_tema,
            "tema": tema,
            "materia_curso": materia_curso,
            "periodo": periodo,
            "subtemas": subtemas,
            "available_subtemas": available_subtemas,
            "covered_subtema_count": len(covered_subtema_ids),
            "pending_subtema_count": pending_subtema_count,
            "temario_progress": round((len(covered_subtema_ids) / len(subtemas)) * 100) if subtemas else 0,
            "available_classes": [self.build_class_slot(clase, tema, subtemas, covered_subtema_ids, tag_catalogs) for clase in assignable_classes],
            "assigned_classes": [self.build_class_slot(clase, tema, subtemas, covered_subtema_ids, tag_catalogs) for clase in assigned_classes],
            "occupied_classes": [self.build_class_slot(clase, tema, subtemas, covered_subtema_ids, tag_catalogs) for clase in occupied_classes],
            "status_counts": status_counts,
            "assigned_count": len(assigned_classes),
            "available_count": len(assignable_classes),
            "raw_available_count": len(available_classes),
            "occupied_count": len(occupied_classes),
            "assignment_blocker": assignment_blocker,
            "can_add_class": not assignment_blocker and bool(assignable_classes),
            "approved_progress": round((status_counts["aprobada"] / len(assigned_classes)) * 100) if assigned_classes else 0,
            "list_url": reverse_lazy("academico:docente_horarios"),
        }

    def get_assignment_blocker(self, tema, assigned_classes, covered_subtema_ids):
        unfinished = [
            clase
            for clase in assigned_classes
            if clase.estado_planificacion in {"pendiente", "rechazada"}
        ]
        if unfinished:
            return "Envia a revision la clase agregada antes de tomar otra clase para este tema."
        subtema_ids = set(tema.subtemas_planificacion.values_list("id", flat=True))
        if subtema_ids and not (subtema_ids - set(covered_subtema_ids)):
            return "Todos los subtemas del tema ya fueron planificados."
        return ""

    def get_inline_tag_catalogs(self):
        return {
            "competencias": list(Competencia.objects.order_by("nombre")),
            "estrategias": list(Estrategia.objects.order_by("nombre")),
            "recursos": list(Recurso.objects.order_by("nombre")),
        }

    def get_planificacion_periodo(self, materia_curso):
        periodos = CursoPeriodo.objects.select_related("periodo").filter(curso=materia_curso.grupo)
        today = timezone.localdate()
        active = periodos.filter(periodo__fecha_inicio__lte=today, periodo__fecha_fin__gte=today).order_by(
            "-periodo__fecha_inicio"
        ).first()
        if active:
            return active.periodo
        latest = periodos.order_by("-periodo__fecha_inicio").first()
        return latest.periodo if latest else None

    def build_class_slot(self, clase, tema, subtemas=None, covered_subtema_ids=None, tag_catalogs=None):
        subtemas = subtemas or []
        covered_subtema_ids = covered_subtema_ids or set()
        tag_catalogs = tag_catalogs or self.get_inline_tag_catalogs()
        horario = clase.horario_aula_curso.horario_dia.horario
        competencias = list(clase.competencias.all())
        estrategias = list(clase.estrategias.all())
        recursos = list(clase.recursos.all())
        selected_subtemas = clase.get_subtemas_planificados()
        selected_subtema_ids = {subtema.pk for subtema in selected_subtemas}
        visible_subtemas = [
            subtema
            for subtema in subtemas
            if subtema.pk in selected_subtema_ids or subtema.pk not in covered_subtema_ids
        ]
        observaciones = clase.observaciones_revision or {}
        steps = [
            {"label": "Tema", "done": bool(clase.tema_id and (selected_subtemas or not subtemas)), "note": observaciones.get("tema", "")},
            {"label": "Competencias", "done": bool(competencias), "note": observaciones.get("competencias", "")},
            {"label": "Estrategias", "done": bool(estrategias), "note": observaciones.get("estrategias", "")},
            {"label": "Recursos", "done": bool(recursos), "note": observaciones.get("recursos", "")},
        ]
        completed_steps = sum(1 for step in steps if step["done"])
        url = str(reverse_lazy("academico:docente_clase_planificar", kwargs={"pk": clase.pk}))
        if clase.tema_id == tema.pk:
            url = f"{url}?{urlencode({'from_planificacion_tema': self.kwargs['pk']})}"
        return {
            "clase": clase,
            "horario": horario,
            "aula": clase.horario_aula_curso.aula_curso.aula,
            "estado": clase.estado_planificacion,
            "estado_label": clase.get_estado_planificacion_display(),
            "tema": clase.tema,
            "subtema": clase.subtema,
            "subtemas": selected_subtemas,
            "subtemas_label": clase.get_subtemas_label(),
            "subtema_options": [
                {
                    "subtema": subtema,
                    "selected": subtema.pk in selected_subtema_ids,
                    "available": True,
                }
                for subtema in visible_subtemas
            ],
            "tag_groups": self.build_inline_tag_groups(clase, tag_catalogs),
            "steps": steps,
            "progress": round((completed_steps / len(steps)) * 100),
            "url": url,
            "action_label": self.get_class_action_label(clase),
            "is_late": clase.fecha < timezone.localdate() and clase.estado_planificacion in {"pendiente", "rechazada"},
            "is_locked": clase.estado_planificacion in CLASS_ASSIGNMENT_LOCK_STATES,
            "can_edit": clase.estado_planificacion != "aprobada",
            "can_unassign": clase.tema_id == tema.pk and clase.estado_planificacion not in CLASS_ASSIGNMENT_LOCK_STATES,
            "revision_note": clase.notas_revision if clase.estado_planificacion == "rechazada" else "",
        }

    def build_inline_tag_groups(self, clase, tag_catalogs):
        specs = (
            ("Competencias", "competencias", "ri-medal-line", set(clase.competencias.values_list("id", flat=True)), "Nueva competencia"),
            ("Estrategias", "estrategias", "ri-route-line", set(clase.estrategias.values_list("id", flat=True)), "Nueva estrategia"),
            ("Recursos", "recursos", "ri-attachment-2", set(clase.recursos.values_list("id", flat=True)), "Nuevo recurso"),
        )
        groups = []
        for title, prefix, icon, selected_ids, placeholder in specs:
            options = [
                {
                    "obj": item,
                    "selected": item.pk in selected_ids,
                }
                for item in tag_catalogs[prefix]
            ]
            groups.append(
                {
                    "title": title,
                    "prefix": prefix,
                    "icon": icon,
                    "placeholder": placeholder,
                    "selected_count": len(selected_ids),
                    "options": options,
                    "selected_options": [item for item in options if item["selected"]],
                }
            )
        return groups

    def get_class_action_label(self, clase):
        return {
            "pendiente": "Editar clase",
            "revision": "Ver clase",
            "rechazada": "Corregir clase",
            "aprobada": "Ver clase",
        }.get(clase.estado_planificacion, "Abrir")


class DireccionHorasDocenteView(DireccionRequiredMixin, View):
    permission_required = "academico.change_clasehoradocente"
    template_name = "academico/direccion_horas_docente.html"

    def get(self, request):
        return render(request, self.template_name, self.get_context())

    def post(self, request):
        clase = get_object_or_404(self.get_clases_queryset(), pk=request.POST.get("clase"))
        registro = self.get_hora_docente(clase)
        original_docente = self.get_docente_programado(clase)
        form = ClaseHoraDocenteForm(
            request.POST,
            instance=registro,
            clase=clase,
            docente_programado=original_docente,
            prefix=self.form_prefix(clase),
        )
        if form.is_valid():
            self.save_hora_docente(form, clase, request.user)
            messages.success(request, "Horas del docente guardadas correctamente.")
        else:
            messages.error(request, "Revisa los datos de horas del docente.")
            return render(request, self.template_name, self.get_context(active_form=form, active_clase=clase))
        return redirect(self.redirect_url())

    def get_context(self, active_form=None, active_clase=None):
        selected_fecha, fecha_value = self.get_date_value("fecha", timezone.localdate())
        selected_grupo = self.get_selected_grupo()
        selected_docente = self.get_selected_docente()
        clases = self.get_clases_queryset().filter(fecha=selected_fecha)
        if selected_grupo:
            clases = clases.filter(materia_curso__grupo=selected_grupo)
        if selected_docente:
            clases = clases.filter(
                docente_responsable_filter(selected_docente) | Q(hora_docente__docente=selected_docente)
            )
        rows = [self.build_row(clase, active_form, active_clase) for clase in clases.distinct()]
        return {
            "title": "Horas docente",
            "fecha_value": fecha_value,
            "selected_fecha": selected_fecha,
            "selected_grupo": selected_grupo,
            "selected_grupo_id": str(selected_grupo.pk) if selected_grupo else "",
            "selected_docente": selected_docente,
            "selected_docente_id": str(selected_docente.pk) if selected_docente else "",
            "grupos": Curso.objects.filter(activo=True).order_by("nombre"),
            "docentes": Partner.objects.filter(es_docente=True, activo=True).order_by("nombre", "apellido"),
            "rows": rows,
            "stats": self.get_stats(rows),
            "report_url": reverse_lazy("academico:direccion_horas_docente_reporte"),
        }

    def get_clases_queryset(self):
        return (
            Clase.objects.select_related(
                "materia_curso__materia",
                "materia_curso__grupo",
                "docente",
                "hora_docente__docente",
                "hora_docente__docente_reemplazado",
                "horario_aula_curso__aula_curso__aula",
                "horario_aula_curso__horario_dia__horario",
            )
            .prefetch_related("materia_curso__profesor_materia_cursos__partner")
            .order_by(
                "fecha",
                "horario_aula_curso__horario_dia__horario__hora_inicio",
                "materia_curso__grupo__nombre",
                "materia_curso__materia__nombre",
            )
        )

    def build_row(self, clase, active_form=None, active_clase=None):
        registro = self.get_hora_docente(clase)
        original_docentes = get_clase_docentes(clase)
        original_docente = self.get_docente_programado(clase, original_docentes)
        form = active_form if active_clase and active_clase.pk == clase.pk else self.get_form(clase, registro, original_docente)
        horario = clase.horario_aula_curso.horario_dia.horario
        estado_value = form["estado"].value()
        return {
            "clase": clase,
            "registro": registro,
            "form": form,
            "form_prefix": self.form_prefix(clase),
            "docente_editable": estado_value == "reemplazo",
            "horario": horario,
            "horas_programadas": self.horas_programadas(clase),
            "docentes_programados": original_docentes,
            "estado_key": registro.estado if registro else "pendiente",
            "estado_label": registro.get_estado_display() if registro else "Pendiente",
            "docente_pagable": registro.docente if registro else None,
        }

    def get_form(self, clase, registro, original_docente):
        initial = {}
        if not registro:
            initial = {
                "estado": "asistio",
                "docente": original_docente,
                "horas": self.horas_programadas(clase),
            }
        return ClaseHoraDocenteForm(
            instance=registro,
            clase=clase,
            docente_programado=original_docente,
            prefix=self.form_prefix(clase),
            initial=initial,
        )

    def save_hora_docente(self, form, clase, user):
        registro = form.save(commit=False)
        original_docentes = get_clase_docentes(clase)
        original_docente = self.get_docente_programado(clase, original_docentes)
        registro.clase = clase
        if registro.estado == "asistio":
            registro.docente = original_docente
            registro.docente_reemplazado = None
        elif registro.estado == "reemplazo":
            registro.docente_reemplazado = original_docente if original_docente and original_docente != registro.docente else None
        else:
            registro.docente = None
            registro.docente_reemplazado = None
        registro.registrado_por = user
        registro.fecha_registro = timezone.now()
        registro.usuario_updated = user
        registro.full_clean()
        registro.save()
        return registro

    @staticmethod
    def get_hora_docente(clase):
        try:
            return clase.hora_docente
        except ClaseHoraDocente.DoesNotExist:
            return None

    @staticmethod
    def get_docente_programado(clase, docentes=None):
        docentes = docentes if docentes is not None else get_clase_docentes(clase)
        return docentes[0] if docentes else None

    @staticmethod
    def form_prefix(clase):
        return f"hora_{clase.pk}"

    @staticmethod
    def horas_programadas(clase):
        horario = clase.horario_aula_curso.horario_dia.horario
        inicio = horario.hora_inicio.hour * 60 + horario.hora_inicio.minute
        fin = horario.hora_fin.hour * 60 + horario.hora_fin.minute
        return (Decimal(fin - inicio) / Decimal("60")).quantize(Decimal("0.01"))

    def get_stats(self, rows):
        stats = {
            "clases": len(rows),
            "registradas": 0,
            "pendientes": 0,
            "reemplazos": 0,
            "horas": Decimal("0.00"),
        }
        for row in rows:
            registro = row["registro"]
            if not registro or registro.estado == "pendiente":
                stats["pendientes"] += 1
                continue
            stats["registradas"] += 1
            if registro.estado == "reemplazo":
                stats["reemplazos"] += 1
            if registro.estado in {"asistio", "reemplazo"}:
                stats["horas"] += registro.horas
        return stats

    def get_date_value(self, key, default):
        value = self.request.GET.get(key) or self.request.POST.get(key) or default.isoformat()
        try:
            parsed = date.fromisoformat(str(value))
        except ValueError:
            parsed = default
            value = default.isoformat()
        return parsed, value

    def get_selected_grupo(self):
        grupo_id = self.request.GET.get("grupo") or self.request.POST.get("grupo") or ""
        if not grupo_id:
            return None
        try:
            return Curso.objects.filter(pk=grupo_id, activo=True).first()
        except (TypeError, ValueError):
            return None

    def get_selected_docente(self):
        docente_id = self.request.GET.get("docente") or self.request.POST.get("docente") or ""
        if not docente_id:
            return None
        try:
            return Partner.objects.filter(pk=docente_id, es_docente=True, activo=True).first()
        except (TypeError, ValueError):
            return None

    def redirect_url(self):
        params = {"fecha": self.request.POST.get("fecha") or timezone.localdate().isoformat()}
        grupo = self.request.POST.get("grupo") or ""
        docente = self.request.POST.get("docente") or ""
        if grupo:
            params["grupo"] = grupo
        if docente:
            params["docente"] = docente
        return f"{reverse_lazy('academico:direccion_horas_docente')}?{urlencode(params)}"


class DireccionReporteHorasDocenteView(DireccionRequiredMixin, View):
    permission_required = "academico.report_clasehoradocente"
    template_name = "academico/direccion_reporte_horas_docente.html"

    def get(self, request):
        context = self.get_context()
        if request.GET.get("export") == "excel":
            return self.export_excel(context)
        return render(request, self.template_name, context)

    def get_context(self):
        today = timezone.localdate()
        desde, desde_value = self.get_date_value("desde", today.replace(day=1))
        hasta, hasta_value = self.get_date_value("hasta", today)
        if desde > hasta:
            desde, hasta = hasta, desde
            desde_value, hasta_value = desde.isoformat(), hasta.isoformat()
        selected_docente = self.get_selected_docente()
        rows = self.get_rows(desde, hasta, selected_docente)
        return {
            "title": "Reporte horas docente",
            "desde_value": desde_value,
            "hasta_value": hasta_value,
            "selected_docente": selected_docente,
            "selected_docente_id": str(selected_docente.pk) if selected_docente else "",
            "docentes": Partner.objects.filter(es_docente=True, activo=True).order_by("nombre", "apellido"),
            "rows": rows,
            "stats": self.get_stats(rows),
            "export_url": self.export_url(desde_value, hasta_value, selected_docente),
            "daily_url": reverse_lazy("academico:direccion_horas_docente"),
        }

    def get_rows(self, desde, hasta, selected_docente):
        queryset = (
            ClaseHoraDocente.objects.select_related(
                "clase__materia_curso__materia",
                "clase__materia_curso__grupo",
                "clase__horario_aula_curso__aula_curso__aula",
                "clase__horario_aula_curso__horario_dia__horario",
                "docente",
                "docente_reemplazado",
                "registrado_por",
            )
            .filter(
                clase__fecha__gte=desde,
                clase__fecha__lte=hasta,
                estado__in=["asistio", "reemplazo"],
                horas__gt=0,
            )
            .order_by("clase__fecha", "docente__nombre", "clase__horario_aula_curso__horario_dia__horario__hora_inicio")
        )
        if selected_docente:
            queryset = queryset.filter(docente=selected_docente)
        return [self.build_row(registro) for registro in queryset]

    def build_row(self, registro):
        clase = registro.clase
        horario = clase.horario_aula_curso.horario_dia.horario
        return {
            "registro": registro,
            "clase": clase,
            "fecha": clase.fecha,
            "hora": f"{horario.hora_inicio:%H:%M} - {horario.hora_fin:%H:%M}",
            "grupo": clase.materia_curso.grupo,
            "aula": clase.horario_aula_curso.aula_curso.aula,
            "materia": clase.materia_curso.materia,
            "docente": registro.docente,
            "docente_reemplazado": registro.docente_reemplazado,
            "estado": registro.get_estado_display(),
            "horas": registro.horas,
            "observacion": registro.observacion or "",
        }

    def get_stats(self, rows):
        docentes = {}
        total_horas = Decimal("0.00")
        reemplazos = 0
        for row in rows:
            total_horas += row["horas"]
            if row["registro"].estado == "reemplazo":
                reemplazos += 1
            if row["docente"]:
                docentes.setdefault(row["docente"].pk, Decimal("0.00"))
                docentes[row["docente"].pk] += row["horas"]
        return {
            "registros": len(rows),
            "horas": total_horas,
            "docentes": len(docentes),
            "reemplazos": reemplazos,
        }

    def export_excel(self, context):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = safe_sheet_title("Horas docente")
        headers = ["Fecha", "Horario", "Grupo", "Aula", "Materia", "Docente", "Estado", "Horas", "Reemplaza a", "Observacion"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.alignment = Alignment(horizontal="center")

        current_row = 2
        for row in context["rows"]:
            values = [
                row["fecha"].strftime("%d/%m/%Y"),
                row["hora"],
                row["grupo"].nombre,
                str(row["aula"]),
                row["materia"].nombre,
                row["docente"].nombre_completo() if row["docente"] else "",
                row["estado"],
                float(row["horas"]),
                row["docente_reemplazado"].nombre_completo() if row["docente_reemplazado"] else "",
                row["observacion"],
            ]
            for col, value in enumerate(values, start=1):
                sheet.cell(row=current_row, column=col, value=value)
            current_row += 1

        sheet.cell(row=current_row + 1, column=7, value="Total horas").font = Font(bold=True)
        sheet.cell(row=current_row + 1, column=8, value=float(context["stats"]["horas"])).font = Font(bold=True)
        widths = [14, 18, 24, 18, 24, 28, 14, 10, 28, 36]
        for col, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col)].width = width
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="reporte_horas_docente.xlsx"'
        return response

    def export_url(self, desde_value, hasta_value, selected_docente):
        params = {"desde": desde_value, "hasta": hasta_value, "export": "excel"}
        if selected_docente:
            params["docente"] = selected_docente.pk
        return f"{reverse_lazy('academico:direccion_horas_docente_reporte')}?{urlencode(params)}"

    def get_date_value(self, key, default):
        value = self.request.GET.get(key) or default.isoformat()
        try:
            parsed = date.fromisoformat(str(value))
        except ValueError:
            parsed = default
            value = default.isoformat()
        return parsed, value

    def get_selected_docente(self):
        docente_id = self.request.GET.get("docente") or ""
        if not docente_id:
            return None
        try:
            return Partner.objects.filter(pk=docente_id, es_docente=True, activo=True).first()
        except (TypeError, ValueError):
            return None


class DocenteClaseAsistenciaView(LoginRequiredMixin, View):
    template_name = "academico/docente_clase_asistencia.html"

    def get_docente(self):
        docente = getattr(self.request.user, "partner", None)
        if docente and docente.es_docente:
            return docente
        return None

    def get_clase(self):
        docente = self.get_docente()
        if not docente:
            return get_object_or_404(Clase.objects.none(), pk=self.kwargs["pk"])
        return get_object_or_404(
            Clase.objects.select_related(
                "materia_curso__materia",
                "materia_curso__grupo",
                "docente",
                "horario_aula_curso__aula_curso__aula",
                "horario_aula_curso__horario_dia__horario",
            )
            .filter(docente_responsable_filter(docente))
            .distinct(),
            pk=self.kwargs["pk"],
        )

    def get(self, request, pk):
        clase = self.get_clase()
        date_response = self.ensure_attendance_date(request, clase)
        if date_response:
            return date_response
        return render(request, self.template_name, self.get_context(clase))

    def post(self, request, pk):
        clase = self.get_clase()
        date_response = self.ensure_attendance_date(request, clase)
        if date_response:
            return date_response
        action = request.POST.get("attendance_action") or "save"
        if action == "close":
            return self.handle_close(request, clase)
        if action != "save":
            messages.error(request, "Accion no valida.")
            return redirect("academico:docente_clase_asistencia", pk=clase.pk)
        if clase.asistencia_cerrada:
            messages.error(request, "La asistencia ya esta cerrada y no se puede modificar.")
            return redirect("academico:docente_clase_asistencia", pk=clase.pk)
        rows, moved_out_rows = self.get_roster_rows(clase)
        valid_states = {choice[0] for choice in ClaseAsistencia.ESTADO_CHOICES}
        registrado_por = self.get_docente()
        with transaction.atomic():
            if moved_out_rows:
                ClaseAsistencia.objects.filter(
                    clase=clase,
                    estudiante_id__in=[row["asignacion"].estudiante_id for row in moved_out_rows],
                ).delete()
            for row in rows:
                asignacion = row["asignacion"]
                estado = request.POST.get(f"estado_{asignacion.pk}", "presente")
                if estado not in valid_states:
                    estado = "presente"
                observacion = (request.POST.get(f"observacion_{asignacion.pk}") or "").strip()
                ClaseAsistencia.objects.update_or_create(
                    clase=clase,
                    estudiante=asignacion.estudiante,
                    defaults={
                        "estado": estado,
                        "observacion": observacion,
                        "registrado_por": registrado_por,
                        "usuario_updated": request.user,
                    },
                )
        messages.success(request, f"Asistencia guardada para {len(rows)} estudiante(s).")
        return redirect("academico:docente_clase_asistencia", pk=clase.pk)

    def ensure_attendance_date(self, request, clase):
        if clase.fecha == timezone.localdate():
            return None
        messages.error(request, "La asistencia solo se habilita el dia de la clase.")
        return redirect("academico:docente_horarios")

    def handle_close(self, request, clase):
        if clase.asistencia_cerrada:
            messages.info(request, "La asistencia ya estaba cerrada.")
            return redirect("academico:docente_clase_asistencia", pk=clase.pk)
        rows, _ = self.get_roster_rows(clase)
        if not rows:
            messages.error(request, "No hay estudiantes para cerrar la asistencia.")
            return redirect("academico:docente_clase_asistencia", pk=clase.pk)
        saved_student_ids = set(ClaseAsistencia.objects.filter(clase=clase).values_list("estudiante_id", flat=True))
        pending_rows = [row for row in rows if row["asignacion"].estudiante_id not in saved_student_ids]
        if pending_rows:
            messages.error(request, "Guarda la asistencia antes de cerrarla.")
            return redirect("academico:docente_clase_asistencia", pk=clase.pk)
        clase.asistencia_cerrada = True
        clase.asistencia_cerrada_por = self.get_docente()
        clase.fecha_cierre_asistencia = timezone.now()
        clase.save(update_fields=["asistencia_cerrada", "asistencia_cerrada_por", "fecha_cierre_asistencia"])
        messages.success(request, "Asistencia cerrada correctamente.")
        return redirect("academico:docente_clase_asistencia", pk=clase.pk)

    def get_context(self, clase):
        rows, moved_out_rows = self.get_roster_rows(clase)
        horario = clase.horario_aula_curso.horario_dia.horario
        has_saved_attendance = ClaseAsistencia.objects.filter(clase=clase).exists()
        can_edit_attendance = clase.fecha == timezone.localdate() and not clase.asistencia_cerrada
        return {
            "title": "Asistencia",
            "clase": clase,
            "horario": horario,
            "rows": rows,
            "moved_out_rows": moved_out_rows,
            "estado_choices": ClaseAsistencia.ESTADO_CHOICES,
            "total_estudiantes": len(rows),
            "can_edit_attendance": can_edit_attendance,
            "can_close_attendance": can_edit_attendance and bool(rows) and has_saved_attendance,
            "has_saved_attendance": has_saved_attendance,
            "planificacion_url": reverse_lazy("academico:docente_clase_planificar", kwargs={"pk": clase.pk}),
            "cancel_url": reverse_lazy("academico:docente_horarios"),
        }

    def get_roster_rows(self, clase):
        base_asignaciones = list(
            GrupoEstudiante.objects.select_related("estudiante", "ficha_inscripcion", "grupo")
            .filter(grupo=clase.materia_curso.grupo, estado="activo")
            .order_by("estudiante__nombre", "ficha_inscripcion__numero")
        )
        movimientos = list(
            ClaseEstudianteMovimiento.objects.select_related(
                "asignacion__estudiante",
                "asignacion__ficha_inscripcion",
                "asignacion__grupo",
                "clase_origen__materia_curso__materia",
                "clase_origen__materia_curso__grupo",
                "clase_origen__horario_aula_curso__aula_curso__aula",
                "clase_origen__horario_aula_curso__horario_dia__horario",
                "clase_destino__materia_curso__materia",
                "clase_destino__materia_curso__grupo",
                "clase_destino__horario_aula_curso__aula_curso__aula",
                "clase_destino__horario_aula_curso__horario_dia__horario",
            )
            .filter(
                activo=True,
                fecha_inicio__lte=clase.fecha,
            )
            .filter(
                Q(clase_origen__materia_curso=clase.materia_curso)
                | Q(clase_destino__materia_curso=clase.materia_curso)
            )
        )
        moved_out_by_assignment = {
            movimiento.asignacion_id: movimiento
            for movimiento in movimientos
            if movimiento.clase_origen.materia_curso_id == clase.materia_curso_id
        }
        incoming_by_assignment = {
            movimiento.asignacion_id: movimiento
            for movimiento in movimientos
            if movimiento.clase_destino.materia_curso_id == clase.materia_curso_id
        }
        visible_assignments_by_id = {
            asignacion.pk: asignacion
            for asignacion in base_asignaciones
            if asignacion.pk not in moved_out_by_assignment
        }
        visible_assignments_by_id.update(
            {
                movimiento.asignacion_id: movimiento.asignacion
                for movimiento in incoming_by_assignment.values()
            }
        )
        visible_assignments = sorted(
            visible_assignments_by_id.values(),
            key=lambda asignacion: (
                asignacion.estudiante.nombre or "",
                asignacion.ficha_inscripcion.numero or "",
            ),
        )
        attendance_by_student = {
            attendance.estudiante_id: attendance
            for attendance in ClaseAsistencia.objects.filter(
                clase=clase,
                estudiante_id__in=[asignacion.estudiante_id for asignacion in visible_assignments],
            )
        }
        rows = []
        for asignacion in visible_assignments:
            attendance = attendance_by_student.get(asignacion.estudiante_id)
            rows.append(
                {
                    "asignacion": asignacion,
                    "estudiante": asignacion.estudiante,
                    "ficha": asignacion.ficha_inscripcion,
                    "attendance": attendance,
                    "estado": attendance.estado if attendance else "presente",
                    "observacion": attendance.observacion if attendance else "",
                    "incoming": incoming_by_assignment.get(asignacion.pk),
                }
            )
        moved_out_rows = [
            {
                "movimiento": movimiento,
                "asignacion": movimiento.asignacion,
                "estudiante": movimiento.asignacion.estudiante,
                "destino_label": self.clase_label(movimiento.clase_destino),
            }
            for movimiento in moved_out_by_assignment.values()
        ]
        return rows, sorted(moved_out_rows, key=lambda row: row["estudiante"].nombre)

    @staticmethod
    def clase_label(clase):
        horario = clase.horario_aula_curso.horario_dia.horario
        return (
            f"{clase.fecha:%d/%m/%Y} - "
            f"{clase.horario_aula_curso.aula_curso.aula} - "
            f"{horario.hora_inicio:%H:%M}-{horario.hora_fin:%H:%M}"
        )


class CoordinacionRevisionAsistenciaView(CoordinacionRequiredMixin, View):
    permission_required = "academico.view_claseasistencia"
    template_name = "academico/coordinacion_revision_asistencia.html"
    status_filters = (
        {"key": "todas", "label": "Todas", "icon": "ri-list-check-3"},
        {"key": "pendientes", "label": "Pendientes", "icon": "ri-time-line"},
        {"key": "ausentes", "label": "Con ausentes", "icon": "ri-user-unfollow-line"},
        {"key": "observaciones", "label": "Observaciones", "icon": "ri-chat-3-line"},
        {"key": "cerradas", "label": "Cerradas", "icon": "ri-lock-2-line"},
    )
    attendance_state_labels = {**dict(ClaseAsistencia.ESTADO_CHOICES), "pendiente": "Pendiente"}

    def get(self, request):
        selected_grupo = self.get_selected_grupo()
        selected_docente = self.get_selected_docente()
        requested_estado = request.GET.get("estado", "")
        clases_queryset = self.get_clases_queryset()

        if selected_grupo:
            clases_queryset = clases_queryset.filter(materia_curso__grupo=selected_grupo)
        if selected_docente:
            clases_queryset = clases_queryset.filter(docente_responsable_filter(selected_docente))

        cards = [self.build_attendance_card(clase) for clase in clases_queryset.distinct()]
        stats = self.get_attendance_stats(cards)
        selected_estado = self.get_selected_filter(requested_estado)
        attendance_cards = self.filter_cards(cards, selected_estado)
        return render(
            request,
            self.template_name,
            {
                "title": "Revision de asistencia",
                "attendance_cards": attendance_cards,
                "attendance_stats": stats,
                "grupos": Curso.objects.filter(activo=True).order_by("nombre"),
                "docentes": Partner.objects.filter(es_docente=True, activo=True).order_by("nombre"),
                "selected_grupo": selected_grupo,
                "selected_grupo_id": str(selected_grupo.pk) if selected_grupo else "",
                "selected_docente": selected_docente,
                "selected_docente_id": str(selected_docente.pk) if selected_docente else "",
                "selected_estado": selected_estado,
                "selected_estado_label": self.get_filter_label(selected_estado),
                "status_tabs": self.get_status_tabs(selected_estado, selected_grupo, selected_docente, stats),
            },
        )

    def get_clases_queryset(self):
        return (
            Clase.objects.select_related(
                "materia_curso__materia",
                "materia_curso__grupo",
                "docente",
                "asistencia_cerrada_por",
                "horario_aula_curso__aula_curso__aula",
                "horario_aula_curso__horario_dia__horario",
            )
            .prefetch_related("materia_curso__profesor_materia_cursos__partner")
            .order_by("fecha", "horario_aula_curso__horario_dia__horario__hora_inicio", "materia_curso__grupo__nombre")
        )

    def get_selected_grupo(self):
        grupo_id = self.request.GET.get("grupo") or ""
        if not grupo_id:
            return None
        try:
            return Curso.objects.filter(pk=grupo_id, activo=True).first()
        except (TypeError, ValueError):
            return None

    def get_selected_docente(self):
        docente_id = self.request.GET.get("docente") or ""
        if not docente_id:
            return None
        try:
            return Partner.objects.filter(pk=docente_id, es_docente=True, activo=True).first()
        except (TypeError, ValueError):
            return None

    def build_attendance_card(self, clase):
        rows, moved_out_rows = DocenteClaseAsistenciaView().get_roster_rows(clase)
        horario = clase.horario_aula_curso.horario_dia.horario
        counts = {key: 0 for key, _ in ClaseAsistencia.ESTADO_CHOICES}
        counts["pendiente"] = 0
        fichas = []
        observation_rows = []
        for row in rows:
            attendance = row["attendance"]
            estado = attendance.estado if attendance else "pendiente"
            counts[estado] = counts.get(estado, 0) + 1
            observacion = row["observacion"] or ""
            ficha_row = {
                "ficha": row["ficha"],
                "estudiante": row["estudiante"],
                "estado": estado,
                "estado_label": self.attendance_state_labels.get(estado, estado),
                "observacion": observacion,
                "incoming": row["incoming"],
            }
            fichas.append(ficha_row)
            if observacion:
                observation_rows.append(ficha_row)

        pending_count = counts.get("pendiente", 0)
        saved_count = len(rows) - pending_count
        status_key, status_label = self.get_card_status(clase, len(rows), pending_count)
        return {
            "clase": clase,
            "horario": horario,
            "grupo": clase.materia_curso.grupo,
            "materia": clase.materia_curso.materia,
            "aula": clase.horario_aula_curso.aula_curso.aula,
            "docentes": get_clase_docentes(clase),
            "counts": counts,
            "count_items": self.get_count_items(counts),
            "fichas": fichas,
            "moved_out_rows": moved_out_rows,
            "observation_rows": observation_rows,
            "observation_count": len(observation_rows),
            "total_estudiantes": len(rows),
            "saved_count": saved_count,
            "pending_count": pending_count,
            "status_key": status_key,
            "status_label": status_label,
            "is_today": clase.fecha == timezone.localdate(),
            "report_url": reverse_lazy("academico:coordinacion_reporte_asistencia_clase", kwargs={"pk": clase.pk}),
        }

    def get_count_items(self, counts):
        return (
            {"key": "presente", "label": "Presentes", "count": counts.get("presente", 0), "icon": "ri-user-follow-line"},
            {"key": "ausente", "label": "Ausentes", "count": counts.get("ausente", 0), "icon": "ri-user-unfollow-line"},
            {"key": "atraso", "label": "Atrasos", "count": counts.get("atraso", 0), "icon": "ri-time-line"},
            {"key": "justificado", "label": "Justificados", "count": counts.get("justificado", 0), "icon": "ri-file-check-line"},
            {"key": "pendiente", "label": "Pendientes", "count": counts.get("pendiente", 0), "icon": "ri-checkbox-blank-circle-line"},
        )

    def get_card_status(self, clase, total_estudiantes, pending_count):
        if not total_estudiantes:
            return "sin-estudiantes", "Sin estudiantes"
        if clase.asistencia_cerrada:
            return "cerrada", "Cerrada"
        if pending_count:
            return "pendiente", "Pendiente"
        return "completa", "Completa"

    def get_attendance_stats(self, cards):
        stats = {
            "total": len(cards),
            "estudiantes": 0,
            "presentes": 0,
            "ausentes": 0,
            "atrasos": 0,
            "justificados": 0,
            "pendientes_registro": 0,
            "observaciones": 0,
            "cerradas": 0,
            "clases_pendientes": 0,
            "con_ausentes": 0,
            "con_observaciones": 0,
        }
        for card in cards:
            counts = card["counts"]
            stats["estudiantes"] += card["total_estudiantes"]
            stats["presentes"] += counts.get("presente", 0)
            stats["ausentes"] += counts.get("ausente", 0)
            stats["atrasos"] += counts.get("atraso", 0)
            stats["justificados"] += counts.get("justificado", 0)
            stats["pendientes_registro"] += card["pending_count"]
            stats["observaciones"] += card["observation_count"]
            if card["clase"].asistencia_cerrada:
                stats["cerradas"] += 1
            if card["pending_count"]:
                stats["clases_pendientes"] += 1
            if counts.get("ausente", 0):
                stats["con_ausentes"] += 1
            if card["observation_count"]:
                stats["con_observaciones"] += 1
        return stats

    def get_selected_filter(self, requested_filter):
        valid_filters = {item["key"] for item in self.status_filters}
        return requested_filter if requested_filter in valid_filters else "todas"

    def get_filter_label(self, selected_filter):
        filter_item = next((item for item in self.status_filters if item["key"] == selected_filter), None)
        return filter_item["label"] if filter_item else ""

    def get_status_tabs(self, selected_filter, selected_grupo, selected_docente, stats):
        base_url = reverse_lazy("academico:coordinacion_revision_asistencia")
        count_map = {
            "todas": stats["total"],
            "pendientes": stats["clases_pendientes"],
            "ausentes": stats["con_ausentes"],
            "observaciones": stats["con_observaciones"],
            "cerradas": stats["cerradas"],
        }
        tabs = []
        for item in self.status_filters:
            params = {"estado": item["key"]}
            if selected_grupo:
                params["grupo"] = selected_grupo.pk
            if selected_docente:
                params["docente"] = selected_docente.pk
            tabs.append(
                {
                    **item,
                    "count": count_map.get(item["key"], 0),
                    "url": f"{base_url}?{urlencode(params)}",
                    "is_active": item["key"] == selected_filter,
                }
            )
        return tabs

    def filter_cards(self, cards, selected_filter):
        if selected_filter == "pendientes":
            return [card for card in cards if card["pending_count"]]
        if selected_filter == "ausentes":
            return [card for card in cards if card["counts"].get("ausente", 0)]
        if selected_filter == "observaciones":
            return [card for card in cards if card["observation_count"]]
        if selected_filter == "cerradas":
            return [card for card in cards if card["clase"].asistencia_cerrada]
        return cards


class CoordinacionReporteAsistenciaClaseView(CoordinacionRevisionAsistenciaView):
    template_name = "academico/coordinacion_reporte_asistencia_clase.html"

    def get(self, request, pk):
        clase = self.get_clase()
        card = self.build_attendance_card(clase)
        return render(
            request,
            self.template_name,
            {
                "title": "Reporte de asistencia",
                "card": card,
                "clase": clase,
                "list_url": reverse_lazy("academico:coordinacion_revision_asistencia"),
            },
        )

    def get_clase(self):
        return get_object_or_404(
            self.get_clases_queryset().prefetch_related(
                Prefetch("asistencias_clase", queryset=ClaseAsistencia.objects.select_related("estudiante", "registrado_por")),
            ),
            pk=self.kwargs["pk"],
        )


class CoordinacionReporteAsistenciaAlumnoView(CoordinacionRequiredMixin, View):
    permission_required = "academico.view_claseasistencia"
    template_name = "academico/coordinacion_reporte_asistencia_alumno.html"
    attendance_state_labels = {**dict(ClaseAsistencia.ESTADO_CHOICES), "pendiente": "Pendiente"}

    def get(self, request):
        context = self.get_context()
        if request.GET.get("export") == "excel":
            if not context["selected_estudiante"]:
                messages.error(request, "Selecciona un alumno para exportar el reporte.")
                return redirect("academico:coordinacion_reporte_asistencia_alumno")
            return self.export_excel(context)
        return render(request, self.template_name, context)

    def get_context(self):
        selected_grupo = self.get_selected_grupo()
        selected_estudiante = self.get_selected_estudiante(selected_grupo)
        fecha_desde, desde_value = self.get_date_value("desde")
        fecha_hasta, hasta_value = self.get_date_value("hasta")
        if fecha_desde and fecha_hasta and fecha_desde > fecha_hasta:
            fecha_desde, fecha_hasta = fecha_hasta, fecha_desde
            desde_value, hasta_value = hasta_value, desde_value

        rows = []
        if selected_estudiante:
            rows = self.get_attendance_rows(selected_estudiante, selected_grupo, fecha_desde, fecha_hasta)
        stats = self.get_stats(rows)
        return {
            "title": "Reporte de asistencia del alumno",
            "grupos": Curso.objects.filter(activo=True).order_by("nombre"),
            "estudiantes": self.get_students_queryset(selected_grupo),
            "selected_grupo": selected_grupo,
            "selected_grupo_id": str(selected_grupo.pk) if selected_grupo else "",
            "selected_estudiante": selected_estudiante,
            "selected_estudiante_id": str(selected_estudiante.pk) if selected_estudiante else "",
            "desde_value": desde_value,
            "hasta_value": hasta_value,
            "rows": rows,
            "stats": stats,
            "count_items": self.get_count_items(stats),
            "export_url": self.get_export_url(selected_grupo, selected_estudiante, desde_value, hasta_value),
        }

    def get_selected_grupo(self):
        grupo_id = self.request.GET.get("grupo") or ""
        if not grupo_id:
            return None
        try:
            return Curso.objects.filter(pk=grupo_id, activo=True).first()
        except (TypeError, ValueError):
            return None

    def get_selected_estudiante(self, selected_grupo=None):
        estudiante_id = self.request.GET.get("estudiante") or ""
        if not estudiante_id:
            return None
        queryset = Partner.objects.filter(pk=estudiante_id, es_estudiante=True, activo=True)
        if selected_grupo:
            queryset = queryset.filter(grupo_asignaciones__grupo=selected_grupo, grupo_asignaciones__estado="activo")
        return queryset.distinct().first()

    def get_students_queryset(self, selected_grupo=None):
        queryset = Partner.objects.filter(
            es_estudiante=True,
            activo=True,
            grupo_asignaciones__estado="activo",
        )
        if selected_grupo:
            queryset = queryset.filter(grupo_asignaciones__grupo=selected_grupo)
        return queryset.distinct().order_by("nombre", "identificacion")

    def get_date_value(self, field_name):
        value = (self.request.GET.get(field_name) or "").strip()
        if not value:
            return None, ""
        try:
            return date.fromisoformat(value), value
        except ValueError:
            return None, ""

    def get_attendance_rows(self, estudiante, selected_grupo=None, fecha_desde=None, fecha_hasta=None):
        assignments = list(
            GrupoEstudiante.objects.select_related("grupo", "ficha_inscripcion", "estudiante")
            .filter(estudiante=estudiante, estado="activo")
            .order_by("fecha_asignacion", "grupo__nombre")
        )
        if selected_grupo:
            assignments = [assignment for assignment in assignments if assignment.grupo_id == selected_grupo.pk]
        if not assignments:
            return []

        group_ids = {assignment.grupo_id for assignment in assignments}
        movements = ClaseEstudianteMovimiento.objects.filter(asignacion__in=assignments, activo=True)
        destination_materia_curso_ids = set(movements.values_list("clase_destino__materia_curso_id", flat=True))
        clase_filter = Q(materia_curso__grupo_id__in=group_ids)
        if destination_materia_curso_ids:
            clase_filter |= Q(materia_curso_id__in=destination_materia_curso_ids)

        clases = (
            Clase.objects.select_related(
                "materia_curso__materia",
                "materia_curso__grupo",
                "docente",
                "horario_aula_curso__aula_curso__aula",
                "horario_aula_curso__horario_dia__horario",
            )
            .prefetch_related("materia_curso__profesor_materia_cursos__partner")
            .filter(clase_filter)
            .distinct()
        )
        if fecha_desde:
            clases = clases.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            clases = clases.filter(fecha__lte=fecha_hasta)

        roster_view = DocenteClaseAsistenciaView()
        report_rows = []
        for clase in clases.order_by("fecha", "horario_aula_curso__horario_dia__horario__hora_inicio", "materia_curso__materia__nombre"):
            roster_rows, _ = roster_view.get_roster_rows(clase)
            student_row = next((row for row in roster_rows if row["estudiante"].pk == estudiante.pk), None)
            if not student_row:
                continue
            asignacion = student_row["asignacion"]
            if asignacion.fecha_asignacion and asignacion.fecha_asignacion > clase.fecha:
                continue
            report_rows.append(self.build_attendance_row(clase, student_row))
        return report_rows

    def build_attendance_row(self, clase, student_row):
        horario = clase.horario_aula_curso.horario_dia.horario
        attendance = student_row["attendance"]
        estado = attendance.estado if attendance else "pendiente"
        docentes = get_clase_docentes(clase)
        return {
            "clase": clase,
            "fecha": clase.fecha,
            "materia": clase.materia_curso.materia,
            "grupo": clase.materia_curso.grupo,
            "aula": clase.horario_aula_curso.aula_curso.aula,
            "horario": horario,
            "docente_label": ", ".join(docente.nombre for docente in docentes) if docentes else "Sin docente asignado",
            "estado": estado,
            "estado_label": self.attendance_state_labels.get(estado, estado),
            "observacion": (attendance.observacion or "") if attendance else "",
            "registrado_por": attendance.registrado_por if attendance else None,
            "registrado_en": attendance.updated_at if attendance else None,
            "asistencia_cerrada": clase.asistencia_cerrada,
            "incoming": student_row["incoming"],
        }

    def get_stats(self, rows):
        stats = {
            "total": len(rows),
            "presente": 0,
            "ausente": 0,
            "atraso": 0,
            "justificado": 0,
            "pendiente": 0,
            "observaciones": 0,
        }
        for row in rows:
            stats[row["estado"]] = stats.get(row["estado"], 0) + 1
            if row["observacion"]:
                stats["observaciones"] += 1
        stats["registradas"] = stats["total"] - stats["pendiente"]
        return stats

    def get_count_items(self, stats):
        return (
            {"key": "total", "label": "Clases", "count": stats["total"], "icon": "ri-calendar-check-line"},
            {"key": "presente", "label": "Presentes", "count": stats["presente"], "icon": "ri-user-follow-line"},
            {"key": "ausente", "label": "Ausentes", "count": stats["ausente"], "icon": "ri-user-unfollow-line"},
            {"key": "atraso", "label": "Atrasos", "count": stats["atraso"], "icon": "ri-time-line"},
            {"key": "justificado", "label": "Justificados", "count": stats["justificado"], "icon": "ri-file-check-line"},
            {"key": "pendiente", "label": "Pendientes", "count": stats["pendiente"], "icon": "ri-checkbox-blank-circle-line"},
            {"key": "observaciones", "label": "Observaciones", "count": stats["observaciones"], "icon": "ri-chat-3-line"},
        )

    def get_export_url(self, selected_grupo, selected_estudiante, desde_value, hasta_value):
        if not selected_estudiante:
            return ""
        params = {"estudiante": selected_estudiante.pk, "export": "excel"}
        if selected_grupo:
            params["grupo"] = selected_grupo.pk
        if desde_value:
            params["desde"] = desde_value
        if hasta_value:
            params["hasta"] = hasta_value
        return f"{reverse_lazy('academico:coordinacion_reporte_asistencia_alumno')}?{urlencode(params)}"

    def export_excel(self, context):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = safe_sheet_title("Asistencia alumno")
        self.write_excel_sheet(sheet, context)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        estudiante = context["selected_estudiante"]
        filename = f"asistencia_{estudiante.identificacion or estudiante.pk}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def write_excel_sheet(self, sheet, context):
        estudiante = context["selected_estudiante"]
        selected_grupo = context["selected_grupo"]
        rows = context["rows"]
        stats = context["stats"]
        title_fill = PatternFill("solid", fgColor="DCEBFF")
        header_fill = PatternFill("solid", fgColor="0F766E")
        label_fill = PatternFill("solid", fgColor="F1F5F9")
        thin = Side(style="thin", color="D7DEE8")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="top", wrap_text=True)

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        title_cell = sheet.cell(row=1, column=1, value="Reporte de asistencia del alumno")
        title_cell.font = Font(bold=True, size=14, color="111827")
        title_cell.fill = title_fill
        title_cell.alignment = center

        info_rows = [
            ("Alumno", estudiante.nombre),
            ("Identificacion", estudiante.identificacion or ""),
            ("Grupo", selected_grupo.nombre if selected_grupo else "Todos los grupos"),
            ("Rango", self.get_range_label(context["desde_value"], context["hasta_value"])),
            ("Generado", timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")),
        ]
        current_row = 3
        for label, value in info_rows:
            sheet.cell(row=current_row, column=1, value=label)
            sheet.cell(row=current_row, column=2, value=value)
            sheet.cell(row=current_row, column=1).font = Font(bold=True, color="334155")
            sheet.cell(row=current_row, column=1).fill = label_fill
            sheet.cell(row=current_row, column=1).border = border
            sheet.cell(row=current_row, column=2).border = border
            current_row += 1

        current_row += 1
        summary = [
            ("Clases", stats["total"]),
            ("Presentes", stats["presente"]),
            ("Ausentes", stats["ausente"]),
            ("Atrasos", stats["atraso"]),
            ("Justificados", stats["justificado"]),
            ("Pendientes", stats["pendiente"]),
            ("Observaciones", stats["observaciones"]),
        ]
        for col, (label, value) in enumerate(summary, start=1):
            label_cell = sheet.cell(row=current_row, column=col, value=label)
            value_cell = sheet.cell(row=current_row + 1, column=col, value=value)
            label_cell.font = Font(bold=True, color="FFFFFF")
            label_cell.fill = header_fill
            value_cell.font = Font(bold=True, size=13, color="111827")
            value_cell.fill = label_fill
            label_cell.alignment = center
            value_cell.alignment = center
            label_cell.border = border
            value_cell.border = border

        current_row += 4
        headers = ["Fecha", "Materia", "Grupo", "Aula", "Horario", "Docente", "Asistencia", "Observacion", "Cierre"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        current_row += 1
        if rows:
            for row in rows:
                values = [
                    row["fecha"].strftime("%d/%m/%Y"),
                    row["materia"].nombre,
                    row["grupo"].nombre,
                    str(row["aula"]),
                    f"{row['horario'].hora_inicio:%H:%M} - {row['horario'].hora_fin:%H:%M}",
                    row["docente_label"],
                    row["estado_label"],
                    row["observacion"] or "Sin observacion",
                    "Cerrada" if row["asistencia_cerrada"] else "Abierta",
                ]
                for col, value in enumerate(values, start=1):
                    cell = sheet.cell(row=current_row, column=col, value=value)
                    cell.border = border
                    cell.alignment = left if col in {2, 6, 8} else center
                current_row += 1
        else:
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
            cell = sheet.cell(row=current_row, column=1, value="Sin asistencias para el filtro seleccionado.")
            cell.alignment = center
            cell.border = border

        widths = [14, 30, 22, 18, 16, 28, 16, 42, 14]
        for col, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col)].width = width
        sheet.freeze_panes = "A12"

    def get_range_label(self, desde_value, hasta_value):
        if desde_value and hasta_value:
            return f"{desde_value} a {hasta_value}"
        if desde_value:
            return f"Desde {desde_value}"
        if hasta_value:
            return f"Hasta {hasta_value}"
        return "Todas las fechas"


class DocenteClasePlanificacionView(LoginRequiredMixin, View):
    template_name = "academico/docente_clase_planificacion.html"

    def get_clase(self):
        docente = getattr(self.request.user, "partner", None)
        if not docente or not docente.es_docente:
            return get_object_or_404(Clase.objects.none(), pk=self.kwargs["pk"])
        return get_object_or_404(
            Clase.objects.select_related(
                "materia_curso__materia",
                "materia_curso__grupo",
                "docente",
                "tema",
                "subtema",
                "horario_aula_curso__aula_curso__aula",
                "horario_aula_curso__horario_dia__horario",
            )
            .prefetch_related(
                "competencias",
                "estrategias",
                "recursos",
                "clase_subtemas__subtema",
                Prefetch("clase_recursos", queryset=ClaseRecurso.objects.select_related("recurso")),
            )
            .filter(docente_responsable_filter(docente)),
            pk=self.kwargs["pk"],
        )

    def get(self, request, pk):
        clase = self.get_clase()
        form = DocenteClasePlanificacionForm(
            instance=clase,
            clase=clase,
            unavailable_subtema_ids=self.get_unavailable_subtema_ids_by_tema(clase),
        )
        topic_initial = self.get_topic_initial_from_request(clase)
        if topic_initial:
            form.initial.update(topic_initial)
        return render(request, self.template_name, self.get_context(form, clase))

    def post(self, request, pk):
        clase = self.get_clase()
        action = request.POST.get("plan_action", "send")
        form = DocenteClasePlanificacionForm(
            request.POST,
            request.FILES,
            instance=clase,
            clase=clase,
            unavailable_subtema_ids=self.get_unavailable_subtema_ids_by_tema(clase),
        )
        if action not in {"draft", "send"}:
            messages.error(request, "Accion no valida.")
            return render(request, self.template_name, self.get_context(form, clase))
        if clase.estado_planificacion == "aprobada":
            messages.error(request, "La planificacion aprobada no se puede editar desde el perfil docente.")
            return render(request, self.template_name, self.get_context(form, clase))
        if form.is_valid():
            return_planificacion_tema = self.get_return_planificacion_tema(clase)
            if return_planificacion_tema and form.cleaned_data.get("tema") != return_planificacion_tema.tema:
                form.add_error("tema", "Esta clase debe mantenerse dentro del tema de la planificacion abierta.")
                return render(request, self.template_name, self.get_context(form, clase))
            new_existing_subtemas, new_subtema_names, subtema_error = self.resolve_class_new_subtemas(
                form.cleaned_data.get("tema"),
                clase,
            )
            if subtema_error:
                form.add_error("subtemas_seleccionados", subtema_error)
                return render(request, self.template_name, self.get_context(form, clase))
            if action == "send":
                ready_errors = self.get_ready_errors(form)
                if ready_errors:
                    for error in ready_errors:
                        form.add_error(None, error)
                    return render(request, self.template_name, self.get_context(form, clase))
            with transaction.atomic():
                clase = form.save()
                created_subtemas = self.create_class_subtemas(form.cleaned_data.get("tema"), new_subtema_names)
                extra_subtemas = [*new_existing_subtemas, *created_subtemas]
                if extra_subtemas:
                    selected_subtemas = list(clase.get_subtemas_planificados())
                    selected_subtema_ids = {subtema.pk for subtema in selected_subtemas}
                    for subtema in extra_subtemas:
                        if subtema.pk not in selected_subtema_ids:
                            selected_subtemas.append(subtema)
                            selected_subtema_ids.add(subtema.pk)
                    clase.sync_subtemas_planificados(selected_subtemas)
                self.sync_tags(clase.competencias, Competencia, "competencias")
                self.sync_tags(clase.estrategias, Estrategia, "estrategias")
                recurso_ids, new_resources_by_index = self.sync_tags(clase.recursos, Recurso, "recursos")
                self.sync_resource_files(clase, recurso_ids, new_resources_by_index)
                if action == "send":
                    clase.estado_planificacion = "revision"
                    clase.notas_revision = ""
                    clase.observaciones_revision = {}
                    clase.revisado_por = None
                    clase.fecha_revision = None
                    clase.revision_tema_ok = False
                    clase.revision_detalle_ok = False
                    clase.revision_competencias_ok = False
                    clase.revision_estrategias_ok = False
                    clase.revision_recursos_ok = False
                    clase.save(update_fields=[
                        "estado_planificacion",
                        "notas_revision",
                        "observaciones_revision",
                        "revisado_por",
                        "fecha_revision",
                        "revision_tema_ok",
                        "revision_detalle_ok",
                        "revision_competencias_ok",
                        "revision_estrategias_ok",
                        "revision_recursos_ok",
                    ])
                elif clase.estado_planificacion == "revision":
                    clase.estado_planificacion = "pendiente"
                    clase.save(update_fields=["estado_planificacion"])
            if action == "draft":
                messages.success(request, "Borrador guardado correctamente.")
                class_url = reverse_lazy("academico:docente_clase_planificar", kwargs={"pk": clase.pk})
                return_planificacion_tema = self.get_return_planificacion_tema(clase)
                if return_planificacion_tema:
                    topic_url = reverse_lazy("academico:docente_tema_planificar", kwargs={"pk": return_planificacion_tema.pk})
                    return redirect(f"{topic_url}#clase-{clase.pk}")
                return redirect(class_url)
            messages.success(request, "Planificacion enviada a revision.")
            return_planificacion_tema = self.get_return_planificacion_tema(clase)
            if return_planificacion_tema:
                topic_url = reverse_lazy("academico:docente_tema_planificar", kwargs={"pk": return_planificacion_tema.pk})
                return redirect(f"{topic_url}#clase-{clase.pk}")
            return redirect("academico:docente_horarios")
        return render(request, self.template_name, self.get_context(form, clase))

    def get_topic_initial_from_request(self, clase):
        if clase.tema_id or clase.estado_planificacion == "aprobada":
            return {}
        planificacion_tema = self.get_return_planificacion_tema(clase)
        tema = planificacion_tema.tema if planificacion_tema else None
        if not tema:
            tema_id = self.request.GET.get("tema") or self.request.GET.get("from_tema")
            if not str(tema_id or "").isdigit():
                return {}
            tema = Tema.objects.filter(pk=tema_id, planificacion__materia_curso=clase.materia_curso).first()
        if not tema:
            return {}
        initial = {"tema": tema.pk}
        subtema_ids = self.request.GET.getlist("subtema_ids")
        legacy_subtema_id = self.request.GET.get("subtema")
        if not subtema_ids and legacy_subtema_id:
            subtema_ids = [legacy_subtema_id]
        parsed_ids = []
        for subtema_id in subtema_ids:
            if str(subtema_id or "").isdigit():
                parsed_ids.append(int(subtema_id))
        if parsed_ids:
            subtemas = Subtema.objects.filter(pk__in=parsed_ids, tema=tema)
            subtemas_by_id = {subtema.pk: subtema for subtema in subtemas}
            selected = [subtemas_by_id[subtema_id] for subtema_id in parsed_ids if subtema_id in subtemas_by_id]
            if selected:
                initial["subtema"] = selected[0].pk
                initial["subtemas_seleccionados"] = [subtema.pk for subtema in selected]
        return initial

    def get_return_planificacion_tema(self, clase):
        docente = getattr(self.request.user, "partner", None)
        if not docente or not docente.es_docente:
            return None
        planificacion_tema_id = self.request.POST.get("from_planificacion_tema") or self.request.GET.get("from_planificacion_tema")
        queryset = PlanificacionTema.objects.select_related(
            "tema",
            "profesor_materia_curso__materia_curso",
        ).filter(
            profesor_materia_curso__partner=docente,
            profesor_materia_curso__materia_curso=clase.materia_curso,
        )
        if str(planificacion_tema_id or "").isdigit():
            planificacion_tema = queryset.filter(pk=planificacion_tema_id).first()
            if planificacion_tema:
                return planificacion_tema
        return_tema = self.get_return_tema(clase)
        if return_tema:
            planificacion_tema = queryset.filter(tema=return_tema).first()
            if planificacion_tema:
                return planificacion_tema
        if clase.tema_id:
            return queryset.filter(tema_id=clase.tema_id).first()
        return None

    def get_return_tema(self, clase):
        tema_id = self.request.POST.get("from_tema") or self.request.GET.get("from_tema")
        if not str(tema_id or "").isdigit():
            planificacion_tema_id = self.request.POST.get("from_planificacion_tema") or self.request.GET.get("from_planificacion_tema")
            if str(planificacion_tema_id or "").isdigit():
                planificacion_tema = PlanificacionTema.objects.select_related("tema").filter(
                    pk=planificacion_tema_id,
                    profesor_materia_curso__materia_curso=clase.materia_curso,
                ).first()
                if planificacion_tema:
                    return planificacion_tema.tema
            return None
        return Tema.objects.filter(pk=tema_id, planificacion__materia_curso=clase.materia_curso).first()

    def get_form_selected_subtema_ids(self, form, clase):
        values = []
        if form.is_bound:
            values = form.data.getlist("subtemas_seleccionados")
            if not values:
                values = form.data.getlist("subtema_ids")
            legacy_subtema_id = form.data.get("subtema")
            if not values and legacy_subtema_id:
                values = [legacy_subtema_id]
        else:
            initial = form.initial.get("subtemas_seleccionados") or []
            if initial:
                values = initial
            else:
                values = [subtema.pk for subtema in clase.get_subtemas_planificados()]
        selected = []
        for value in values:
            value = getattr(value, "pk", value)
            if str(value or "").isdigit():
                subtema_id = int(value)
                if subtema_id not in selected:
                    selected.append(subtema_id)
        return selected

    def get_unavailable_subtema_ids_by_tema(self, clase):
        used_by_tema = self.get_used_subtema_ids_by_tema(clase)
        ids = set()
        for subtema_ids in used_by_tema.values():
            ids.update(subtema_ids)
        return ids

    def get_used_subtema_ids_by_tema(self, clase):
        used_by_tema = {}
        clases = (
            Clase.objects.select_related("tema", "subtema")
            .prefetch_related("clase_subtemas__subtema")
            .filter(materia_curso=clase.materia_curso, tema__isnull=False)
            .exclude(pk=clase.pk)
        )
        for other_clase in clases:
            if not other_clase.tema_id:
                continue
            used_by_tema.setdefault(str(other_clase.tema_id), set()).update(
                subtema.pk for subtema in other_clase.get_subtemas_planificados()
            )
        return used_by_tema

    def tema_requires_subtemas(self, tema):
        return bool(tema and tema.subtemas_planificacion.exists())

    def form_has_selected_subtemas(self, form):
        subtemas = list(form.cleaned_data.get("subtemas_seleccionados") or [])
        legacy_subtema = form.cleaned_data.get("subtema")
        if legacy_subtema and legacy_subtema not in subtemas:
            subtemas.append(legacy_subtema)
        return bool(subtemas or self.get_new_subtema_names())

    def get_ready_errors(self, form):
        errors = []
        tema = form.cleaned_data.get("tema")
        if not tema:
            errors.append("Selecciona el tema de la clase.")
        elif self.tema_requires_subtemas(tema) and not self.form_has_selected_subtemas(form):
            errors.append("Selecciona al menos un subtema de la clase.")
        if not self.post_has_tag_items("competencias"):
            errors.append("Selecciona o agrega al menos una competencia.")
        if not self.post_has_tag_items("estrategias"):
            errors.append("Selecciona o agrega al menos una estrategia.")
        if not self.post_has_tag_items("recursos"):
            errors.append("Selecciona o agrega al menos un recurso.")
        return errors

    def post_has_tag_items(self, prefix):
        return bool(self.get_selected_tag_ids(prefix) or self.get_new_tag_names(prefix))

    def parse_written_names(self, raw_value):
        names = []
        seen = set()
        for value in (raw_value or "").replace("\r", "\n").replace(",", "\n").splitlines():
            name = value.strip()
            key = name.lower()
            if name and key not in seen:
                names.append(name)
                seen.add(key)
        return names

    def get_new_subtema_names(self):
        return self.parse_written_names(self.request.POST.get("subtemas_nuevos") or "")

    def resolve_class_new_subtemas(self, tema, clase):
        names = self.get_new_subtema_names()
        if not names:
            return [], [], ""
        if not tema:
            return [], [], "Selecciona el tema para crear subtemas nuevos."
        unavailable_subtema_ids = self.get_unavailable_subtema_ids_by_tema(clase)
        existing_subtemas = []
        existing_ids = set()
        new_names = []
        for name in names:
            existing = Subtema.objects.filter(tema=tema, nombre__iexact=name).first()
            if existing:
                if existing.pk in unavailable_subtema_ids:
                    return [], [], f"El subtema {existing.nombre} ya esta asignado a otra clase del tema."
                if existing.pk not in existing_ids:
                    existing_subtemas.append(existing)
                    existing_ids.add(existing.pk)
                continue
            new_names.append(name)
        return existing_subtemas, new_names, ""

    def create_class_subtemas(self, tema, names):
        if not tema or not names:
            return []
        next_order = (Subtema.objects.filter(tema=tema).aggregate(max_order=Max("orden"))["max_order"] or 0) + 1
        created = []
        for offset, name in enumerate(names):
            created.append(Subtema.objects.create(tema=tema, nombre=name, orden=next_order + offset))
        return created

    def get_selected_tag_ids(self, prefix):
        return {
            int(value)
            for value in self.request.POST.getlist(f"{prefix}_existentes")
            if str(value).isdigit()
        }

    def sync_tags(self, relation, model, prefix):
        selected_ids = self.get_selected_tag_ids(prefix)
        new_items_by_index = {}
        for item in self.get_posted_new_tags(prefix):
            obj, _ = model.objects.get_or_create(nombre=item["nombre"])
            selected_ids.add(obj.pk)
            new_items_by_index[item["index"]] = obj
        relation.set(selected_ids)
        return selected_ids, new_items_by_index

    def sync_resource_files(self, clase, recurso_ids, new_resources_by_index):
        ClaseRecurso.objects.filter(clase=clase).exclude(recurso_id__in=recurso_ids).delete()
        for recurso_id in recurso_ids:
            clase_recurso, _ = ClaseRecurso.objects.get_or_create(clase=clase, recurso_id=recurso_id)
            uploaded = self.request.FILES.get(f"recurso_archivo_{recurso_id}")
            if uploaded:
                clase_recurso.archivo = uploaded
                clase_recurso.save(update_fields=["archivo"])
        for index, recurso in new_resources_by_index.items():
            uploaded = self.request.FILES.get(f"recursos-{index}-archivo")
            if uploaded:
                clase_recurso, _ = ClaseRecurso.objects.get_or_create(clase=clase, recurso=recurso)
                clase_recurso.archivo = uploaded
                clase_recurso.save(update_fields=["archivo"])

    def get_new_tag_names(self, prefix):
        return [item["nombre"] for item in self.get_posted_new_tags(prefix)]

    def get_posted_new_tags(self, prefix):
        try:
            total = int(self.request.POST.get(f"{prefix}-TOTAL_FORMS", 0))
        except (TypeError, ValueError):
            total = 0
        items = []
        for index in range(total):
            name = (self.request.POST.get(f"{prefix}-{index}-nombre") or "").strip()
            delete = self.request.POST.get(f"{prefix}-{index}-DELETE") == "on"
            if name and not delete:
                items.append({"index": index, "nombre": name})
        existing_names = {item["nombre"].lower() for item in items}
        for index, name in enumerate(self.parse_written_names(self.request.POST.get(f"{prefix}_nuevos") or "")):
            if name.lower() in existing_names:
                continue
            items.append({"index": f"typed-{index}", "nombre": name})
            existing_names.add(name.lower())
        return items

    def get_context(self, form, clase):
        horario = clase.horario_aula_curso.horario_dia.horario
        temas = Tema.objects.filter(planificacion__materia_curso=clase.materia_curso).prefetch_related(
            "subtemas_planificacion"
        ).order_by("orden", "nombre")
        used_subtema_ids_by_tema = self.get_used_subtema_ids_by_tema(clase)
        subtemas_by_tema = {
            str(tema.pk): [
                {
                    "id": subtema.pk,
                    "nombre": subtema.nombre,
                    "available": subtema.pk not in used_subtema_ids_by_tema.get(str(tema.pk), set()),
                }
                for subtema in tema.subtemas_planificacion.order_by("orden", "nombre")
            ]
            for tema in temas
        }
        observaciones = clase.observaciones_revision or {}
        clase_recursos_by_id = {
            item.recurso_id: item
            for item in clase.clase_recursos.select_related("recurso")
        }
        can_edit = clase.estado_planificacion != "aprobada"
        if not can_edit:
            for field in form.fields.values():
                field.widget.attrs["disabled"] = "disabled"
        planning_items = self.get_planning_items(clase, form)
        completed_planning_items = sum(1 for item in planning_items if item["done"])
        planning_progress = round((completed_planning_items / len(planning_items)) * 100) if planning_items else 0
        tag_groups = (
            self.build_tag_group(
                "Competencias",
                "competencias",
                Competencia.objects.order_by("nombre"),
                set(clase.competencias.values_list("id", flat=True)),
                "Agregar competencia",
                "Nueva competencia",
                observaciones.get("competencias", ""),
                "ri-medal-line",
                "seleccionadas",
            ),
            self.build_tag_group(
                "Estrategias",
                "estrategias",
                Estrategia.objects.order_by("nombre"),
                set(clase.estrategias.values_list("id", flat=True)),
                "Agregar estrategia",
                "Nueva estrategia",
                observaciones.get("estrategias", ""),
                "ri-route-line",
                "seleccionadas",
            ),
            self.build_tag_group(
                "Recursos",
                "recursos",
                Recurso.objects.order_by("nombre"),
                set(clase.recursos.values_list("id", flat=True)),
                "Agregar recurso",
                "Nuevo recurso",
                observaciones.get("recursos", ""),
                "ri-attachment-2",
                "seleccionados",
                clase_recursos_by_id,
            ),
        )
        return_planificacion_tema = self.get_return_planificacion_tema(clase)
        return_tema = return_planificacion_tema.tema if return_planificacion_tema else self.get_return_tema(clase)
        cancel_url = reverse_lazy("academico:docente_horarios")
        if return_planificacion_tema:
            cancel_url = reverse_lazy("academico:docente_tema_planificar", kwargs={"pk": return_planificacion_tema.pk})
        return {
            "title": "Planificar clase",
            "form": form,
            "clase": clase,
            "horario": horario,
            "temas": temas,
            "subtemas_by_tema_json": json.dumps(subtemas_by_tema),
            "selected_subtema_ids_json": json.dumps(self.get_form_selected_subtema_ids(form, clase)),
            "tag_groups": tag_groups,
            "clase_recursos_by_id": clase_recursos_by_id,
            "observaciones_revision": observaciones,
            "planning_items": planning_items,
            "planning_completed": completed_planning_items,
            "planning_total": len(planning_items),
            "planning_progress": planning_progress,
            "is_late": clase.fecha < timezone.localdate() and clase.estado_planificacion in {"pendiente", "rechazada"},
            "is_observed": clase.estado_planificacion == "rechazada",
            "is_submitted": clase.estado_planificacion == "revision",
            "can_edit": can_edit,
            "return_tema": return_tema,
            "return_planificacion_tema": return_planificacion_tema,
            "cancel_url": cancel_url,
            "attendance_url": reverse_lazy("academico:docente_clase_asistencia", kwargs={"pk": clase.pk}),
        }

    def get_planning_items(self, clase, form):
        if form.is_bound:
            tema_id = (form.data.get("tema") or "").strip()
            tema_has_subtemas = str(tema_id).isdigit() and Tema.objects.filter(
                pk=tema_id,
                subtemas_planificacion__isnull=False,
            ).exists()
            tema_done = bool(tema_id) and bool(
                form.data.getlist("subtemas_seleccionados")
                or form.data.getlist("subtema_ids")
                or form.data.get("subtema")
                or self.get_new_subtema_names()
                or not tema_has_subtemas
            )
            competencias_done = self.post_has_tag_items("competencias")
            estrategias_done = self.post_has_tag_items("estrategias")
            recursos_done = self.post_has_tag_items("recursos")
        else:
            subtemas = clase.get_subtemas_planificados()
            tema_done = bool(clase.tema_id and (subtemas or not clase.tema.subtemas_planificacion.exists()))
            competencias_done = clase.competencias.exists()
            estrategias_done = clase.estrategias.exists()
            recursos_done = clase.recursos.exists()
        observaciones = clase.observaciones_revision or {}
        return [
            {"key": "tema", "label": "Tema", "done": tema_done, "note": observaciones.get("tema", "")},
            {"key": "competencias", "label": "Competencias", "done": competencias_done, "note": observaciones.get("competencias", "")},
            {"key": "estrategias", "label": "Estrategias", "done": estrategias_done, "note": observaciones.get("estrategias", "")},
            {"key": "recursos", "label": "Recursos", "done": recursos_done, "note": observaciones.get("recursos", "")},
        ]

    def build_tag_group(
        self,
        title,
        prefix,
        queryset,
        selected_ids,
        add_label,
        placeholder,
        revision_note,
        icon,
        count_label,
        clase_recursos_by_id=None,
    ):
        if self.request.method == "POST":
            selected_ids = self.get_selected_tag_ids(prefix)
        items = []
        for obj in queryset:
            clase_recurso = (clase_recursos_by_id or {}).get(obj.pk)
            file_meta = file_attachment_meta(clase_recurso.archivo) if clase_recurso else None
            items.append(
                {
                    "obj": obj,
                    "selected": obj.pk in selected_ids,
                    "clase_recurso": clase_recurso,
                    "file_meta": file_meta,
                }
            )
        return {
            "title": title,
            "prefix": prefix,
            "items": items,
            "selected_count": len(selected_ids),
            "add_label": add_label,
            "placeholder": placeholder,
            "revision_note": revision_note,
            "new_entries": self.get_posted_new_tags(prefix) if self.request.method == "POST" else [],
            "new_text": "\n".join(self.get_new_tag_names(prefix)) if self.request.method == "POST" else "",
            "icon": icon,
            "count_label": count_label,
        }


class CoordinacionRevisionPlanificacionesView(CoordinacionRequiredMixin, View):
    permission_required = "academico.review_planificacionclase"
    template_name = "academico/coordinacion_revision_planificaciones.html"
    status_filters = (
        {
            "key": "revision",
            "label": "Enviadas",
            "states": ("revision",),
            "icon": "ri-send-plane-line",
        },
        {
            "key": "pendiente",
            "label": "Sin enviar",
            "states": ("pendiente",),
            "icon": "ri-draft-line",
        },
        {
            "key": "atrasadas",
            "label": "Atrasadas",
            "states": None,
            "late": True,
            "icon": "ri-alarm-warning-line",
        },
        {
            "key": "rechazada",
            "label": "Observadas",
            "states": ("rechazada",),
            "icon": "ri-error-warning-line",
        },
        {
            "key": "aprobada",
            "label": "Aprobadas",
            "states": ("aprobada",),
            "icon": "ri-checkbox-circle-line",
        },
        {
            "key": "todas",
            "label": "Todas",
            "states": None,
            "icon": "ri-list-check-3",
        },
    )

    def get(self, request):
        estado = request.GET.get("estado", "")
        q = (request.GET.get("q") or "").strip()
        selected_docente = self.get_selected_docente()
        clases_queryset = self.get_clases_queryset()

        if selected_docente:
            clases_queryset = clases_queryset.filter(docente_responsable_filter(selected_docente))
        if q:
            clases_queryset = clases_queryset.filter(
                Q(materia_curso__materia__nombre__icontains=q)
                | Q(materia_curso__grupo__nombre__icontains=q)
                | Q(tema__nombre__icontains=q)
                | Q(subtema__nombre__icontains=q)
                | Q(clase_subtemas__subtema__nombre__icontains=q)
                | docente_responsable_search_filter(q)
            )
        clases = list(clases_queryset.distinct())
        stats = self.get_revision_stats(clases)
        selected_estado = self.get_selected_filter(estado, stats)
        revision_cards = [
            self.build_revision_card(clase, selected_docente)
            for clase in self.filter_clases(clases, selected_estado)
        ]
        docente_history = self.build_docente_history(clases, selected_docente)
        return render(
            request,
            self.template_name,
            {
                "title": "Revision de planificaciones",
                "clases": clases,
                "revision_cards": revision_cards,
                "revision_stats": stats,
                "docentes": Partner.objects.filter(es_docente=True, activo=True).order_by("nombre"),
                "selected_docente": selected_docente,
                "selected_docente_id": selected_docente.pk if selected_docente else "",
                "selected_estado_label": self.get_filter_label(selected_estado),
                "status_tabs": self.get_status_tabs(selected_estado, selected_docente, q, stats),
                "estado_choices": Clase.ESTADO_PLANIFICACION_CHOICES,
                "selected_estado": selected_estado,
                "search_query": q,
                "docente_history": docente_history,
                "unassigned_alert": clases_sin_docente_alert(),
            },
        )

    def get_clases_queryset(self):
        return (
            Clase.objects.select_related(
                "materia_curso__materia",
                "materia_curso__grupo",
                "docente",
                "tema",
                "subtema",
                "horario_aula_curso__aula_curso__aula",
                "horario_aula_curso__horario_dia__horario",
            )
            .prefetch_related(
                "competencias",
                "estrategias",
                "recursos",
                "materia_curso__profesor_materia_cursos__partner",
                "clase_subtemas__subtema",
            )
            .order_by("fecha", "horario_aula_curso__horario_dia__horario__hora_inicio", "materia_curso__grupo__nombre")
        )

    def get_selected_docente(self):
        docente_id = self.request.GET.get("docente") or ""
        if not docente_id:
            return None
        return Partner.objects.filter(pk=docente_id, es_docente=True, activo=True).first()

    def empty_revision_stats(self):
        return {
            "total": 0,
            "pendiente": 0,
            "revision": 0,
            "rechazada": 0,
            "aprobada": 0,
            "atrasadas": 0,
        }

    def get_revision_stats(self, clases):
        stats = self.empty_revision_stats()
        stats["total"] = len(clases)
        for clase in clases:
            if clase.estado_planificacion in stats:
                stats[clase.estado_planificacion] += 1
            if self.is_late(clase):
                stats["atrasadas"] += 1
        return stats

    def is_late(self, clase):
        return clase.fecha < timezone.localdate() and clase.estado_planificacion in {"pendiente", "rechazada"}

    def get_selected_filter(self, requested_filter, stats):
        valid_filters = {item["key"] for item in self.status_filters}
        if requested_filter in valid_filters:
            return requested_filter
        if stats["revision"]:
            return "revision"
        if stats["atrasadas"]:
            return "atrasadas"
        if stats["pendiente"]:
            return "pendiente"
        return "todas"

    def get_filter_label(self, selected_filter):
        filter_item = next((item for item in self.status_filters if item["key"] == selected_filter), None)
        return filter_item["label"] if filter_item else ""

    def get_status_tabs(self, selected_filter, selected_docente, q, stats):
        base_url = reverse_lazy("academico:coordinacion_revision_planificaciones")
        tabs = []
        for item in self.status_filters:
            if item.get("late"):
                count = stats["atrasadas"]
            elif item["states"] is None:
                count = stats["total"]
            else:
                count = sum(stats[state] for state in item["states"])
            params = {"estado": item["key"]}
            if selected_docente:
                params["docente"] = selected_docente.pk
            if q:
                params["q"] = q
            tabs.append(
                {
                    **item,
                    "count": count,
                    "url": f"{base_url}?{urlencode(params)}",
                    "is_active": item["key"] == selected_filter,
                }
            )
        return tabs

    def filter_clases(self, clases, selected_filter):
        filter_item = next((item for item in self.status_filters if item["key"] == selected_filter), None)
        if not filter_item or selected_filter == "todas":
            return clases
        if filter_item.get("late"):
            return [clase for clase in clases if self.is_late(clase)]
        allowed_states = set(filter_item["states"] or [])
        return [clase for clase in clases if clase.estado_planificacion in allowed_states]

    def build_revision_card(self, clase, selected_docente=None):
        horario = clase.horario_aula_curso.horario_dia.horario
        docentes = get_clase_docentes(clase)
        metric_docente = selected_docente if selected_docente in docentes else None
        competencias = list(clase.competencias.all())
        estrategias = list(clase.estrategias.all())
        recursos = list(clase.recursos.all())
        subtemas = clase.get_subtemas_planificados()
        topic_progress = topic_temario_progress(clase.materia_curso, clase.tema, metric_docente)
        materia_progress = materia_temario_progress(clase.materia_curso, metric_docente)
        observaciones = clase.observaciones_revision or {}
        steps = [
            {"label": "Tema", "done": bool(clase.tema_id and (subtemas or not clase.tema.subtemas_planificacion.exists())), "note": observaciones.get("tema", "")},
            {"label": "Competencias", "done": bool(competencias), "note": observaciones.get("competencias", "")},
            {"label": "Estrategias", "done": bool(estrategias), "note": observaciones.get("estrategias", "")},
            {"label": "Recursos", "done": bool(recursos), "note": observaciones.get("recursos", "")},
        ]
        completed_steps = sum(1 for step in steps if step["done"])
        action_labels = {
            "pendiente": "Ver pendiente",
            "revision": "Revisar",
            "rechazada": "Ver observacion",
            "aprobada": "Ver aprobada",
        }
        return {
            "clase": clase,
            "horario": horario,
            "estado": clase.estado_planificacion,
            "estado_label": clase.get_estado_planificacion_display(),
            "grupo": clase.materia_curso.grupo,
            "materia": clase.materia_curso.materia,
            "aula": clase.horario_aula_curso.aula_curso.aula,
            "docentes": docentes,
            "tema": clase.tema,
            "subtema": clase.subtema,
            "subtemas": subtemas,
            "topic_progress": topic_progress,
            "materia_progress": materia_progress,
            "steps": steps,
            "progress": round((completed_steps / len(steps)) * 100),
            "is_late": self.is_late(clase),
            "is_unsent": clase.estado_planificacion == "pendiente",
            "revision_note": clase.notas_revision if clase.estado_planificacion == "rechazada" else "",
            "action_label": action_labels.get(clase.estado_planificacion, "Abrir"),
            "url": reverse_lazy("academico:coordinacion_revision_planificacion_detalle", kwargs={"pk": clase.pk}),
        }

    def build_docente_history(self, clases, selected_docente):
        if not selected_docente:
            return None
        materia_cursos = {}
        for clase in clases:
            materia_cursos[clase.materia_curso_id] = clase.materia_curso
        rows = []
        for materia_curso in sorted(
            materia_cursos.values(),
            key=lambda item: (item.grupo.nombre, item.materia.nombre),
        ):
            clases_materia = [clase for clase in clases if clase.materia_curso_id == materia_curso.pk]
            state_counts = {
                "pendiente": 0,
                "revision": 0,
                "rechazada": 0,
                "aprobada": 0,
            }
            topic_ids = set()
            for clase in clases_materia:
                if clase.estado_planificacion in state_counts:
                    state_counts[clase.estado_planificacion] += 1
                if clase.tema_id:
                    topic_ids.add(clase.tema_id)
            progress = materia_temario_progress(materia_curso, selected_docente)
            rows.append(
                {
                    "materia_curso": materia_curso,
                    "progress": progress,
                    "state_counts": state_counts,
                    "class_count": len(clases_materia),
                    "topic_count": len(topic_ids),
                }
            )
        total_subtemas = sum(row["progress"]["total"] for row in rows)
        covered_subtemas = sum(row["progress"]["covered"] for row in rows)
        return {
            "docente": selected_docente,
            "rows": rows,
            "class_count": len(clases),
            "topic_count": len({clase.tema_id for clase in clases if clase.tema_id}),
            "progress": {
                "covered": covered_subtemas,
                "total": total_subtemas,
                "progress": planning_percent(covered_subtemas, total_subtemas),
            },
        }


class CoordinacionRevisionPlanificacionDetalleView(CoordinacionRequiredMixin, View):
    template_name = "academico/coordinacion_revision_planificacion_detalle.html"

    def get_clase(self):
        return get_object_or_404(
            Clase.objects.select_related(
                "materia_curso__materia",
                "materia_curso__grupo",
                "docente",
                "tema",
                "subtema",
                "horario_aula_curso__aula_curso__aula",
                "horario_aula_curso__horario_dia__horario",
            ).prefetch_related(
                "competencias",
                "estrategias",
                "recursos",
                "clase_subtemas__subtema",
                Prefetch("clase_recursos", queryset=ClaseRecurso.objects.select_related("recurso")),
                "materia_curso__profesor_materia_cursos__partner",
            ),
            pk=self.kwargs["pk"],
        )

    def get(self, request, pk):
        return render(request, self.template_name, self.get_context(self.get_clase()))

    def post(self, request, pk):
        clase = self.get_clase()
        action = request.POST.get("review_action")
        notas = request.POST.get("notas_revision", "").strip()
        if action == "rechazar" and not notas:
            messages.error(request, "Escribe observaciones para rechazar la planificacion.")
            return render(request, self.template_name, self.get_context(clase))
        if action not in {"aprobar", "rechazar"}:
            messages.error(request, "Accion no valida.")
            return render(request, self.template_name, self.get_context(clase))
        if action == "aprobar" and not clase_tiene_docente(clase):
            messages.error(request, "Asigna un docente a la materia del grupo antes de aprobar la planificacion.")
            return render(request, self.template_name, self.get_context(clase))
        clase.revision_tema_ok = request.POST.get("revision_tema_ok") == "on"
        clase.revision_detalle_ok = request.POST.get("revision_detalle_ok") == "on"
        clase.revision_competencias_ok = request.POST.get("revision_competencias_ok") == "on"
        clase.revision_estrategias_ok = request.POST.get("revision_estrategias_ok") == "on"
        clase.revision_recursos_ok = request.POST.get("revision_recursos_ok") == "on"
        section_comments = self.get_section_comments(request)
        unchecked_sections = [
            label
            for key, label in self.review_sections()
            if not getattr(clase, f"revision_{key}_ok")
        ]
        if action == "aprobar" and unchecked_sections:
            messages.error(request, "Marca todos los puntos como correctos antes de aprobar.")
            return render(request, self.template_name, self.get_context(clase))
        if action == "rechazar":
            if not unchecked_sections:
                messages.error(request, "Para devolver la planificacion, deja al menos un punto observado.")
                return render(request, self.template_name, self.get_context(clase))
            missing = [
                label
                for key, label in self.review_sections()
                if not getattr(clase, f"revision_{key}_ok") and not section_comments.get(key)
            ]
            if missing:
                messages.error(request, "Agrega comentario en: " + ", ".join(missing))
                return render(request, self.template_name, self.get_context(clase))
        clase.estado_planificacion = "aprobada" if action == "aprobar" else "rechazada"
        clase.notas_revision = notas
        clase.observaciones_revision = {} if action == "aprobar" else section_comments
        clase.revisado_por = getattr(request.user, "partner", None)
        clase.fecha_revision = timezone.now()
        clase.save(update_fields=[
            "estado_planificacion",
            "notas_revision",
            "observaciones_revision",
            "revisado_por",
            "fecha_revision",
            "revision_tema_ok",
            "revision_detalle_ok",
            "revision_competencias_ok",
            "revision_estrategias_ok",
            "revision_recursos_ok",
        ])
        messages.success(request, "Revision registrada correctamente.")
        return redirect("academico:coordinacion_revision_planificaciones")

    def review_sections(self):
        return (
            ("tema", "Tema y subtemas"),
            ("competencias", "Competencias"),
            ("estrategias", "Estrategias"),
            ("recursos", "Recursos"),
        )

    def get_section_comments(self, request):
        comments = {}
        for key, _ in self.review_sections():
            if request.POST.get(f"revision_{key}_ok") == "on":
                continue
            value = request.POST.get(f"observacion_{key}", "").strip()
            if value:
                comments[key] = value
        return comments

    def get_context(self, clase):
        horario = clase.horario_aula_curso.horario_dia.horario
        docentes = get_clase_docentes(clase)
        metric_docente = docentes[0] if len(docentes) == 1 else None
        subtemas = clase.get_subtemas_planificados()
        observaciones = clase.observaciones_revision or {}
        review_items = [
            {
                "key": key,
                "label": label,
                "ok": getattr(clase, f"revision_{key}_ok"),
                "comment": observaciones.get(key, ""),
            }
            for key, label in self.review_sections()
        ]
        planning_items = [
            {"label": "Tema", "done": bool(clase.tema_id and (subtemas or not clase.tema.subtemas_planificacion.exists()))},
            {"label": "Competencias", "done": clase.competencias.exists()},
            {"label": "Estrategias", "done": clase.estrategias.exists()},
            {"label": "Recursos", "done": clase.recursos.exists()},
        ]
        completed_planning_items = sum(1 for item in planning_items if item["done"])
        checked_review_items = sum(1 for item in review_items if item["ok"])
        return {
            "title": "Revisar planificacion",
            "clase": clase,
            "horario": horario,
            "docentes": docentes,
            "subtemas": subtemas,
            "tiene_docente": bool(docentes),
            "resource_items": self.get_resource_items(clase),
            "review_items": review_items,
            "review_total": len(review_items),
            "review_checked": checked_review_items,
            "review_progress": round((checked_review_items / len(review_items)) * 100) if review_items else 0,
            "planning_items": planning_items,
            "planning_progress": round((completed_planning_items / len(planning_items)) * 100) if planning_items else 0,
            "topic_progress": topic_temario_progress(clase.materia_curso, clase.tema, metric_docente),
            "materia_progress": materia_temario_progress(clase.materia_curso, metric_docente),
            "is_late": clase.fecha < timezone.localdate() and clase.estado_planificacion in {"pendiente", "rechazada"},
            "is_unsent": clase.estado_planificacion == "pendiente",
            "observaciones_revision": observaciones,
            "list_url": reverse_lazy("academico:coordinacion_revision_planificaciones"),
            "planificacion_docente_url": reverse_lazy("academico:planificacion_docente"),
        }

    def get_resource_items(self, clase):
        return [
            {
                "clase_recurso": item,
                "recurso": item.recurso,
                "archivo": item.archivo,
                "file_meta": file_attachment_meta(item.archivo),
            }
            for item in clase.clase_recursos.all()
        ]


class AsignaturaListView(InstitutoListView):
    model = Asignatura
    title = "Asignaturas"
    create_url_name = "academico:asignatura_nueva"
    columns = (("Codigo", "codigo"), ("Nombre", "nombre"), ("Activo", "activo"))


class AsignaturaCreateView(InstitutoCreateView):
    model = Asignatura
    form_class = AsignaturaForm
    title = "Nueva asignatura"
    success_url = reverse_lazy("academico:asignatura_list")
    cancel_url = reverse_lazy("academico:asignatura_list")


class AsignaturaUpdateView(InstitutoUpdateView):
    model = Asignatura
    form_class = AsignaturaForm
    title = "Editar asignatura"
    success_url = reverse_lazy("academico:asignatura_list")
    cancel_url = reverse_lazy("academico:asignatura_list")


class TemarioListView(InstitutoListView):
    model = Temario
    title = "Temarios"
    create_url_name = "academico:temario_nuevo"
    columns = (("Asignatura", "asignatura"), ("Nombre", "nombre"), ("Periodo", "periodo_academico"), ("Estado", "estado"))

    def get_queryset(self):
        return super().get_queryset().select_related("asignatura", "periodo_academico", "empresa")


class TemarioCreateView(InstitutoCreateView):
    model = Temario
    form_class = TemarioForm
    title = "Nuevo temario"
    success_url = reverse_lazy("academico:temario_list")
    cancel_url = reverse_lazy("academico:temario_list")


class TemarioUpdateView(InstitutoUpdateView):
    model = Temario
    form_class = TemarioForm
    title = "Editar temario"
    success_url = reverse_lazy("academico:temario_list")
    cancel_url = reverse_lazy("academico:temario_list")


class HorarioClaseListView(InstitutoListView):
    model = HorarioClase
    title = "Horarios de clase"
    create_url_name = "academico:horario_asignacion"
    update_url_name = "academico:horario_editar"
    columns = (
        ("Fecha", "fecha"),
        ("Hora", "rango_hora"),
        ("Aula", "aula"),
        ("Asignatura", "asignatura"),
        ("Docente", "docente"),
        ("Tutor", "tutor"),
        ("Periodo", "periodo_academico"),
        ("Estado", "estado"),
    )

    def get_queryset(self):
        queryset = (
            super()
            .get_queryset()
            .select_related("empresa", "periodo_academico", "aula", "asignatura", "docente", "tutor")
        )
        user = self.request.user
        if not can_view_all_horarios(user):
            queryset = queryset.filter(docente__usuario=user)
        q = self.request.GET.get("q")
        if q:
            queryset = queryset.filter(
                Q(aula__nombre__icontains=q)
                | Q(asignatura__nombre__icontains=q)
                | Q(docente__nombre__icontains=q)
                | Q(tutor__nombre__icontains=q)
                | Q(periodo_academico__nombre__icontains=q)
            )
        return queryset

    def get_column_value(self, obj, attr):
        if attr == "rango_hora":
            return f"{obj.hora_inicio:%H:%M} - {obj.hora_fin:%H:%M}"
        return super().get_column_value(obj, attr)


class HorarioClaseCreateView(InstitutoCreateView):
    model = HorarioClase
    form_class = HorarioClaseForm
    title = "Nuevo horario de clase"
    success_url = reverse_lazy("academico:horario_list")
    cancel_url = reverse_lazy("academico:horario_list")


class HorarioClaseUpdateView(InstitutoUpdateView):
    model = HorarioClase
    form_class = HorarioClaseForm
    title = "Editar horario de clase"
    success_url = reverse_lazy("academico:horario_list")
    cancel_url = reverse_lazy("academico:horario_list")


class HorarioAsignacionView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "academico.add_horarioclase"
    template_name = "academico/horario_asignacion.html"

    def get_empresa(self):
        return Empresa.objects.filter(activa=True).first() or Empresa.objects.first()

    def get(self, request):
        empresa = self.get_empresa()
        initial = {}
        if empresa:
            periodo = empresa.periodoacademico_set.filter(activo=True).order_by("-fecha_inicio", "nombre").first()
            initial["periodo_academico"] = getattr(periodo, "pk", None)
            if periodo and periodo.fecha_inicio:
                initial["fecha_inicio"] = periodo.fecha_inicio.isoformat()
            if periodo and periodo.fecha_fin:
                initial["fecha_fin"] = periodo.fecha_fin.isoformat()
            initial["aula"] = getattr(empresa.aula_set.filter(activo=True).first(), "pk", None)
        selected_date = request.GET.get("date")
        if selected_date:
            try:
                parsed_date = date.fromisoformat(selected_date)
                initial["fecha_inicio"] = parsed_date.isoformat()
                initial["fecha_fin"] = parsed_date.isoformat()
                initial["dias_semana"] = [str(parsed_date.weekday())]
            except ValueError:
                pass
        if request.GET.get("aula"):
            initial["aula"] = request.GET.get("aula")
        if request.GET.get("periodo"):
            initial["periodo_academico"] = request.GET.get("periodo")
        base_form = HorarioAsignacionBaseForm(empresa=empresa, initial=initial)
        formset = HorarioAsignacionFormSet(form_kwargs={"empresa": empresa})
        return render(request, self.template_name, self.get_context(base_form, formset))

    def post(self, request):
        empresa = self.get_empresa()
        base_form = HorarioAsignacionBaseForm(request.POST, empresa=empresa)
        formset = HorarioAsignacionFormSet(request.POST, form_kwargs={"empresa": empresa})
        saved = 0
        skipped = 0
        if base_form.is_valid():
            periodo = base_form.cleaned_data["periodo_academico"]
            aula = base_form.cleaned_data["aula"]
            current_date = base_form.cleaned_data["fecha_inicio"]
            end_date = base_form.cleaned_data["fecha_fin"]
            weekdays = {int(day) for day in base_form.cleaned_data["dias_semana"]}
            while current_date <= end_date:
                if current_date.weekday() not in weekdays:
                    current_date += timedelta(days=1)
                    continue
                horario = HorarioClase(
                    empresa=empresa,
                    periodo_academico=periodo,
                    aula=aula,
                    fecha=current_date,
                    hora_inicio=base_form.cleaned_data["hora_inicio"],
                    hora_fin=base_form.cleaned_data["hora_fin"],
                    asignatura=base_form.cleaned_data["asignatura"],
                    docente=base_form.cleaned_data["docente"],
                    tutor=base_form.cleaned_data.get("tutor"),
                    estado="programada",
                    activo=True,
                    usuario_updated=request.user,
                )
                try:
                    horario.full_clean()
                    horario.save()
                    PlanificacionClase.objects.get_or_create(
                        horario_clase=horario,
                        defaults={
                            "empresa": empresa,
                            "docente": horario.docente,
                            "aula": horario.aula,
                            "asignatura": horario.asignatura,
                            "fecha_planificada": horario.fecha,
                            "objetivo": "",
                            "estado": "pendiente",
                            "activo": True,
                            "usuario_updated": request.user,
                        },
                    )
                    saved += 1
                except ValidationError:
                    skipped += 1
                current_date += timedelta(days=1)
            if saved:
                message = f"{saved} clase(s) programada(s)."
                if skipped:
                    message += f" {skipped} no se crearon por cruces de aula o docente."
                messages.success(request, message)
                return redirect("academico:horario_calendario")
            if skipped:
                base_form.add_error(None, "No se crearon horarios porque todos cruzan con otra clase del aula o docente.")
        return render(request, self.template_name, self.get_context(base_form, formset))

    def get_context(self, base_form, formset):
        return {
            "title": "Asignar horarios",
            "base_form": base_form,
            "formset": formset,
            "calendar_url": reverse_lazy("academico:horario_calendario"),
            "cancel_url": reverse_lazy("academico:horario_list"),
        }


class HorarioCalendarioView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "academico.view_horarioclase"
    template_name = "academico/horario_calendario.html"
    month_names = (
        "",
        "Enero",
        "Febrero",
        "Marzo",
        "Abril",
        "Mayo",
        "Junio",
        "Julio",
        "Agosto",
        "Septiembre",
        "Octubre",
        "Noviembre",
        "Diciembre",
    )

    def get_month_date(self, request, queryset):
        today = timezone.localdate()
        first_horario = queryset.order_by("fecha", "hora_inicio").first()
        default_date = first_horario.fecha if first_horario else today
        try:
            year = int(request.GET.get("year") or default_date.year)
            month = int(request.GET.get("month") or default_date.month)
            return date(year, month, 1)
        except (TypeError, ValueError):
            return date(default_date.year, default_date.month, 1)

    def get_focus_date(self, request, queryset):
        month_date = self.get_month_date(request, queryset)
        try:
            return date.fromisoformat(request.GET.get("date") or month_date.isoformat())
        except ValueError:
            return month_date

    def build_calendar_url(self, target_date, filters, view_mode="month"):
        params = {"view": view_mode, "date": target_date.isoformat(), "year": target_date.year, "month": target_date.month}
        params.update({key: value for key, value in filters.items() if value})
        return f"?{urlencode(params)}"

    def can_view_all(self, user):
        return can_view_all_horarios(user)

    def get(self, request):
        can_add_horario = request.user.has_perm("academico.add_horarioclase")
        can_change_horario = request.user.has_perm("academico.change_horarioclase")
        queryset = (
            HorarioClase.objects.select_related("periodo_academico", "aula", "asignatura", "docente", "tutor")
            .filter(activo=True)
            .order_by("fecha", "hora_inicio", "aula__nombre")
        )
        can_view_all = self.can_view_all(request.user)
        if not can_view_all:
            queryset = queryset.filter(docente__usuario=request.user)
        options_queryset = queryset
        periodo_id = request.GET.get("periodo")
        aula_id = request.GET.get("aula")
        if periodo_id:
            queryset = queryset.filter(periodo_academico_id=periodo_id)
        if aula_id:
            queryset = queryset.filter(aula_id=aula_id)
        sidebar_horarios = list(queryset)
        for index, horario in enumerate(sidebar_horarios):
            horario.color_index = (index % 5) + 1

        view_mode = request.GET.get("view") if request.GET.get("view") in {"day", "week", "month"} else "month"
        focus_date = self.get_focus_date(request, queryset)
        month_date = focus_date.replace(day=1)
        next_month = (month_date.replace(day=28) + timedelta(days=4)).replace(day=1)
        previous_month = (month_date - timedelta(days=1)).replace(day=1)
        month_end = next_month - timedelta(days=1)
        month_horarios = list(queryset.filter(fecha__range=(month_date, month_end)))

        horarios_by_date = {}
        for index, horario in enumerate(month_horarios):
            horario.color_index = (index % 5) + 1
            horarios_by_date.setdefault(horario.fecha, []).append(horario)

        calendar_weeks = []
        month_calendar = calendar_module.Calendar(firstweekday=0)
        for week in month_calendar.monthdatescalendar(month_date.year, month_date.month):
            calendar_weeks.append(
                [
                    {
                        "date": day,
                        "in_month": day.month == month_date.month,
                        "is_today": day == timezone.localdate(),
                        "is_selected": day == focus_date,
                        "url": self.build_calendar_url(day, {"periodo": periodo_id or "", "aula": aula_id or ""}, view_mode),
                        "horarios": horarios_by_date.get(day, []),
                    }
                    for day in week
                ]
            )

        today_month = timezone.localdate().replace(day=1)
        filters = {"periodo": periodo_id or "", "aula": aula_id or ""}
        week_start = focus_date - timedelta(days=focus_date.weekday())
        week_days = []
        for offset in range(7):
            day = week_start + timedelta(days=offset)
            week_horarios = list(queryset.filter(fecha=day))
            for index, horario in enumerate(week_horarios):
                horario.color_index = (index % 5) + 1
            week_days.append(
                {
                    "date": day,
                    "is_today": day == timezone.localdate(),
                    "is_selected": day == focus_date,
                    "url": self.build_calendar_url(day, filters, "week"),
                    "horarios": week_horarios,
                }
            )
        day_horarios = list(queryset.filter(fecha=focus_date))
        for index, horario in enumerate(day_horarios):
            horario.color_index = (index % 5) + 1
        selected_date_label = f"{focus_date.day} {self.month_names[focus_date.month]} {focus_date.year}"
        assign_selected_url = reverse_lazy("academico:horario_asignacion")
        assign_params = {"date": focus_date.isoformat(), "periodo": periodo_id or "", "aula": aula_id or ""}
        assign_selected_url = f"{assign_selected_url}?{urlencode({k: v for k, v in assign_params.items() if v})}"
        if view_mode == "day":
            previous_url = self.build_calendar_url(focus_date - timedelta(days=1), filters, view_mode)
            next_url = self.build_calendar_url(focus_date + timedelta(days=1), filters, view_mode)
        elif view_mode == "week":
            previous_url = self.build_calendar_url(focus_date - timedelta(days=7), filters, view_mode)
            next_url = self.build_calendar_url(focus_date + timedelta(days=7), filters, view_mode)
        else:
            previous_url = self.build_calendar_url(previous_month, filters, view_mode)
            next_url = self.build_calendar_url(next_month, filters, view_mode)
        event_classes = ("important", "success", "info", "chill", "")
        calendar_events = []
        for index, horario in enumerate(queryset):
            calendar_events.append(
                {
                    "id": horario.pk,
                    "title": f"{horario.asignatura} · {horario.aula}",
                    "start": f"{horario.fecha.isoformat()}T{horario.hora_inicio.strftime('%H:%M:%S')}",
                    "end": f"{horario.fecha.isoformat()}T{horario.hora_fin.strftime('%H:%M:%S')}",
                    "allDay": False,
                    "className": event_classes[index % len(event_classes)],
                    "url": str(reverse_lazy("academico:horario_editar", kwargs={"pk": horario.pk})) if can_change_horario else "",
                    "description": f"{horario.hora_inicio:%H:%M} - {horario.hora_fin:%H:%M} · {horario.asignatura} · {horario.aula} · {horario.docente}",
                }
            )

        return render(
            request,
            self.template_name,
            {
                "title": "Calendario de horarios",
                "view_mode": view_mode,
                "month_label": f"{self.month_names[month_date.month]} {month_date.year}",
                "selected_date_label": selected_date_label,
                "selected_date_iso": focus_date.isoformat(),
                "calendar_weeks": calendar_weeks,
                "week_days": week_days,
                "day_horarios": day_horarios,
                "month_horarios": month_horarios[:14],
                "calendar_sidebar_horarios": sidebar_horarios,
                "periodos": options_queryset.values_list("periodo_academico_id", "periodo_academico__nombre")
                .distinct()
                .order_by("periodo_academico__nombre"),
                "aulas": options_queryset.values_list("aula_id", "aula__nombre").distinct().order_by("aula__nombre"),
                "selected_periodo": periodo_id or "",
                "selected_aula": aula_id or "",
                "can_view_all_horarios": can_view_all,
                "can_add_horario": can_add_horario,
                "can_change_horario": can_change_horario,
                "day_view_url": self.build_calendar_url(focus_date, filters, "day"),
                "week_view_url": self.build_calendar_url(focus_date, filters, "week"),
                "month_view_url": self.build_calendar_url(focus_date, filters, "month"),
                "previous_month_url": previous_url,
                "next_month_url": next_url,
                "today_url": self.build_calendar_url(today_month, filters, view_mode),
                "assign_selected_url": assign_selected_url,
                "calendar_events_json": json.dumps(calendar_events),
                "calendar_default_view": {"day": "agendaDay", "week": "agendaWeek", "month": "month"}[view_mode],
                "calendar_default_date": focus_date.isoformat(),
                "current_year": month_date.year,
                "current_month": month_date.month,
            },
        )


class TemaListView(InstitutoListView):
    model = Tema
    title = "Temas"
    create_url_name = "academico:tema_nuevo"
    columns = (("Planificacion", "planificacion"), ("Orden", "orden"), ("Nombre", "nombre"), ("Detalle", "detalle"))

    def get_queryset(self):
        return super().get_queryset().select_related("planificacion", "planificacion__materia_curso")


class TemaCreateView(InstitutoCreateView):
    model = Tema
    form_class = TemaForm
    title = "Nuevo tema"
    success_url = reverse_lazy("academico:tema_list")
    cancel_url = reverse_lazy("academico:tema_list")


class TemaUpdateView(InstitutoUpdateView):
    model = Tema
    form_class = TemaForm
    title = "Editar tema"
    success_url = reverse_lazy("academico:tema_list")
    cancel_url = reverse_lazy("academico:tema_list")


class PlanificacionClaseListView(InstitutoListView):
    model = PlanificacionClase
    template_name = "academico/planificacion_list.html"
    title = "Planificacion de clases"
    update_url_name = "academico:planificacion_editar"
    columns = (
        ("Fecha", "fecha_planificada"),
        ("Docente", "docente"),
        ("Aula", "aula"),
        ("Asignatura", "asignatura"),
        ("Tema", "tema"),
        ("Estado", "estado"),
    )

    def get_queryset(self):
        queryset = super().get_queryset().select_related("docente", "aula", "asignatura", "tema", "empresa", "horario_clase")
        user = self.request.user
        if not user.is_superuser and not user.groups.filter(name="Director").exists():
            queryset = queryset.filter(docente__usuario=user)
        estado = self.request.GET.get("estado")
        if estado:
            queryset = queryset.filter(estado=estado)
        q = self.request.GET.get("q")
        if q:
            queryset = queryset.filter(
                Q(docente__nombre__icontains=q)
                | Q(aula__nombre__icontains=q)
                | Q(asignatura__nombre__icontains=q)
                | Q(tema__nombre__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["estado_choices"] = PlanificacionClase.ESTADO_CHOICES
        context["selected_estado"] = self.request.GET.get("estado", "")
        return context

    def get_action_label(self, obj):
        return "Revisar"


class PlanificacionClaseCreateView(InstitutoCreateView):
    model = PlanificacionClase
    form_class = PlanificacionClaseForm
    title = "Nueva planificacion"
    success_url = reverse_lazy("academico:planificacion_list")
    cancel_url = reverse_lazy("academico:planificacion_list")


class PlanificacionClaseUpdateView(InstitutoUpdateView):
    model = PlanificacionClase
    form_class = PlanificacionClaseForm
    title = "Editar planificacion"
    success_url = reverse_lazy("academico:planificacion_list")
    cancel_url = reverse_lazy("academico:planificacion_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["planificacion_resumen"] = self.object
        user = self.request.user
        can_review = (
            user.is_superuser
            or user.groups.filter(name="Director").exists()
            or user.has_perm("academico.review_planificacionclase")
        )
        context["review_enabled"] = can_review and self.object.estado == "revision"
        context["show_save_button"] = (not can_review) and self.object.estado in {"pendiente", "rechazada"}
        return context

    def get_queryset(self):
        queryset = super().get_queryset().select_related("docente")
        user = self.request.user
        if not user.is_superuser and not user.groups.filter(name="Director").exists():
            queryset = queryset.filter(docente__usuario=user)
        return queryset

    def form_valid(self, form):
        user = self.request.user
        if not user.is_superuser and not user.groups.filter(name="Director").exists():
            if form.instance.estado in {"pendiente", "rechazada"}:
                form.instance.estado = "revision"
        return super().form_valid(form)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        review_action = request.POST.get("review_action")
        user = request.user
        can_review = (
            user.is_superuser
            or user.groups.filter(name="Director").exists()
            or user.has_perm("academico.review_planificacionclase")
        )
        if review_action and can_review:
            self.object.notas_revision = request.POST.get("notas_revision", "")
            self.object.fecha_revision = timezone.now()
            if review_action == "aprobar":
                self.object.estado = "aprobada"
                messages.success(request, "Planificacion aprobada.")
            elif review_action == "rechazar":
                if not self.object.notas_revision.strip():
                    messages.error(request, "Escribe observaciones para devolver la planificacion.")
                    return redirect(request.path)
                self.object.estado = "rechazada"
                messages.success(request, "Planificacion rechazada y devuelta para subsanar.")
            self.object.usuario_updated = request.user
            self.object.save(update_fields=["notas_revision", "fecha_revision", "estado", "usuario_updated", "updated"])
            return redirect(self.success_url)
        return super().post(request, *args, **kwargs)


class BancoPreguntaListView(InstitutoListView):
    model = BancoPregunta
    title = "Bancos de preguntas"
    create_url_name = "academico:banco_nuevo"
    columns = (("Asignatura", "asignatura"), ("Tema", "tema"), ("Tipo", "tipo"), ("Meta", "meta_preguntas"), ("Revisado", "revisado_coordinacion"))

    def get_queryset(self):
        return super().get_queryset().select_related("asignatura", "tema", "subtema", "empresa")


class BancoPreguntaCreateView(InstitutoCreateView):
    model = BancoPregunta
    form_class = BancoPreguntaForm
    title = "Nuevo banco de preguntas"
    success_url = reverse_lazy("academico:banco_list")
    cancel_url = reverse_lazy("academico:banco_list")


class BancoPreguntaUpdateView(InstitutoUpdateView):
    model = BancoPregunta
    form_class = BancoPreguntaForm
    title = "Editar banco de preguntas"
    success_url = reverse_lazy("academico:banco_list")
    cancel_url = reverse_lazy("academico:banco_list")


class PreguntaListView(InstitutoListView):
    model = Pregunta
    title = "Preguntas"
    create_url_name = "academico:pregunta_nueva"
    columns = (("Banco", "banco_pregunta"), ("Enunciado", "enunciado"), ("Dificultad", "dificultad"), ("Estado", "estado"), ("Creado por", "creado_por"))

    def get_queryset(self):
        return super().get_queryset().select_related("banco_pregunta", "creado_por")

    def get_column_value(self, obj, attr):
        value = super().get_column_value(obj, attr)
        if attr == "enunciado" and len(value) > 90:
            return f"{value[:90]}..."
        return value


class PreguntaCreateView(InstitutoCreateView):
    model = Pregunta
    form_class = PreguntaForm
    title = "Nueva pregunta"
    success_url = reverse_lazy("academico:pregunta_list")
    cancel_url = reverse_lazy("academico:pregunta_list")


class PreguntaUpdateView(InstitutoUpdateView):
    model = Pregunta
    form_class = PreguntaForm
    title = "Editar pregunta"
    success_url = reverse_lazy("academico:pregunta_list")
    cancel_url = reverse_lazy("academico:pregunta_list")
